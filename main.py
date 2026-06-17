"""CLI entry point for Product Research Agent.

Usage:
  python main.py                    # scrape all niches
  python main.py --niche capsule    # one niche
  python main.py --no-headless      # show the browser (debug)
  python main.py --diff             # compare latest 2 runs, output new/retired ads
"""
from __future__ import annotations

import argparse
import json
import logging
import logging.handlers
import os
import sys
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from scraper import run
from scorer import rank_ads, top_products
import db
import creatives
import filters
import enrichment
import notifications
import tiktok_scraper


ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "config.json"
OUTPUT_DIR = ROOT / "output"
LOGS_DIR = ROOT / "logs"

log = logging.getLogger("pra.main")


def setup_logging() -> None:
    LOGS_DIR.mkdir(exist_ok=True)
    fmt = logging.Formatter("%(asctime)s %(levelname)-7s %(name)s | %(message)s")
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    root.handlers.clear()

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(fmt)
    root.addHandler(console)

    file = logging.handlers.RotatingFileHandler(
        LOGS_DIR / "agent.log",
        maxBytes=5_000_000,
        backupCount=5,
        encoding="utf-8",
    )
    file.setFormatter(fmt)
    root.addHandler(file)


def load_config() -> dict:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def write_xlsx(rows: list[dict], path: Path, sheet_title: str = "Ads") -> None:
    """Write rows to xlsx with header styling, autofilter, and color-scale on score/days columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    if not rows:
        wb.save(path)
        return

    fields = list(rows[0].keys())
    ws.append(fields)

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(fields) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r.get(f, "") for f in fields])

    col_widths = {
        "ad_text": 60, "sample_ad_text": 60,
        "landing_url": 50, "sample_landing_url": 50,
        "page_name": 32, "page_names": 50, "brand": 28,
        "keyword": 22, "niche": 12,
        "library_id": 20, "scraped_at": 22,
        "start_date": 14, "end_date": 14,
        "is_active": 12, "any_active": 12,
        "days_running": 14, "max_days_running": 16,
        "score": 10, "total_score": 12, "ad_count": 10,
        "variants_from_brand": 18, "platforms": 18,
    }
    for col_idx, name in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(name, 18)

    last_row = len(rows) + 1
    gradient = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    for col_name in ("score", "total_score", "days_running", "max_days_running"):
        if col_name in fields:
            letter = get_column_letter(fields.index(col_name) + 1)
            ws.conditional_formatting.add(f"{letter}2:{letter}{last_row}", gradient)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def load_ads_from_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    out: list[dict] = []
    for r in rows[1:]:
        d = dict(zip(header, r))
        if d.get("library_id"):
            out.append(d)
    return out


def run_diff() -> int:
    files = sorted(OUTPUT_DIR.glob("all_ads_*.xlsx"))
    if len(files) < 2:
        log.warning("Need >=2 prior all_ads runs to diff. Currently have %d.", len(files))
        return 1

    prev_path, curr_path = files[-2], files[-1]
    log.info("Diff: %s -> %s", prev_path.name, curr_path.name)
    prev = load_ads_from_xlsx(prev_path)
    curr = load_ads_from_xlsx(curr_path)

    prev_ids = {a.get("library_id") for a in prev}
    curr_ids = {a.get("library_id") for a in curr}

    new_ads = [a for a in curr if a.get("library_id") not in prev_ids]
    retired_ads = [a for a in prev if a.get("library_id") not in curr_ids]

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_path = OUTPUT_DIR / f"diff_new_{stamp}.xlsx"
    retired_path = OUTPUT_DIR / f"diff_retired_{stamp}.xlsx"
    write_xlsx(new_ads, new_path, sheet_title="New ads")
    write_xlsx(retired_ads, retired_path, sheet_title="Retired ads")

    log.info("Diff result: %d new ads, %d retired ads", len(new_ads), len(retired_ads))
    log.info("  new     -> %s", new_path)
    log.info("  retired -> %s", retired_path)
    return 0


def download_creatives_parallel(all_rows: list[dict], max_workers: int = 8) -> int:
    """Download missing creatives in parallel. Returns count downloaded."""
    import concurrent.futures
    pending = []
    for r in all_rows:
        lib_id = r.get("library_id")
        url = r.get("media_url")
        mtype = r.get("media_type")
        if lib_id and url and mtype and not db.has_creative(lib_id):
            pending.append((lib_id, mtype, url))
    if not pending:
        return 0
    log.info("Downloading %d creatives (parallel)...", len(pending))
    downloaded = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(creatives.download, lib, mt, url): lib for lib, mt, url in pending}
        for fut in concurrent.futures.as_completed(futures):
            if fut.result():
                downloaded += 1
    log.info("Creatives saved: %d / %d attempted", downloaded, len(pending))
    return downloaded


def import_legacy_xlsx() -> int:
    """One-shot: backfill the SQLite DB from existing all_ads_*.xlsx files."""
    db.init_db()
    files = sorted(OUTPUT_DIR.glob("all_ads_*.xlsx"))
    if not files:
        log.info("No all_ads_*.xlsx files to import.")
        return 0
    imported = 0
    for path in files:
        rows = load_ads_from_xlsx(path)
        if not rows:
            continue
        # Skip if we already have a run with this many ads (very rough dedup)
        existing_runs = db.list_runs(limit=200)
        if any(r.get("notes") == str(path.name) for r in existing_runs):
            log.info("Skipping already-imported %s", path.name)
            continue
        run_id = db.start_run(niches=["imported"], source="xlsx", notes=str(path.name))
        ranked = rank_ads(rows)
        ann = filters.annotate(ranked)
        for r in ranked:
            extra = ann.get(r.get("library_id"), {})
            r["geo_signal"] = extra.get("geo_signal", "unknown")
            r["niche_relevance"] = extra.get("niche_relevance", "match")
        db.insert_ads(run_id, ranked)
        db.finish_run(run_id, len(ranked))
        log.info("Imported %s -> run %d (%d ads)", path.name, run_id, len(rows))
        imported += 1
    return imported


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Product Research Agent — scrape PH Meta Ads Library for winning products."
    )
    p.add_argument("--niche", default="all", help="capsule | cream | oil | coffee | all (default: all)")
    p.add_argument("--no-headless", action="store_true", help="Show the browser window (for debugging)")
    p.add_argument("--diff", action="store_true", help="Compare the two most recent all_ads runs and write a new/retired report")
    p.add_argument("--import-xlsx", action="store_true", help="Backfill the SQLite DB from existing all_ads_*.xlsx files")
    p.add_argument("--no-creatives", action="store_true", help="Skip downloading ad creatives (faster runs)")
    p.add_argument("--no-enrich", action="store_true", help="Skip marketplace enrichment (Shopee/Lazada)")
    p.add_argument("--no-llm-classify", action="store_true", help="Skip auto Claude classification after scrape (saves ~$1/run)")
    p.add_argument("--enrich", action="store_true", help="Run fast marketplace enrichment (requests-based; often blocked)")
    p.add_argument("--enrich-browser", action="store_true", help="Run Playwright-based marketplace enrichment (slower, reliable; caps top 80 ads)")
    p.add_argument("--reannotate", action="store_true", help="Re-tag all DB ads with geo_signal / niche_relevance flags")
    p.add_argument("--tiktok", action="store_true", help="Scrape TikTok Creative Center Top Ads (all niches)")
    p.add_argument("--tiktok-niche", default="all", help="When used with --tiktok, restrict to one niche")
    p.add_argument("--notify", action="store_true", help="Send Telegram daily summary based on the latest DB run")
    p.add_argument("--rescore", action="store_true", help="Recompute brand clustering + niche-normalized scores for all DB ads")
    p.add_argument("--classify", action="store_true", help="LLM-classify ads (hook angle, claim type, demographic) via Claude API. Needs ANTHROPIC_API_KEY env var.")
    p.add_argument("--classify-max", type=int, default=200, help="Max ads to classify per --classify run (default 200)")
    return p.parse_args()


def main() -> int:
    setup_logging()
    args = parse_args()

    if args.diff:
        OUTPUT_DIR.mkdir(exist_ok=True)
        return run_diff()

    if args.import_xlsx:
        OUTPUT_DIR.mkdir(exist_ok=True)
        count = import_legacy_xlsx()
        log.info("Imported %d xlsx runs into DB.", count)
        return 0

    if args.enrich:
        db.init_db()
        latest = db.latest_run_id(only_meta=True)
        if not latest:
            log.error("No runs in DB. Run a scrape first.")
            return 1
        enrichment.enrich_run(latest)
        return 0

    if args.enrich_browser:
        db.init_db()
        latest = db.latest_run_id(only_meta=True)
        if not latest:
            log.error("No runs in DB. Run a scrape first.")
            return 1
        enrichment.enrich_run_browser(latest)
        return 0

    if args.tiktok:
        db.init_db()
        if args.tiktok_niche == "all":
            niches = list(load_config()["niches"].keys())
        else:
            niches = [args.tiktok_niche]
        total = tiktok_scraper.run(niches=niches, country="PH", period=7)
        log.info("TikTok scrape complete: %d ads stored", total)
        return 0

    if args.notify:
        db.init_db()
        ok, msg = notifications.send_daily_summary()
        log.info("Notify: %s", msg)
        return 0 if ok else 1

    if args.reannotate:
        db.init_db()
        runs = db.list_runs(limit=200)
        total = 0
        for r in runs:
            ads = db.ads_for_run(r["run_id"])
            ann = filters.annotate(ads)
            db.update_ad_flags_bulk(r["run_id"], ann)
            log.info("Re-annotated run %d (%d ads)", r["run_id"], len(ads))
            total += len(ads)
        log.info("Re-annotated %d total ads across %d runs.", total, len(runs))
        return 0

    if args.classify:
        db.init_db()
        latest = db.latest_run_id(only_meta=True)
        if not latest:
            log.error("No runs in DB. Run a scrape first.")
            return 1
        import llm_classifier
        llm_classifier.classify_run(latest, max_ads=args.classify_max)
        return 0

    if args.rescore:
        db.init_db()
        runs = db.list_runs(limit=200)
        total = 0
        for r in runs:
            if r.get("source") == "tiktok":
                continue
            ads = db.ads_for_run(r["run_id"])
            if not ads:
                continue
            ranked = rank_ads(ads)
            updates = {
                a["library_id"]: {
                    "brand": a.get("brand", ""),
                    "score": float(a.get("score") or 0),
                    "score_normalized": float(a.get("score_normalized") or 0),
                    "variants_from_brand": int(a.get("variants_from_brand") or 0),
                }
                for a in ranked if a.get("library_id")
            }
            db.update_ad_flags_bulk(r["run_id"], updates)
            log.info("Re-scored run %d (%d ads)", r["run_id"], len(ranked))
            total += len(ranked)
        log.info("Re-scored %d total ads across %d runs.", total, len(runs))
        return 0

    config = load_config()
    if args.no_headless:
        config["headless"] = False

    if args.niche == "all":
        niches = list(config["niches"].keys())
    elif args.niche in config["niches"]:
        niches = [args.niche]
    else:
        log.error("Unknown niche: %r. Available: %s", args.niche, list(config["niches"]))
        return 1

    OUTPUT_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_id = db.start_run(niches=niches, source="scrape")
    log.info("DB run_id=%d", run_id)

    log.info("Country: %s  |  active_status: %s", config["country"], config["active_status"])
    log.info("Niches: %s", niches)
    results = run(config, niches)

    all_rows: list[dict] = []
    for niche, ads in results.items():
        rows = [a.to_row() for a in ads]
        ranked = rank_ads(rows)
        path = OUTPUT_DIR / f"{niche}_{stamp}.xlsx"
        write_xlsx(ranked, path, sheet_title=f"{niche} ads")
        log.info("[%s] %d ads -> %s", niche, len(ranked), path.name)
        all_rows.extend(ranked)

    if all_rows:
        combined = rank_ads(all_rows)
        write_xlsx(combined, OUTPUT_DIR / f"all_ads_{stamp}.xlsx", sheet_title="All ads")
        log.info("[combined] %d ads -> all_ads_%s.xlsx", len(combined), stamp)

        winners = top_products(combined, top_n=30)
        write_xlsx(winners, OUTPUT_DIR / f"top_products_{stamp}.xlsx", sheet_title="Top brands")
        log.info("[winners] top %d brands -> top_products_%s.xlsx", len(winners), stamp)

        annotations = filters.annotate(combined)
        for r in combined:
            extra = annotations.get(r.get("library_id"), {})
            r["geo_signal"] = extra.get("geo_signal", "unknown")
            r["niche_relevance"] = extra.get("niche_relevance", "match")

        # Phase 14 — auto-tag category, sub_category, location on every scraped ad
        import categorization, location_detector
        cat_ann = categorization.annotate(combined)
        loc_ann = location_detector.annotate(combined)
        for r in combined:
            lib = r.get("library_id")
            if not lib:
                continue
            cat_extra = cat_ann.get(lib, {})
            loc_extra = loc_ann.get(lib, {})
            r["category"] = cat_extra.get("category")
            r["sub_category"] = cat_extra.get("sub_category")
            r["location"] = loc_extra.get("location")

        db.insert_ads(run_id, combined)
        db.finish_run(run_id, len(combined))
        log.info("DB: persisted %d ads to run %d", len(combined), run_id)

        ph_conf = sum(1 for r in combined if r.get("geo_signal") == "ph-confident")
        off_niche = sum(1 for r in combined if r.get("niche_relevance") == "no_match")
        log.info("Filters: %d PH-confident, %d off-niche flagged", ph_conf, off_niche)

        # Phase 15.6 — auto-LLM-classify if ANTHROPIC_API_KEY is set and not disabled
        if (os.environ.get("ANTHROPIC_API_KEY")
                and not getattr(args, "no_llm_classify", False)):
            try:
                import llm_classifier
                # Cap at 200 ads per run to bound cost (~$1/run at claude-haiku rates)
                _attempted, _ok = llm_classifier.classify_run(run_id, max_ads=200)
                log.info("Auto-classify: %d/%d ads tagged via Claude", _ok, _attempted)
            except Exception as _e_llm:
                log.warning("Auto-classify failed (non-fatal): %s", _e_llm)

        if not args.no_creatives:
            download_creatives_parallel(combined)

        if not args.no_enrich:
            enrichment.enrich_run(run_id)
    else:
        log.warning("No ads parsed in this run.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
