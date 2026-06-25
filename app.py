"""Streamlit dashboard for Product Research Agent.

Launch:  streamlit run app.py    (or double-click launch_app.bat)
"""
from __future__ import annotations

import os
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

import db
import suggestions
import notifications


ROOT = Path(__file__).resolve().parent
CREATIVES_DIR = ROOT / "creatives"
CONFIG_PATH = ROOT / "config.json"
SAVED_SEARCHES_PATH = ROOT / "saved_searches.json"


# ---- Phase 19.1: Cloud deployment helpers ----
# Streamlit Community Cloud uses st.secrets instead of env vars. Bridge the two
# so the same code works locally (env vars) and in the cloud (secrets.toml).
def _hydrate_secrets_to_env() -> None:
    """Copy known st.secrets keys into os.environ so existing env-var checks work."""
    try:
        if hasattr(st, "secrets") and st.secrets is not None:
            for k in ("ANTHROPIC_API_KEY", "APP_PASSWORD", "DEPLOYMENT_LABEL"):
                if k in st.secrets and not os.environ.get(k):
                    os.environ[k] = str(st.secrets[k])
    except Exception:
        # st.secrets raises if no secrets.toml present locally — fine, fall through
        pass


_hydrate_secrets_to_env()


# Detect cloud mode: Streamlit Community Cloud sets STREAMLIT_RUNTIME=cloud
# or we explicitly mark via DEPLOYMENT_LABEL. Used to hide local-only features
# (Playwright scrape buttons, Windows Task Scheduler integration, etc.).
IS_CLOUD = bool(
    os.environ.get("STREAMLIT_RUNTIME") == "cloud"
    or os.environ.get("DEPLOYMENT_LABEL")
    or os.environ.get("ORBIT_CLOUD_MODE")
)


def _require_password_gate() -> bool:
    """If APP_PASSWORD is configured, show a login screen before the app renders.
    Returns True if user is authenticated (or no password configured)."""
    pw = os.environ.get("APP_PASSWORD")
    if not pw:
        return True
    if st.session_state.get("_authed"):
        return True
    # Render login screen
    _l, _c, _r = st.columns([1, 2, 1])
    with _c:
        st.markdown(
            "<style>"
            "@import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,600;9..144,700&display=swap');"
            ".stApp{background:"
            "radial-gradient(720px 340px at 50% -8%,rgba(212,175,55,0.16),transparent 60%),"
            "radial-gradient(640px 320px at 88% 114%,rgba(124,196,160,0.10),transparent 55%),"
            "#0A1813 !important;}"
            "[data-testid='stForm']{background:rgba(255,255,255,0.035) !important;"
            "border:1px solid rgba(212,175,55,0.18) !important;border-radius:16px !important;"
            "padding:24px 26px 18px !important;max-width:440px;margin:0 auto !important;"
            "box-shadow:0 22px 60px rgba(0,0,0,0.38) !important;}"
            "[data-testid='stForm'] input{background:rgba(10,24,19,0.55) !important;"
            "border:1px solid rgba(212,175,55,0.20) !important;border-radius:9px !important;color:#E7E1D3 !important;}"
            "</style><div style='padding-top:4cm'></div>",
            unsafe_allow_html=True,
        )
        _llg1, _llg2, _llg3 = st.columns([1, 2, 1])
        with _llg2:
            _login_logo = ROOT / "static" / "orbit_logo.png"
            if _login_logo.exists():
                st.image(str(_login_logo), width="stretch")
        st.markdown(
            "<div style='text-align:center;color:#9DB3A5;font-size:0.72rem;letter-spacing:0.2em;"
            "text-transform:uppercase;margin-top:-12px;margin-bottom:18px'>Product Research Hunter</div>",
            unsafe_allow_html=True,
        )
        with st.form("auth_form"):
            attempt = st.text_input("Access password", type="password",
                                     placeholder="Enter team password")
            if st.form_submit_button("Sign in", type="primary", width="stretch"):
                if attempt == pw:
                    st.session_state["_authed"] = True
                    st.rerun()
                else:
                    st.error("Wrong password.")
    st.stop()
    return False


def load_config() -> dict:
    import json
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"niches": {}}


def load_saved_searches() -> dict:
    import json
    if not SAVED_SEARCHES_PATH.exists():
        return {}
    try:
        return json.loads(SAVED_SEARCHES_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_saved_searches(searches: dict) -> None:
    import json
    SAVED_SEARCHES_PATH.write_text(json.dumps(searches, indent=2), encoding="utf-8")


ONBOARDING_FLAG_PATH = ROOT / ".onboarding_seen"
INBOX_LAST_SEEN_PATH = ROOT / ".inbox_last_seen"


def onboarding_done() -> bool:
    return ONBOARDING_FLAG_PATH.exists()


def mark_onboarding_done() -> None:
    ONBOARDING_FLAG_PATH.write_text("seen", encoding="utf-8")


def inbox_last_seen() -> str:
    """ISO timestamp of when user last opened the inbox (or empty)."""
    try:
        return INBOX_LAST_SEEN_PATH.read_text(encoding="utf-8").strip()
    except Exception:
        return ""


def mark_inbox_seen() -> None:
    from datetime import datetime as _dt_ib
    INBOX_LAST_SEEN_PATH.write_text(
        _dt_ib.now().isoformat(timespec="seconds"), encoding="utf-8"
    )


# ---- Recent activity log (Phase 12.2) ----
ACTIVITY_LOG_PATH = ROOT / ".activity_log.json"
_ACTIVITY_MAX = 30  # keep last 30 entries


def log_activity(action: str, target: str = "") -> None:
    """Append a user-action event to the local activity log.

    action: short verb, e.g. 'starred', 'unstarred', 'added_to_testing'
    target: object name, e.g. 'GlowLab', 'Slimming capsule'
    """
    import json as _json_al
    from datetime import datetime as _dt_al
    try:
        if ACTIVITY_LOG_PATH.exists():
            entries = _json_al.loads(ACTIVITY_LOG_PATH.read_text(encoding="utf-8"))
            if not isinstance(entries, list):
                entries = []
        else:
            entries = []
        entries.insert(0, {
            "ts": _dt_al.now().isoformat(timespec="seconds"),
            "action": action,
            "target": target,
        })
        entries = entries[:_ACTIVITY_MAX]
        ACTIVITY_LOG_PATH.write_text(_json_al.dumps(entries, indent=2), encoding="utf-8")
    except Exception:
        pass  # never let activity logging break the app


# ---- Per-page first-visit hints (Phase 12.5) ----
PAGE_HINTS_SEEN_PATH = ROOT / ".page_hints_seen.json"
_PAGE_HINTS = {
    "dashboard": (
        "Welcome to Dashboard",
        "The cross-source overview. The **Daily Brief** at the top shows what changed since "
        "your last scrape. Below it: source counts and the top brands aggregated across "
        "FB Ads, Shopee, Lazada, and TikTok.",
    ),
    "fb_ads": (
        "Welcome to FB Ads Library",
        "The main intelligence page. Click any brand row to open the detail panel — star, "
        "set pipeline status, view brand timeline, fetch Google Trends, find similar brands. "
        "Use **Cards** view for a more visual scan. Run the **Claude classifier** to add AI tags.",
    ),
    "shopee": (
        "Welcome to Shopee",
        "Shopee-linked FB ads enriched with marketplace data (price, units sold, rating, "
        "reviews). Click **Run enrichment** to fetch live sales data via Playwright.",
    ),
    "lazada": (
        "Welcome to Lazada",
        "Lazada-linked FB ads with marketplace data. Same enrichment workflow as Shopee.",
    ),
    "tiktok": (
        "Welcome to TikTok",
        "Top-performing PH TikTok ads from Creative Center. Engagement signals (likes, plays, CTR) "
        "are TikTok's proxy for what's converting on their platform.",
    ),
    "supplier": (
        "Welcome to Supplier",
        "Search 1688.com for Chinese suppliers and convert RMB prices to PHP. The Margin "
        "Calculator below helps you sanity-check pricing before placing an order.",
    ),
    "copy_studio": (
        "Welcome to Copy Studio",
        "AI-powered ad copy generator. Picks winning hook phrases from your scraped corpus "
        "and weaves them into 6 fresh Taglish ad variations. Needs ANTHROPIC_API_KEY.",
    ),
    "testing": (
        "Welcome to Testing",
        "Your product testing lineup. Pipeline-first layout shows queued → testing → "
        "passed/launched. Edit any card to log ROI (revenue, ad spend, COGS). The "
        "**ROI Summary panel** at top shows your win rate and aggregate net profit.",
    ),
    "gallery": (
        "Welcome to Creative Gallery",
        "Visual gallery of all captured ad creatives (images + videos), sorted by score. "
        "Useful for inspiration when building your own ads.",
    ),
    "trends": (
        "Welcome to Trends",
        "Active-ad counts per niche, plotted across all your runs. Rising lines = competition "
        "heating up. Falling = market cooling or scraper drift.",
    ),
    "hooks": (
        "Welcome to Hook Patterns",
        "NLP analysis of winning ad copy. Shows 2–4 word phrases that appear in 2+ different "
        "brands' ads — battle-tested copy patterns you can adapt.",
    ),
    "notifications": (
        "Welcome to Notifications",
        "Configure the Telegram bot to receive daily digests: new winners, retirements, "
        "niche heat, top sustained brands. Free, no email setup needed.",
    ),
    "history": (
        "Welcome to Run History",
        "Complete log of all scrapes (including failed ones). Useful for debugging or "
        "comparing scrape sizes over time.",
    ),
    "settings": (
        "Welcome to Settings",
        "Configure API keys, defaults, theme, and manage data. The danger-zone actions "
        "(cleanup empty runs, reset onboarding) live here.",
    ),
}


def page_hint_seen(page_id: str) -> bool:
    import json as _json_ph
    try:
        if not PAGE_HINTS_SEEN_PATH.exists():
            return False
        data = _json_ph.loads(PAGE_HINTS_SEEN_PATH.read_text(encoding="utf-8"))
        return page_id in data.get("seen", [])
    except Exception:
        return False


def mark_page_hint_seen(page_id: str) -> None:
    import json as _json_ph
    try:
        if PAGE_HINTS_SEEN_PATH.exists():
            data = _json_ph.loads(PAGE_HINTS_SEEN_PATH.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                data = {"seen": []}
        else:
            data = {"seen": []}
        seen = set(data.get("seen", []))
        seen.add(page_id)
        data["seen"] = sorted(seen)
        PAGE_HINTS_SEEN_PATH.write_text(_json_ph.dumps(data, indent=2), encoding="utf-8")
    except Exception:
        pass


def render_page_hint(page_id: str) -> None:
    """Render a dismissable hint card on first visit to a page.
    The dismiss control lives inside the card as a small bottom-right link,
    so the card doesn't get visually broken by an external ✕ column."""
    if page_id not in _PAGE_HINTS:
        return
    if page_hint_seen(page_id):
        return
    title, body = _PAGE_HINTS[page_id]
    # Slim one-line tip strip with a compact dismiss (premium, low-clutter)
    _hint_col, _x_col = st.columns([24, 1])
    with _hint_col:
        st.markdown(
            f"<div style='background:var(--pra-subtle-bg);border:1px solid var(--pra-border);"
            f"border-left:2px solid var(--pra-accent);border-radius:7px;padding:9px 15px;"
            f"margin-bottom:12px;font-size:0.82rem;line-height:1.5'>"
            f"<span style='color:var(--pra-accent);font-weight:600'>💡 {title}</span>"
            f"<span style='color:var(--pra-text-dim)'> — {body}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with _x_col:
        if st.button("✕", key=f"hint_dismiss_{page_id}", help="Hide this tip (won't show again)"):
            mark_page_hint_seen(page_id)
            st.rerun()


def get_activity_today() -> list[dict]:
    """Return today's activity entries (most recent first)."""
    import json as _json_al
    from datetime import datetime as _dt_al
    try:
        if not ACTIVITY_LOG_PATH.exists():
            return []
        entries = _json_al.loads(ACTIVITY_LOG_PATH.read_text(encoding="utf-8"))
        today = _dt_al.now().date().isoformat()
        return [e for e in entries if (e.get("ts") or "").startswith(today)]
    except Exception:
        return []


_favicon_path = Path(__file__).resolve().parent / "static" / "favicon.png"
st.set_page_config(
    page_title="Orbit — Product Research Hunter",
    page_icon=str(_favicon_path) if _favicon_path.exists() else "🛍️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Gate everything behind APP_PASSWORD if set (Phase 19.1)
_require_password_gate()


# ---------- Theme palettes ----------

_DARK = {
    "bg":            "#0A1813",
    "panel":         "#10241D",
    "subtle_bg":     "#0C1E18",
    "border":        "#1E3A30",
    "border_hover":  "#2E4D40",
    "text":          "#E7E1D3",
    "text_strong":   "#F6F1E6",
    "text_muted":    "#9DB3A5",
    "text_dim":      "#6E8278",
    "accent":        "#D4AF37",
    "accent_bright": "#E6CC73",
    "accent_dim":    "rgba(212,175,55,0.10)",
    "success":       "#7CC4A0",
    "info":          "#8FC2B4",
    "warning":       "#E6CC73",
    "danger":        "#E0909F",
    "input_bg":      "#0C1E18",
    "scrollbar":     "#1E3A30",
    "dot":           "rgba(212, 175, 55, 0.055)",
}

_LIGHT = {
    "bg":            "#F6F1E7",
    "panel":         "#FFFDF8",
    "subtle_bg":     "#EEE7D7",
    "border":        "#DCD2BF",
    "border_hover":  "#C7B998",
    "text":          "#2A2620",
    "text_strong":   "#141109",
    "text_muted":    "#6E6655",
    "text_dim":      "#9A9078",
    "accent":        "#9A7B16",
    "accent_bright": "#C9A227",
    "accent_dim":    "rgba(154,123,22,0.10)",
    "success":       "#2E7D52",
    "info":          "#3A7D8C",
    "warning":       "#9A6700",
    "danger":        "#B3404E",
    "input_bg":      "#FFFDF8",
    "scrollbar":     "#DCD2BF",
    "dot":           "rgba(42, 38, 32, 0.06)",
}

if "theme_mode" not in st.session_state:
    st.session_state.theme_mode = "dark"

if "current_page" not in st.session_state:
    st.session_state.current_page = "dashboard"

if "show_palette" not in st.session_state:
    st.session_state.show_palette = False

if "show_inbox" not in st.session_state:
    st.session_state.show_inbox = False

if "show_shortcuts" not in st.session_state:
    st.session_state.show_shortcuts = False

P = _LIGHT if st.session_state.theme_mode == "light" else _DARK


# ---------- Nav structure ----------
# Each entry: (page_id, label, icon glyph). Sections group them.
_NAV_SECTIONS = [
    ("Overview", [
        ("dashboard",     "Dashboard",       "⊞"),
        ("shortlist",     "Shortlist",       "★"),
    ]),
    ("Hunt", [
        ("radar",         "Product Radar",   "🛰️"),
        ("fb_ads",        "FB Ads Library",  "◉"),
        ("shopee",        "Shopee",          "▢"),
        ("lazada",        "Lazada",          "▣"),
        ("tiktok",        "TikTok",          "♪"),
        ("bestsellers",   "Bestsellers",     "🏆"),
    ]),
    ("Validate", [
        ("competitors",   "Competitors",     "◎"),
    ]),
    ("Source", [
        ("supplier",      "Supplier",        "⚙"),
    ]),
    ("System", [
        ("notifications", "Notifications",   "◎"),
        ("history",       "Run History",     "◷"),
        ("settings",      "Settings",        "⚙"),
    ]),
]

_PAGE_TITLES = {
    pid: label
    for _section, items in _NAV_SECTIONS
    for pid, label, _icon in items
}

# Material Symbols for the sidebar nav (fixed-width, monochrome — perfect alignment)
_NAV_MAT = {
    "dashboard": ":material/dashboard:",
    "shortlist": ":material/bookmarks:",
    "radar": ":material/radar:",
    "fb_ads": ":material/ads_click:",
    "shopee": ":material/storefront:",
    "lazada": ":material/shopping_bag:",
    "tiktok": ":material/music_note:",
    "bestsellers": ":material/emoji_events:",
    "competitors": ":material/groups:",
    "supplier": ":material/factory:",
    "notifications": ":material/notifications:",
    "history": ":material/history:",
    "settings": ":material/settings:",
}


_CSS = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600;700&display=swap');

:root {{
    --pra-bg: {P['bg']};
    --pra-panel: {P['panel']};
    --pra-subtle-bg: {P['subtle_bg']};
    --pra-border: {P['border']};
    --pra-border-hover: {P['border_hover']};
    --pra-text: {P['text']};
    --pra-text-strong: {P['text_strong']};
    --pra-text-muted: {P['text_muted']};
    --pra-text-dim: {P['text_dim']};
    --pra-accent: {P['accent']};
    --pra-accent-bright: {P['accent_bright']};
    --pra-accent-dim: {P['accent_dim']};
    --pra-success: {P['success']};
    --pra-info: {P['info']};
    --pra-warning: {P['warning']};
    --pra-danger: {P['danger']};
    --pra-input-bg: {P['input_bg']};
    --pra-dot: {P['dot']};
}}

/* App background with subtle dot grid (Linear/Vercel style) */
html, body {{
    background-color: var(--pra-bg) !important;
}}
.stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] {{
    background-color: var(--pra-bg) !important;
    background-image:
        radial-gradient(1100px 620px at 50% -18%, var(--pra-accent-dim), transparent 60%),
        radial-gradient(900px 500px at 100% 100%, rgba(124,196,160,0.05), transparent 55%),
        radial-gradient(circle at center, var(--pra-dot) 1.5px, transparent 1.5px) !important;
    background-size: 100% 100%, 100% 100%, 24px 24px !important;
    background-position: 0 0, 0 0, 0 0 !important;
    background-attachment: fixed, fixed, fixed !important;
    color: var(--pra-text) !important;
}}
.main, .main > div, .block-container {{
    background: transparent !important;
}}

/* ---- Typography baseline ---- */
html, body, [class*="css"], .stApp, .stMarkdown, button, input, select, textarea {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    -webkit-font-smoothing: antialiased;
}}
body, .stApp, .stMarkdown p, .stMarkdown li, .stMarkdown div {{
    color: var(--pra-text);
}}

/* ---- Layout polish ---- */
.main .block-container {{
    padding-top: 1.5rem;
    padding-bottom: 2.5rem;
    max-width: 1600px;
}}
section[data-testid="stSidebar"], section[data-testid="stSidebar"] > div {{
    background: var(--pra-subtle-bg) !important;
    border-right: 1px solid var(--pra-border);
}}

/* ---- Compact brand bar (single line, premium minimal) ---- */
.pra-header {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 14px 22px;
    margin: 0 0 22px 0;
    background: var(--pra-panel);
    border: 1px solid var(--pra-border);
    border-radius: 6px;
    border-left: 2px solid var(--pra-accent);
}}
.pra-header-title {{
    font-size: 0.82rem;
    font-weight: 600;
    color: var(--pra-text);
    letter-spacing: 0.16em;
    text-transform: uppercase;
}}
.pra-header-right {{
    color: var(--pra-text-muted);
    font-size: 0.78rem;
    font-variant-numeric: tabular-nums;
    letter-spacing: 0.02em;
}}
.pra-header-right strong {{ color: var(--pra-accent); font-weight: 600; }}

/* ---- Metric cards ---- */
[data-testid="stMetric"] {{
    background: var(--pra-panel);
    border: 1px solid var(--pra-border);
    border-radius: 4px;
    padding: 16px 20px;
    transition: border-color 0.2s ease;
}}
[data-testid="stMetric"]:hover {{ border-color: var(--pra-border-hover); }}
[data-testid="stMetricLabel"] {{
    color: var(--pra-text-muted);
    font-size: 0.72rem;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.08em;
}}
[data-testid="stMetricValue"] {{
    font-family: 'Fraunces', Georgia, serif !important;
    font-size: 2rem;
    font-weight: 600;
    color: var(--pra-text-strong);
    letter-spacing: 0;
}}
[data-testid="stMetricDelta"] {{
    font-size: 0.75rem;
    color: var(--pra-text-muted) !important;
    font-weight: 400;
}}
[data-testid="stMetricDelta"] svg {{ display: none; }}

/* ---- Sidebar typography ---- */
section[data-testid="stSidebar"] h1 {{
    font-size: 0.85rem;
    font-weight: 600;
    color: var(--pra-accent);
    letter-spacing: 0.12em;
    text-transform: uppercase;
}}
section[data-testid="stSidebar"] h3, section[data-testid="stSidebar"] .stSubheader {{
    font-size: 0.7rem;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: var(--pra-text-muted);
    font-weight: 500;
    margin-top: 1.25rem;
    margin-bottom: 0.5rem;
}}

/* ---- Tabs: refined underline ---- */
.stTabs [data-baseweb="tab-list"] {{
    gap: 0;
    border-bottom: 1px solid var(--pra-border);
}}
.stTabs [data-baseweb="tab"] {{
    padding: 12px 20px;
    color: var(--pra-text-muted);
    font-size: 0.85rem;
    font-weight: 500;
    letter-spacing: 0.02em;
    background: transparent;
    border-radius: 0;
}}
.stTabs [data-baseweb="tab"]:hover {{ color: var(--pra-text); }}
.stTabs [aria-selected="true"] {{
    color: var(--pra-text-strong) !important;
    font-weight: 600;
}}
.stTabs [data-baseweb="tab-highlight"] {{ background: var(--pra-accent) !important; height: 2px; }}

/* ---- Buttons ---- */
.stButton > button {{
    border-radius: 4px;
    font-weight: 500;
    font-size: 0.85rem;
    letter-spacing: 0.01em;
    padding: 8px 16px;
    transition: background-color 0.15s ease, border-color 0.15s ease;
}}
.stButton > button[kind="secondary"], .stButton > button:not([kind="primary"]),
button[data-testid="stBaseButton-secondary"] {{
    background: var(--pra-panel) !important;
    border: 1px solid var(--pra-border) !important;
    color: var(--pra-text) !important;
}}
.stButton > button[kind="secondary"]:hover, .stButton > button:not([kind="primary"]):hover,
button[data-testid="stBaseButton-secondary"]:hover {{
    background: var(--pra-subtle-bg) !important;
    border-color: var(--pra-accent) !important;
    color: var(--pra-text-strong) !important;
}}

/* Force text color INSIDE buttons — Streamlit wraps labels in <p>/<span>/<div> that
   sometimes don't inherit the button's color. This makes them inherit explicitly. */
.stButton button p,
.stButton button span,
.stButton button div,
.stButton button [data-testid="stMarkdownContainer"],
.stButton button [data-testid="stMarkdownContainer"] *,
button[data-testid^="stBaseButton"] p,
button[data-testid^="stBaseButton"] span {{
    color: inherit !important;
    background: transparent !important;
}}

/* Primary buttons: ensure their dark text shows on gold bg */
.stButton > button[kind="primary"] p,
.stButton > button[kind="primary"] span,
button[data-testid="stBaseButton-primary"] p,
button[data-testid="stBaseButton-primary"] span {{
    color: {P['bg']} !important;
}}

/* Download button + form submit button — same treatment */
.stDownloadButton > button, button[data-testid="stBaseButton-secondaryFormSubmit"] {{
    background: var(--pra-panel) !important;
    border: 1px solid var(--pra-border) !important;
    color: var(--pra-text) !important;
}}
button[data-testid="stBaseButton-primaryFormSubmit"] {{
    background: var(--pra-accent) !important;
    border: 1px solid var(--pra-accent) !important;
    color: {P['bg']} !important;
}}
.stDownloadButton button p, .stDownloadButton button span,
button[data-testid^="stBaseButton"] [data-testid="stMarkdownContainer"] {{
    color: inherit !important;
}}
.stButton > button[kind="primary"] {{
    background: var(--pra-accent);
    color: {P['bg']};
    border: 1px solid var(--pra-accent);
    font-weight: 600;
}}
.stButton > button[kind="primary"]:hover {{
    background: var(--pra-accent-bright);
    border-color: var(--pra-accent-bright);
}}

/* ---- Dataframe (Glide Data Grid) ---- */
.stDataFrame {{
    border: 1px solid var(--pra-border);
    border-radius: 4px;
    overflow: hidden;
}}
[data-testid="stDataFrame"], [data-testid="stDataFrame"] > div,
[data-testid="stDataFrame"] [data-testid="stTable"] {{
    background: var(--pra-panel) !important;
}}
/* Override Glide Data Grid internal CSS variables */
[data-testid="stDataFrame"] {{
    --gdg-bg-cell: {P['panel']} !important;
    --gdg-bg-cell-medium: {P['subtle_bg']} !important;
    --gdg-bg-header: {P['subtle_bg']} !important;
    --gdg-bg-header-has-focus: {P['border']} !important;
    --gdg-bg-header-hovered: {P['border']} !important;
    --gdg-bg-bubble: {P['panel']} !important;
    --gdg-bg-bubble-selected: {P['accent_dim']} !important;
    --gdg-bg-search-result: {P['accent_dim']} !important;
    --gdg-text-dark: {P['text']} !important;
    --gdg-text-medium: {P['text_muted']} !important;
    --gdg-text-light: {P['text_dim']} !important;
    --gdg-text-header: {P['text']} !important;
    --gdg-text-bubble: {P['text']} !important;
    --gdg-text-group-header: {P['text_muted']} !important;
    --gdg-accent-color: {P['accent']} !important;
    --gdg-accent-fg: {P['bg']} !important;
    --gdg-accent-light: {P['accent_dim']} !important;
    --gdg-border-color: {P['border']} !important;
    --gdg-drilldown-border: {P['border_hover']} !important;
    --gdg-header-bottom-border-color: {P['border']} !important;
    --gdg-horizontal-border-color: {P['border']} !important;
    --gdg-vertical-border-color: {P['border']} !important;
    --gdg-link-color: {P['accent']} !important;
}}
/* Sometimes the inner canvas/iframe needs the bg too */
[data-testid="stDataFrame"] canvas {{
    background: {P['panel']} !important;
}}

/* ---- Headings (Fraunces serif — editorial luxe) ---- */
h1, h2, h3, h4 {{
    font-family: 'Fraunces', Georgia, 'Times New Roman', serif !important;
    letter-spacing: 0;
    color: var(--pra-text-strong);
}}
h1 {{ font-weight: 600; letter-spacing: 0.005em; }}
h2 {{ font-size: 1.3rem; font-weight: 600; margin-top: 0.5rem; }}
h3 {{ font-size: 1.05rem; font-weight: 600; }}
h4, h5 {{ font-size: 0.92rem; font-weight: 600; letter-spacing: 0.01em; }}

/* ---- Expanders / Containers — aggressive override ---- */
div[data-testid="stExpander"],
div[data-testid="stExpanderDetails"],
div[data-testid="stExpander"] details {{
    border: 1px solid var(--pra-border) !important;
    border-radius: 4px !important;
    background: var(--pra-panel) !important;
    color: var(--pra-text) !important;
}}
div[data-testid="stExpander"] summary,
div[data-testid="stExpander"] details > summary,
[data-testid="stExpander"] [data-testid="stExpanderToggleIcon"],
[data-testid="stExpanderHeader"],
[data-testid="stExpander"] details summary {{
    background: var(--pra-panel) !important;
    color: var(--pra-text) !important;
    font-weight: 500 !important;
}}
div[data-testid="stExpander"] summary:hover {{
    background: var(--pra-subtle-bg) !important;
    color: var(--pra-text-strong) !important;
}}
div[data-testid="stExpander"] summary p,
div[data-testid="stExpander"] summary span,
div[data-testid="stExpander"] summary div {{
    color: var(--pra-text) !important;
    background: transparent !important;
}}
div[data-testid="stExpander"] summary svg {{
    color: var(--pra-text-muted) !important;
    fill: var(--pra-text-muted) !important;
}}
/* Expanded body */
div[data-testid="stExpander"] details[open] {{
    background: var(--pra-panel) !important;
}}
div[data-testid="stExpander"] details[open] > div {{
    background: var(--pra-panel) !important;
}}

/* ---- Captions ---- */
.stCaption, [data-testid="stCaption"], [data-testid="stCaptionContainer"] {{
    color: var(--pra-text-muted) !important;
    font-size: 0.78rem;
    font-weight: 400;
    line-height: 1.55;
}}

/* ---- Inputs ---- */
.stTextInput input, .stSelectbox > div > div, .stMultiSelect > div > div,
.stTextArea textarea, .stNumberInput input {{
    background: var(--pra-input-bg) !important;
    border-color: var(--pra-border) !important;
    border-radius: 4px !important;
    font-size: 0.85rem !important;
    color: var(--pra-text) !important;
}}
.stTextInput input:focus, .stSelectbox > div > div:focus-within,
.stTextArea textarea:focus {{
    border-color: var(--pra-accent) !important;
    box-shadow: 0 0 0 1px var(--pra-accent) !important;
}}

/* Placeholder text — keep it muted, not gold */
.stTextInput input::placeholder,
.stTextArea textarea::placeholder,
.stNumberInput input::placeholder {{
    color: var(--pra-text-muted) !important;
    opacity: 0.7 !important;
}}

/* ---- BaseWeb select / multiselect (Niche dropdown etc.) — aggressive override ---- */
[data-baseweb="select"],
[data-baseweb="select"] > div,
[data-baseweb="select"] [role="combobox"],
[data-baseweb="select"] [data-baseweb="input"],
[data-baseweb="select"] input,
[data-baseweb="select-control"],
[data-baseweb="select-content"] {{
    background: var(--pra-input-bg) !important;
    color: var(--pra-text) !important;
    border-color: var(--pra-border) !important;
}}
[data-baseweb="select"] svg {{
    color: var(--pra-text-muted) !important;
    fill: var(--pra-text-muted) !important;
}}

/* Force all text inside selects to be visible (catches placeholders + selected text) */
.stMultiSelect [data-baseweb="select"] *:not(svg):not(path),
.stSelectbox [data-baseweb="select"] *:not(svg):not(path),
[data-baseweb="select"] div[role="combobox"],
[data-baseweb="select"] div[role="combobox"] > div,
[data-baseweb="select"] [data-baseweb="value-container"],
[data-baseweb="select"] [data-baseweb="value-container"] > div {{
    color: var(--pra-text) !important;
}}

/* Make placeholder slightly muted but readable */
[data-baseweb="select"] [class*="placeholder" i],
[data-baseweb="select"] [aria-hidden="true"]:not(svg):not(path),
[data-baseweb="select"] div[role="combobox"] > div:empty + div {{
    color: var(--pra-text-muted) !important;
    opacity: 0.95 !important;
}}

/* Dropdown popover (rendered in a portal at body root — needs aggressive selectors) */
[data-baseweb="popover"],
[data-baseweb="popover"] > div,
[data-baseweb="popover"] > div > div,
[data-baseweb="menu"],
[data-baseweb="menu"] > div,
[data-baseweb="select-dropdown"],
[data-baseweb="select-options"],
[data-baseweb="popover"] [role="listbox"],
[data-baseweb="menu"] [role="listbox"],
[role="listbox"],
[role="listbox"] > div,
[data-baseweb="popover"] ul,
[data-baseweb="menu"] ul {{
    background: var(--pra-panel) !important;
    border-color: var(--pra-border) !important;
    color: var(--pra-text) !important;
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.15) !important;
}}
[data-baseweb="popover"] li,
[data-baseweb="menu"] li,
[data-baseweb="popover"] [role="option"],
[data-baseweb="menu"] [role="option"],
[role="listbox"] li,
[role="listbox"] [role="option"],
[role="option"] {{
    background: transparent !important;
    color: var(--pra-text) !important;
}}
[data-baseweb="popover"] [role="option"]:hover,
[data-baseweb="menu"] [role="option"]:hover,
[data-baseweb="popover"] [role="option"][aria-selected="true"],
[data-baseweb="menu"] [role="option"][aria-selected="true"],
[role="listbox"] [role="option"]:hover,
[role="option"]:hover {{
    background: var(--pra-subtle-bg) !important;
    color: var(--pra-text) !important;
}}
/* Text spans inside options (placeholder rows like "— pick a niche —") */
[data-baseweb="popover"] [role="option"] *,
[data-baseweb="menu"] [role="option"] *,
[role="listbox"] [role="option"] * {{
    color: var(--pra-text) !important;
    background: transparent !important;
}}

/* Selected tags (chips) in multiselect */
[data-baseweb="tag"] {{
    background: var(--pra-accent-dim) !important;
    color: var(--pra-text) !important;
    border-color: var(--pra-accent) !important;
}}
[data-baseweb="tag"] span {{ color: var(--pra-text) !important; }}

/* ---- Widget labels (Streamlit puts them in stWidgetLabel) ---- */
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] label,
.stTextInput > label, .stSelectbox > label, .stMultiSelect > label,
.stSlider > label, .stNumberInput > label, .stTextArea > label,
.stDateInput > label, .stTimeInput > label, .stRadio > label,
.stForm label, [data-testid="stForm"] label {{
    color: var(--pra-text) !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
}}

/* Slider value chip */
.stSlider [role="slider"] {{ color: var(--pra-text) !important; }}
.stSlider [data-baseweb="slider"] {{ color: var(--pra-text) !important; }}

/* ---- Checkboxes ---- */
.stCheckbox label, .stCheckbox label p, .stCheckbox label * {{
    font-size: 0.85rem !important;
    color: var(--pra-text) !important;
}}
.stCheckbox label p {{ font-weight: 400 !important; }}

/* ---- All text in sidebar should respect theme ---- */
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] {{
    color: var(--pra-text) !important;
}}
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {{
    color: var(--pra-text-muted) !important;
}}

/* ---- Theme toggle button — small icon ---- */
.stSidebar button[data-testid="stBaseButton-secondary"]:has-text("☾"),
.stSidebar button[data-testid="stBaseButton-secondary"]:has-text("☀") {{
    font-size: 1.1rem !important;
    padding: 4px !important;
    line-height: 1 !important;
    min-height: 32px !important;
}}

/* ---- Slider ---- */
.stSlider [data-baseweb="slider"] > div > div {{ background: var(--pra-accent) !important; }}

/* ---- Pills ---- */
[data-baseweb="button-group"] button {{ color: var(--pra-text) !important; }}

/* ---- Hide default chrome ---- */
#MainMenu {{ visibility: hidden; }}
footer {{ visibility: hidden; }}
header[data-testid="stHeader"] {{ background: transparent; }}
.stDeployButton {{ display: none !important; }}

/* ---- Dividers ---- */
hr {{ border-color: var(--pra-border) !important; margin: 1.5rem 0 !important; }}

/* ---- Theme toggle icon button (sidebar bottom) ---- */
section[data-testid="stSidebar"] .stButton button {{
    min-height: 38px;
}}
/* Icon-only footer buttons (theme toggle + shortcuts "?") — clean circular icons */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:has(button[title*="Switch to"]) button,
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:has(button[title*="shortcut" i]) button {{
    aspect-ratio: 1 / 1;
    padding: 0 !important;
    font-size: 1.05rem !important;
    line-height: 1 !important;
    border-radius: 50% !important;
    background: var(--pra-panel) !important;
    border: 1px solid var(--pra-border) !important;
    color: var(--pra-accent) !important;
    transition: all 0.2s ease;
    min-height: 34px !important;
    max-height: 34px !important;
    width: 34px !important;
    min-width: 34px !important;
    /* Override the nav-flat styles that would otherwise leak in here */
    text-align: center !important;
    justify-content: center !important;
    display: flex !important;
    align-items: center !important;
    margin: 0 auto !important;
}}
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:has(button[title*="Switch to"]) button:hover,
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:has(button[title*="shortcut" i]) button:hover {{
    background: var(--pra-accent) !important;
    color: var(--pra-bg) !important;
    border-color: var(--pra-accent) !important;
    transform: scale(1.05);
}}
/* Make sure the inner text/span doesn't add a label box around the icon */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:has(button[title*="Switch to"]) button p,
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"]:has(button[title*="shortcut" i]) button p {{
    margin: 0 !important;
    color: inherit !important;
    font-size: inherit !important;
    line-height: 1 !important;
}}

/* ---- Custom theme-aware table (used where st.dataframe canvas doesn't respect CSS vars) ---- */
.pra-table {{
    width: 100%;
    border-collapse: collapse;
    background: var(--pra-panel);
    border: 1px solid var(--pra-border);
    border-radius: 4px;
    overflow: hidden;
    font-size: 0.85rem;
    font-variant-numeric: tabular-nums;
}}
.pra-table thead th {{
    background: var(--pra-subtle-bg);
    color: var(--pra-text-muted);
    padding: 11px 14px;
    text-align: left;
    border-bottom: 1px solid var(--pra-border);
    font-weight: 600;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    white-space: nowrap;
}}
.pra-table tbody td {{
    padding: 11px 14px;
    border-bottom: 1px solid var(--pra-border);
    color: var(--pra-text);
    vertical-align: middle;
}}
.pra-table tbody tr:last-child td {{ border-bottom: none; }}
.pra-table tbody tr:hover td {{ background: var(--pra-subtle-bg); }}
.pra-table a.pra-landing {{
    color: var(--pra-accent);
    text-decoration: none;
    font-size: 0.82rem;
}}
.pra-table a.pra-landing:hover {{ text-decoration: underline; }}
.pra-table .num {{ text-align: right; color: var(--pra-text-muted); }}
.pra-table .brand-cell {{ font-weight: 500; color: var(--pra-text-strong); }}
.pra-table .sources-cell {{
    color: var(--pra-text-muted);
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.06em;
}}
.pra-bar-track {{
    background: var(--pra-border);
    height: 6px;
    border-radius: 2px;
    overflow: hidden;
    flex: 1;
    min-width: 60px;
}}
.pra-bar-fill {{
    background: var(--pra-accent);
    height: 100%;
    border-radius: 2px;
}}
.pra-score-cell {{
    display: flex;
    align-items: center;
    gap: 10px;
    min-width: 140px;
}}
.pra-score-num {{
    color: var(--pra-text-muted);
    font-size: 0.8rem;
    min-width: 50px;
    text-align: right;
}}

/* ---- Pipeline kanban cards (Testing tab) ---- */
.pra-pipeline-col {{
    background: var(--pra-panel);
    border: 1px solid var(--pra-border);
    border-radius: 6px;
    padding: 12px 12px 6px 12px;
    min-height: 120px;
}}
.pra-pipeline-col-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 10px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--pra-border);
}}
.pra-pipeline-status {{
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
}}
.pra-pipeline-count {{
    color: var(--pra-text-muted);
    font-size: 0.75rem;
    font-variant-numeric: tabular-nums;
}}
.pra-pipeline-card {{
    background: var(--pra-subtle-bg);
    border: 1px solid var(--pra-border);
    border-radius: 4px;
    padding: 10px 12px;
    margin-bottom: 8px;
    transition: border-color 0.15s ease;
}}
.pra-pipeline-card:hover {{
    border-color: var(--pra-accent);
}}
.pra-pipeline-card-name {{
    color: var(--pra-text);
    font-size: 0.85rem;
    font-weight: 500;
    line-height: 1.3;
    margin-bottom: 4px;
}}
.pra-pipeline-card-meta {{
    color: var(--pra-text-muted);
    font-size: 0.72rem;
    line-height: 1.4;
}}
.pra-pipeline-card-brand {{
    color: var(--pra-text-muted);
    font-size: 0.72rem;
    font-style: italic;
}}
.pra-pipeline-empty {{
    color: var(--pra-text-dim);
    font-size: 0.78rem;
    text-align: center;
    padding: 20px 8px;
    font-style: italic;
}}

/* ============================================================
   Sidebar premium nav (Brand Foundry style)
   ============================================================ */

/* Wrap the nav (and only the nav) in a marker container so we can scope styles. */
.pra-nav-zone {{ /* injected via st.markdown before nav expanders */ }}

/* ---- Section dropdowns (expanders) styled as section headers ---- */
/* Only nav-section expanders get this treatment — the Filters / Saved searches
   expanders below the nav stay with the default expander style for clarity. */
section[data-testid="stSidebar"] div[data-testid="stExpander"].pra-nav-section-exp,
section[data-testid="stSidebar"] .pra-nav-zone ~ div[data-testid="stExpander"]:not(.pra-utility-exp) {{
    border: none !important;
    background: transparent !important;
    margin: 0 !important;
    padding: 0 !important;
}}

/* Generic style for ALL sidebar expanders — clean borderless dropdowns */
section[data-testid="stSidebar"] div[data-testid="stExpander"] {{
    border: 1px solid transparent !important;
    background: transparent !important;
    margin: 1px 0 !important;
    padding: 0 !important;
    border-radius: 4px !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] details {{
    background: transparent !important;
    border: none !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary {{
    background: transparent !important;
    color: var(--pra-text-dim) !important;
    font-size: 0.66rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.18em !important;
    text-transform: uppercase !important;
    padding: 8px 10px 8px 12px !important;
    border: none !important;
    cursor: pointer;
    list-style: none;
    transition: color 0.15s ease;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary:hover {{
    background: transparent !important;
    color: var(--pra-text-muted) !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary p,
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary span,
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary div {{
    color: inherit !important;
    background: transparent !important;
    font-weight: inherit !important;
    letter-spacing: inherit !important;
    font-size: inherit !important;
    text-transform: inherit !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary svg {{
    color: var(--pra-text-dim) !important;
    fill: var(--pra-text-dim) !important;
    width: 12px !important;
    height: 12px !important;
    opacity: 0.7;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] details[open] > div {{
    background: transparent !important;
    padding: 2px 0 6px 0 !important;
}}

/* ---- Nav buttons (inside nav-section expanders) — left-aligned column ---- */
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton > button[kind="secondary"] {{
    width: 100% !important;
    background: transparent !important;
    border: 1px solid transparent !important;
    border-left: 2px solid transparent !important;
    border-radius: 4px !important;
    color: var(--pra-text-muted) !important;
    font-weight: 500 !important;
    font-size: 0.83rem !important;
    text-align: left !important;
    display: flex !important;
    justify-content: flex-start !important;
    align-items: center !important;
    padding: 6px 10px 6px 14px !important;
    margin: 1px 0 !important;
    min-height: 32px !important;
    letter-spacing: 0.005em !important;
    transition: all 0.12s ease !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton > button[kind="secondary"]:hover {{
    background: var(--pra-panel) !important;
    color: var(--pra-text-strong) !important;
    border-left-color: var(--pra-border-hover) !important;
    transform: none !important;
}}

/* ---- Flat (top-level) nav button — Dashboard ----
   Inactive state: subtle, mirrors the expander-internal style */
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-marker) + [data-testid="stElementContainer"] .stButton > button[kind="secondary"] {{
    width: 100% !important;
    background: transparent !important;
    border: 1px solid transparent !important;
    border-left: 2px solid transparent !important;
    border-radius: 4px !important;
    color: var(--pra-text-muted) !important;
    font-weight: 500 !important;
    font-size: 0.83rem !important;
    text-align: left !important;
    display: flex !important;
    justify-content: flex-start !important;
    align-items: center !important;
    padding: 6px 10px 6px 14px !important;
    margin: 1px 0 !important;
    min-height: 32px !important;
}}
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-marker) + [data-testid="stElementContainer"] .stButton > button[kind="secondary"]:hover {{
    background: var(--pra-panel) !important;
    color: var(--pra-text-strong) !important;
    border-left-color: var(--pra-border-hover) !important;
}}

/* Active state (Dashboard is current page) — subtle accent highlight */
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-active) + [data-testid="stElementContainer"] .stButton > button[kind="secondary"] {{
    background: var(--pra-accent-dim) !important;
    color: var(--pra-text-strong) !important;
    border: 1px solid var(--pra-border) !important;
    border-left: 2px solid var(--pra-accent) !important;
    font-weight: 600 !important;
}}
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-active) + [data-testid="stElementContainer"] .stButton > button[kind="secondary"]:hover {{
    background: var(--pra-accent-dim) !important;
    color: var(--pra-accent-bright) !important;
    border-left-color: var(--pra-accent-bright) !important;
}}

/* Inner label of flat nav buttons */
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-marker) + [data-testid="stElementContainer"] .stButton button p,
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-marker) + [data-testid="stElementContainer"] .stButton button div {{
    color: inherit !important;
    background: transparent !important;
    text-align: left !important;
    font-weight: inherit !important;
    width: 100% !important;
    margin: 0 !important;
}}

/* Collapse the empty marker container so it adds no vertical space above Dashboard */
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-flat-nav-marker) {{
    margin: 0 !important;
    padding: 0 !important;
    height: 0 !important;
    min-height: 0 !important;
    overflow: hidden !important;
}}

/* Inner text wrapper inside nav buttons: force left-align */
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton button [data-testid="stMarkdownContainer"],
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton button [data-testid="stMarkdownContainer"] p,
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton button p,
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton button span,
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stButton button div {{
    color: inherit !important;
    background: transparent !important;
    text-align: left !important;
    font-weight: inherit !important;
    width: 100% !important;
    margin: 0 !important;
    line-height: 1.3 !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}}

/* ---- Sidebar real-CTA buttons (Scrape only) — gold fill, centered ---- */
/* Tightly scoped to stHorizontalBlock so it never bleeds onto nav buttons. */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton > button[kind="primary"] {{
    background: var(--pra-accent) !important;
    color: {P['bg']} !important;
    border: 1px solid var(--pra-accent) !important;
    border-left-width: 1px !important;
    border-left-color: var(--pra-accent) !important;
    text-align: center !important;
    justify-content: center !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    min-height: 36px !important;
    display: flex !important;
    align-items: center !important;
    padding: 6px 12px !important;
}}
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton > button[kind="primary"]:hover {{
    background: var(--pra-accent-bright) !important;
    color: {P['bg']} !important;
    border-color: var(--pra-accent-bright) !important;
}}
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton > button[kind="secondary"] {{
    text-align: center !important;
    justify-content: center !important;
    background: var(--pra-panel) !important;
    border: 1px solid var(--pra-border) !important;
    color: var(--pra-text) !important;
    font-weight: 500 !important;
    min-height: 36px !important;
    padding: 6px 12px !important;
}}
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton button p,
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] .stButton button div {{
    text-align: center !important;
    color: inherit !important;
}}

/* ---- Filters expander interior: tighter widget spacing + edge breathing room ---- */
section[data-testid="stSidebar"] div[data-testid="stExpander"] details[open] > div {{
    padding: 6px 10px 12px 10px !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stWidgetLabel"],
section[data-testid="stSidebar"] div[data-testid="stExpander"] label {{
    font-size: 0.78rem !important;
    color: var(--pra-text) !important;
    font-weight: 500 !important;
    text-transform: none !important;
    letter-spacing: 0 !important;
    margin-bottom: 4px !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stCheckbox {{
    margin: 4px 0 !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stCheckbox label {{
    font-size: 0.82rem !important;
    gap: 8px !important;
}}
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stMultiSelect,
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stSelectbox,
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stTextInput {{
    margin-bottom: 12px !important;
}}

/* ---- Slider polish (used in Filters expander) — muted ticks, no edge-touching chips ---- */
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stSlider {{
    margin-bottom: 14px !important;
    padding: 0 6px !important;
}}
section[data-testid="stSidebar"] .stSlider [data-baseweb="slider"] {{
    margin-top: 6px !important;
    padding: 6px 0 !important;
}}
/* The min/max tick value labels at the slider edges */
section[data-testid="stSidebar"] .stSlider [data-testid="stTickBar"] {{
    padding: 0 !important;
    margin-top: 6px !important;
}}
section[data-testid="stSidebar"] .stSlider [data-testid="stTickBarMin"],
section[data-testid="stSidebar"] .stSlider [data-testid="stTickBarMax"] {{
    background: transparent !important;
    color: var(--pra-text-dim) !important;
    font-size: 0.68rem !important;
    font-weight: 400 !important;
    padding: 2px 0 !important;
}}
/* Current value chip above the thumb */
section[data-testid="stSidebar"] .stSlider [data-baseweb="tooltip"],
section[data-testid="stSidebar"] .stSlider [role="slider"] + div {{
    background: var(--pra-panel) !important;
    color: var(--pra-text) !important;
    border: 1px solid var(--pra-border) !important;
    border-radius: 3px !important;
    font-size: 0.72rem !important;
    padding: 2px 6px !important;
}}
/* Thumb itself — subtle gold dot */
section[data-testid="stSidebar"] .stSlider [role="slider"] {{
    background: var(--pra-accent) !important;
    border: 2px solid var(--pra-bg) !important;
    box-shadow: 0 0 0 1px var(--pra-accent) !important;
    width: 14px !important;
    height: 14px !important;
}}

/* ---- Text input inside filters: subtle inset look ---- */
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stTextInput input,
section[data-testid="stSidebar"] div[data-testid="stExpander"] .stTextArea textarea {{
    background: var(--pra-bg) !important;
    border: 1px solid var(--pra-border) !important;
    padding: 7px 10px !important;
    font-size: 0.82rem !important;
}}

/* ---- Multiselect / selectbox inside filters: same panel feel ---- */
section[data-testid="stSidebar"] div[data-testid="stExpander"] [data-baseweb="select"] > div {{
    background: var(--pra-bg) !important;
    min-height: 36px !important;
    border-radius: 4px !important;
}}

/* ---- Checkbox tick styling (gold, not garish) ---- */
section[data-testid="stSidebar"] .stCheckbox label > div[role="checkbox"][aria-checked="true"],
section[data-testid="stSidebar"] .stCheckbox label > span > div[aria-checked="true"] {{
    background: var(--pra-accent) !important;
    border-color: var(--pra-accent) !important;
}}

/* Sidebar divider tighter (subtle hairline) */
section[data-testid="stSidebar"] hr {{
    margin: 10px 4px !important;
    border-color: var(--pra-border) !important;
    opacity: 0.6;
}}

/* Tighten the sidebar's overall padding so content has more room */
section[data-testid="stSidebar"] > div {{
    padding-top: 14px !important;
    padding-left: 10px !important;
    padding-right: 10px !important;
}}

/* Premium section label (rendered as markdown above top-level sidebar widgets) */
.pra-side-label {{
    color: var(--pra-text-dim) !important;
    font-size: 0.62rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.2em !important;
    text-transform: uppercase !important;
    margin: 8px 4px 6px 8px !important;
}}

/* Top-level sidebar selectbox (Active Run) — clean, minimal, integrated look */
section[data-testid="stSidebar"] [data-testid="stElementContainer"]:has(.pra-side-label) ~ [data-testid="stElementContainer"]:first-of-type [data-baseweb="select"] > div,
section[data-testid="stSidebar"] > div > div > [data-testid="stElementContainer"] > [data-testid="stSelectbox"] [data-baseweb="select"] > div {{
    background: var(--pra-panel) !important;
    border: 1px solid var(--pra-border) !important;
    border-radius: 4px !important;
    min-height: 38px !important;
    padding: 2px 4px 2px 12px !important;
    font-size: 0.82rem !important;
    transition: border-color 0.15s ease;
    box-shadow: none !important;
}}
section[data-testid="stSidebar"] > div > div > [data-testid="stElementContainer"] > [data-testid="stSelectbox"] [data-baseweb="select"] > div:hover {{
    border-color: var(--pra-accent) !important;
}}
section[data-testid="stSidebar"] > div > div > [data-testid="stElementContainer"] > [data-testid="stSelectbox"] [data-baseweb="select"] [data-baseweb="value-container"] > div {{
    color: var(--pra-text) !important;
    font-variant-numeric: tabular-nums;
    letter-spacing: 0.01em;
    font-weight: 500;
    padding: 0 !important;
}}
/* Chevron arrow — make it subtle */
section[data-testid="stSidebar"] > div > div > [data-testid="stElementContainer"] > [data-testid="stSelectbox"] [data-baseweb="select"] svg {{
    color: var(--pra-text-dim) !important;
    fill: var(--pra-text-dim) !important;
    width: 14px !important;
    height: 14px !important;
}}

/* Brand header inside sidebar — centered logo + centered tagline */
section[data-testid="stSidebar"] [data-testid="stImage"] {{
    text-align: center !important;
    padding: 6px 0 4px 0 !important;
}}
section[data-testid="stSidebar"] [data-testid="stImage"] img {{
    display: block !important;
    margin: 0 auto !important;
}}
/* Sidebar brand fallback (no PNG) — centered */
.pra-sidebar-brand {{
    text-align: center;
    padding: 8px 0 4px 0;
}}
/* Tagline below the logo — centered */
.pra-sidebar-tag {{
    text-align: center;
    color: var(--pra-text-dim);
    font-size: 0.62rem;
    letter-spacing: 0.22em;
    text-transform: uppercase;
    font-weight: 600;
    margin: 4px 0 14px 0;
}}

/* ============================================================
   Animation refinements (Phase 12.4) — subtle, premium minimal
   ============================================================ */

/* Page-transition fade-in: applies to the main content area */
[data-testid="stMain"] .block-container {{
    animation: pra-fade-in 0.22s ease-out;
}}
@keyframes pra-fade-in {{
    from {{ opacity: 0; transform: translateY(3px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
}}

/* Metric cards: subtle scale on hover */
[data-testid="stMetric"] {{
    transition: border-color 0.2s ease, transform 0.18s ease, box-shadow 0.18s ease !important;
}}
[data-testid="stMetric"]:hover {{
    transform: translateY(-1px);
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08) !important;
}}

/* Expander chevron — smooth rotation */
section[data-testid="stSidebar"] div[data-testid="stExpander"] summary svg,
div[data-testid="stExpander"] summary svg {{
    transition: transform 0.2s ease !important;
}}

/* Daily Brief panel — subtle slide-in from top */
.pra-brief {{
    animation: pra-slide-in 0.32s ease-out;
}}
@keyframes pra-slide-in {{
    from {{ opacity: 0; transform: translateY(-4px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
}}

/* Brief items — staggered fade-in */
.pra-brief-item {{
    animation: pra-fade-in 0.4s ease-out;
}}

/* Toast notifications — pulse highlight on appear */
[data-testid="stToast"] {{
    animation: pra-toast-in 0.28s ease-out;
}}
@keyframes pra-toast-in {{
    from {{ opacity: 0; transform: translateX(12px); }}
    to   {{ opacity: 1; transform: translateX(0); }}
}}

/* Pipeline cards (Testing) — subtle hover lift */
.pra-pipeline-card {{
    transition: border-color 0.15s ease, transform 0.15s ease, box-shadow 0.15s ease !important;
}}
.pra-pipeline-card:hover {{
    transform: translateY(-1px);
    box-shadow: 0 2px 6px rgba(0, 0, 0, 0.06);
}}

/* Buttons — smoother focus ring */
.stButton button {{
    transition: background-color 0.15s ease, border-color 0.15s ease,
                color 0.15s ease, transform 0.1s ease !important;
}}
.stButton button:active {{
    transform: scale(0.985);
}}

/* Reduce motion for accessibility */
@media (prefers-reduced-motion: reduce) {{
    *, *::before, *::after {{
        animation-duration: 0.01s !important;
        animation-iteration-count: 1 !important;
        transition-duration: 0.01s !important;
    }}
}}

/* ============================================================
   Print-friendly view (Phase 12.3)
   ============================================================ */
@media print {{
    /* Hide chrome: sidebar, header, action buttons */
    section[data-testid="stSidebar"],
    [data-testid="stToolbar"],
    [data-testid="stStatusWidget"],
    header[data-testid="stHeader"],
    .stDeployButton,
    button[data-testid="stBaseButton-primary"],
    button[data-testid="stBaseButton-secondary"],
    [data-testid="stHorizontalBlock"]:has(button[key*="palette"]),
    [data-testid="stHorizontalBlock"]:has(button[key*="inbox"]),
    .pra-brief-cta,
    .pra-flat-nav-marker {{
        display: none !important;
    }}
    /* Force white background + dark text for paper */
    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] {{
        background: #ffffff !important;
        background-image: none !important;
        color: #000000 !important;
    }}
    /* Expand main column to full width */
    .main .block-container {{
        max-width: 100% !important;
        padding: 0 !important;
    }}
    /* Tables — clean printable look */
    .pra-table, .stDataFrame, [data-testid="stDataFrame"] {{
        background: #ffffff !important;
        color: #000000 !important;
        border: 1px solid #888 !important;
    }}
    .pra-table thead th {{ background: #f0f0f0 !important; color: #000 !important; }}
    .pra-table tbody td {{ color: #000 !important; }}
    /* Avoid page breaks inside rows */
    tr, .pra-pipeline-card, .pra-brief-item {{ break-inside: avoid; }}
    /* Page header for printed output */
    body::before {{
        content: "Orbit — Product Research Hunter · printed view";
        display: block;
        font-size: 10pt;
        color: #666;
        border-bottom: 1px solid #ccc;
        padding-bottom: 6px;
        margin-bottom: 14px;
    }}
}}

/* ============================================================
   Brand card view (FB Ads page) — card + attached "View" button
   ============================================================ */
.pra-brand-card {{
    background: var(--pra-panel);
    border: 1px solid var(--pra-border);
    border-bottom: none;
    border-radius: 6px 6px 0 0;
    padding: 14px 16px 12px 16px;
    margin-bottom: 0;
    min-height: 130px;
    transition: border-color 0.15s ease;
}}
/* The button immediately following a card uses the foot-marker as a CSS hook */
[data-testid="stElementContainer"]:has(.pra-brand-card-foot-marker) + [data-testid="stElementContainer"] .stButton {{
    margin-top: 0 !important;
    margin-bottom: 12px !important;
}}
[data-testid="stElementContainer"]:has(.pra-brand-card-foot-marker) + [data-testid="stElementContainer"] .stButton > button {{
    border-top: 0 !important;
    border-radius: 0 0 6px 6px !important;
    min-height: 32px !important;
    font-size: 0.78rem !important;
    color: var(--pra-text-muted) !important;
    font-weight: 500 !important;
    background: var(--pra-subtle-bg) !important;
    transition: background 0.15s ease, color 0.15s ease !important;
}}
[data-testid="stElementContainer"]:has(.pra-brand-card-foot-marker) + [data-testid="stElementContainer"] .stButton > button:hover {{
    background: var(--pra-accent-dim) !important;
    color: var(--pra-accent) !important;
}}
/* Group hover: card border tints when hovering the button below */
[data-testid="stElementContainer"]:has(.pra-brand-card-foot-marker):has(+ [data-testid="stElementContainer"] button:hover) .pra-brand-card,
[data-testid="stElementContainer"]:has(.pra-brand-card:hover) .pra-brand-card {{
    border-color: var(--pra-accent) !important;
}}

/* ============================================================
   Toast notifications — premium polish (Streamlit's bottom-right popups)
   ============================================================ */
[data-testid="stToast"] {{
    background: var(--pra-panel) !important;
    border: 1px solid var(--pra-border) !important;
    border-left: 3px solid var(--pra-accent) !important;
    color: var(--pra-text) !important;
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.25) !important;
    border-radius: 6px !important;
    padding: 12px 16px !important;
    font-size: 0.86rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.005em !important;
}}
[data-testid="stToast"] * {{
    color: var(--pra-text) !important;
    background: transparent !important;
}}
[data-testid="stToastContainer"] {{
    bottom: 24px !important;
    right: 24px !important;
}}

/* ============================================================
   Daily Brief widget (Dashboard top panel)
   ============================================================ */
.pra-brief {{
    background: linear-gradient(135deg, rgba(212,175,55,0.06) 0%, rgba(212,175,55,0.015) 100%);
    border: 1px solid var(--pra-border);
    border-left: 2px solid var(--pra-accent);
    border-radius: 6px;
    padding: 18px 22px 16px 22px;
    margin: 0 0 18px 0;
}}
.pra-brief-header {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 12px;
    padding-bottom: 10px;
    border-bottom: 1px solid var(--pra-border);
}}
.pra-brief-eyebrow {{
    color: var(--pra-accent);
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.18em;
    text-transform: uppercase;
}}
.pra-brief-date {{
    color: var(--pra-text-muted);
    font-size: 0.78rem;
    font-variant-numeric: tabular-nums;
    letter-spacing: 0.02em;
}}
.pra-brief-items {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px 18px;
}}
.pra-brief-item {{
    display: flex;
    align-items: flex-start;
    gap: 10px;
    color: var(--pra-text);
    font-size: 0.86rem;
    line-height: 1.5;
}}
.pra-brief-item-icon {{
    flex-shrink: 0;
    width: 22px;
    height: 22px;
    border-radius: 4px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.78rem;
    font-weight: 700;
    margin-top: 1px;
}}
.pra-brief-item.positive .pra-brief-item-icon {{
    background: rgba(124,196,160,0.14);
    color: var(--pra-success);
}}
.pra-brief-item.warning .pra-brief-item-icon {{
    background: rgba(230,204,115,0.14);
    color: var(--pra-warning);
}}
.pra-brief-item.info .pra-brief-item-icon {{
    background: rgba(143,194,180,0.14);
    color: var(--pra-info);
}}
.pra-brief-item.accent .pra-brief-item-icon {{
    background: rgba(212,175,55,0.16);
    color: var(--pra-accent);
}}
.pra-brief-item-text strong {{
    color: var(--pra-text-strong);
    font-weight: 600;
    font-variant-numeric: tabular-nums;
}}
.pra-brief-item-text .muted {{
    color: var(--pra-text-muted);
    font-size: 0.78rem;
}}
.pra-brief-empty {{
    color: var(--pra-text-muted);
    font-size: 0.82rem;
    font-style: italic;
    padding: 8px 0;
}}
.pra-brief-cta {{
    margin-top: 12px;
    padding-top: 10px;
    border-top: 1px solid var(--pra-border);
    color: var(--pra-text-muted);
    font-size: 0.78rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
}}
.pra-brief-cta-action {{
    color: var(--pra-accent);
    font-weight: 600;
}}
</style>
"""
st.markdown(_CSS, unsafe_allow_html=True)

# ---- Card depth: subtle shadows so windows/content stand out from the background ----
st.markdown(
    "<style>"
    "[data-testid='stMetric']{box-shadow:0 5px 20px rgba(0,0,0,0.38) !important;}"
    ".pra-header{box-shadow:0 7px 26px rgba(0,0,0,0.42) !important;}"
    "[data-testid='stMain'] div[style*='--pra-panel']{box-shadow:0 6px 22px rgba(0,0,0,0.38) !important;}"
    "[data-testid='stMain'] div[style*='rgba(212,175,55,0.13)']{"
    "box-shadow:0 12px 38px rgba(212,175,55,0.12),0 7px 24px rgba(0,0,0,0.42) !important;}"
    "[data-testid='stDataFrame']{box-shadow:0 8px 28px rgba(0,0,0,0.40) !important;border-radius:8px;}"
    ".pra-table{box-shadow:0 8px 28px rgba(0,0,0,0.36) !important;}"
    "[data-testid='stMain'] [data-testid='stExpander']{box-shadow:0 5px 18px rgba(0,0,0,0.30) !important;}"
    "[data-testid='stMain'] .stButton>button{box-shadow:0 3px 12px rgba(0,0,0,0.28) !important;}"
    "[data-testid='stMain'] .stButton>button:hover{box-shadow:0 5px 18px rgba(0,0,0,0.38) !important;}"
    "</style>",
    unsafe_allow_html=True,
)


# ---------- Helpers ----------

@st.cache_data(ttl=60)
def get_runs(limit: int = 50, only_meta: bool = True) -> list[dict]:
    """Default to only_meta=True for the main dashboard. TikTok runs are surfaced in their own tab."""
    return db.list_runs(limit=limit, only_meta=only_meta)


@st.cache_data(ttl=60)
def get_ads(run_id: int) -> list[dict]:
    return db.ads_for_run(run_id)


def _agent_narrative(active_run_id: int, runs_list: list[dict], curr_rows: list[dict]) -> str:
    """Produce an analyst-style 2-3 sentence brief based on the current data."""
    if not curr_rows:
        return ""
    from urllib.parse import urlparse as _up
    n_total = len(curr_rows)
    n_active = sum(1 for r in curr_rows if r.get("is_active"))
    n_ph = sum(1 for r in curr_rows if r.get("geo_signal") == "ph-confident")

    by_niche_brands: dict[str, set] = defaultdict(set)
    for r in curr_rows:
        if not r.get("is_active"):
            continue
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        n, b = r.get("niche"), (r.get("brand") or r.get("page_name"))
        if n and b:
            by_niche_brands[n].add(b)

    by_brand_score: dict[str, dict] = {}
    for r in curr_rows:
        if not r.get("is_active"):
            continue
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        b = r.get("brand") or r.get("page_name")
        if not b:
            continue
        entry = by_brand_score.setdefault(b, {
            "brand": b, "niche": r.get("niche"), "score": 0.0,
            "max_days": 0, "ad_count": 0, "landing": "",
        })
        entry["score"] += float(r.get("score_normalized") or 0)
        entry["max_days"] = max(entry["max_days"], r.get("days_running") or 0)
        entry["ad_count"] += 1
        if not entry["landing"] and r.get("landing_url"):
            entry["landing"] = r["landing_url"]

    top = sorted(by_brand_score.values(), key=lambda x: -x["score"])[:1]
    parts: list[str] = []
    parts.append(
        f"I reviewed **{n_total:,}** ads in this run "
        f"— {n_ph:,} PH-confident, {n_active:,} still active."
    )
    if by_niche_brands:
        hot = max(by_niche_brands.items(), key=lambda x: len(x[1]))
        parts.append(
            f"The **{hot[0]}** niche is hottest with **{len(hot[1])} active brands** competing."
        )
    if top:
        t = top[0]
        landing = ""
        try:
            d = _up(t["landing"]).netloc.replace("www.", "")
            if d:
                landing = f" landing at *{d}*"
        except Exception:
            pass
        parts.append(
            f"My top pick is **{t['brand']}** — "
            f"{t['max_days']:,}-day longevity, {t['ad_count']} active variants{landing}."
        )
    return " ".join(parts)


def _activity_ticker(active_run_id: int, runs_list: list[dict], curr_rows: list[dict]) -> str:
    """One-line activity feed: starred brand changes + biggest events since previous run."""
    parts: list[str] = []
    starred = set(db.list_starred_brands())
    if starred:
        starred_present = sum(
            1 for r in curr_rows
            if (r.get("brand") or r.get("page_name") or "") in starred and r.get("is_active")
        )
        parts.append(f"<strong>{starred_present}</strong> of {len(starred)} shortlist brands active")

    if len(runs_list) >= 2:
        # Find the run that started IMMEDIATELY BEFORE the active one (by timestamp).
        # Without this, users who pick an older run from the dropdown would diff against
        # a NEWER run (whichever was first in the list) — totally wrong baseline.
        _active_meta = next((r for r in runs_list if r["run_id"] == active_run_id), None)
        _active_ts = (_active_meta or {}).get("started_at", "") if _active_meta else ""
        prev_id = None
        if _active_ts:
            _candidates = [r for r in runs_list
                           if r.get("started_at", "") < _active_ts and r["run_id"] != active_run_id]
            _candidates.sort(key=lambda r: r.get("started_at", ""), reverse=True)
            if _candidates:
                prev_id = _candidates[0]["run_id"]
        if prev_id:
            prev_rows = db.ads_for_run(prev_id)
            if len(prev_rows) >= max(50, len(curr_rows) // 2):
                nw = len(suggestions.new_winners_since(prev_rows, curr_rows, threshold_days=30))
                rt = len(suggestions.retired_ads(prev_rows, curr_rows))
                if nw:
                    parts.append(f"<strong>{nw}</strong> new 30-day winner{'s' if nw != 1 else ''}")
                if rt:
                    parts.append(f"<strong>{rt}</strong> retirement{'s' if rt != 1 else ''}")
                # Shortlist-specific changes
                if starred:
                    prev_brand_ads: dict[str, list] = defaultdict(list)
                    curr_brand_ads: dict[str, list] = defaultdict(list)
                    for r in prev_rows:
                        b = r.get("brand") or r.get("page_name")
                        if b in starred:
                            prev_brand_ads[b].append(r)
                    for r in curr_rows:
                        b = r.get("brand") or r.get("page_name")
                        if b in starred:
                            curr_brand_ads[b].append(r)
                    changed = 0
                    for b in starred:
                        if len(prev_brand_ads.get(b, [])) != len(curr_brand_ads.get(b, [])):
                            changed += 1
                    if changed:
                        parts.append(f"<strong>{changed}</strong> starred brand{'s' if changed != 1 else ''} changed")

    if not parts:
        return ""
    return " · ".join(parts)


def _research_brief(active_run_id: int, runs_list: list[dict], curr_rows: list[dict]) -> dict:
    """Compute the at-a-glance brief: new winners, retirements, hot niche, freshness."""
    from datetime import datetime as _dt
    out = {
        "new_winners_count": 0,
        "retired_count": 0,
        "hot_niche": "—",
        "hot_niche_brands": 0,
        "scrape_freshness": "—",
        "scrape_freshness_full": "",
        "first_run": len(runs_list) < 2,
    }
    if len(runs_list) >= 2:
        # Find the run that started IMMEDIATELY BEFORE the active one (by timestamp).
        # Without this, users who pick an older run from the dropdown would diff against
        # a NEWER run (whichever was first in the list) — totally wrong baseline.
        _active_meta = next((r for r in runs_list if r["run_id"] == active_run_id), None)
        _active_ts = (_active_meta or {}).get("started_at", "") if _active_meta else ""
        prev_id = None
        if _active_ts:
            _candidates = [r for r in runs_list
                           if r.get("started_at", "") < _active_ts and r["run_id"] != active_run_id]
            _candidates.sort(key=lambda r: r.get("started_at", ""), reverse=True)
            if _candidates:
                prev_id = _candidates[0]["run_id"]
        if prev_id:
            prev_rows = db.ads_for_run(prev_id)
            if len(prev_rows) >= max(50, len(curr_rows) // 2):
                out["new_winners_count"] = len(suggestions.new_winners_since(prev_rows, curr_rows, threshold_days=30))
                out["retired_count"] = len(suggestions.retired_ads(prev_rows, curr_rows))

    by_niche_brands: dict[str, set] = defaultdict(set)
    for r in curr_rows:
        if not r.get("is_active"):
            continue
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        n = r.get("niche")
        b = r.get("brand") or r.get("page_name")
        if n and b:
            by_niche_brands[n].add(b)
    if by_niche_brands:
        hot = max(by_niche_brands.items(), key=lambda x: len(x[1]))
        out["hot_niche"] = hot[0]
        out["hot_niche_brands"] = len(hot[1])

    run_meta = next((r for r in runs_list if r["run_id"] == active_run_id), None)
    if run_meta and run_meta.get("started_at"):
        try:
            then = _dt.fromisoformat(run_meta["started_at"])
            out["scrape_freshness_full"] = then.strftime("%b %d, %H:%M")
            delta = _dt.now() - then
            if delta.days >= 1:
                out["scrape_freshness"] = f"{delta.days}d ago"
            elif delta.seconds >= 3600:
                out["scrape_freshness"] = f"{delta.seconds // 3600}h ago"
            else:
                out["scrape_freshness"] = f"{max(1, delta.seconds // 60)}m ago"
        except Exception:
            pass
    return out


@st.cache_data(ttl=60)
def get_stats() -> dict:
    return db.stats()


def _creative_path_url(rel_path: str) -> str | None:
    if not rel_path:
        return None
    p = ROOT / rel_path
    return str(p) if p.exists() else None


@st.cache_data(ttl=300)
def _brands_trendlines(brands: tuple[str, ...], n_recent_runs: int = 12) -> dict[str, list[int]]:
    """Compute compact active-ad-count time series per brand across recent runs.

    Returns: {brand_lower: [count_run1, count_run2, ...]} — list length up to n_recent_runs.
    Used to render a sparkline column in the Top Brands table.
    Cached so we don't hammer the DB on every rerun.
    """
    if not brands:
        return {}
    out: dict[str, list[int]] = {b.lower(): [] for b in brands}
    try:
        with db.connect() as conn:
            # Last N META runs ordered chronologically
            recent_runs = conn.execute(
                """SELECT r.run_id, r.started_at FROM runs r
                   JOIN ads a ON a.run_id = r.run_id
                   GROUP BY r.run_id
                   ORDER BY r.started_at DESC
                   LIMIT ?""",
                (n_recent_runs,),
            ).fetchall()
            recent_runs = list(reversed(recent_runs))  # oldest -> newest

            # For each brand, count ads per run
            placeholders = ",".join("?" for _ in brands)
            counts_by_brand_run = conn.execute(
                f"""SELECT LOWER(COALESCE(brand, page_name)) AS bkey,
                          run_id,
                          SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) AS active_ads
                   FROM ads
                   WHERE LOWER(COALESCE(brand, page_name)) IN ({placeholders})
                   GROUP BY bkey, run_id""",
                tuple(b.lower() for b in brands),
            ).fetchall()
            # Build lookup
            lookup: dict[tuple[str, int], int] = {}
            for r in counts_by_brand_run:
                lookup[(r["bkey"], r["run_id"])] = int(r["active_ads"] or 0)
            # Fill timeseries
            for b in brands:
                bk = b.lower()
                out[bk] = [lookup.get((bk, rr["run_id"]), 0) for rr in recent_runs]
    except Exception as _e_tl:
        pass
    return out


@st.cache_data(ttl=300)
def _brand_price_history(brand: str) -> list[dict]:
    """Pull marketplace price points for a brand across runs (Phase 20.4).

    Uses the mp_price already captured per run during enrichment. Returns a
    chronological list of {date, price} — empty if the brand was never enriched.
    """
    out: list[dict] = []
    if not brand:
        return out
    try:
        with db.connect() as conn:
            rows = conn.execute(
                """SELECT r.started_at, a.mp_price
                   FROM ads a JOIN runs r ON r.run_id = a.run_id
                   WHERE LOWER(COALESCE(a.brand, a.page_name)) = LOWER(?)
                     AND a.mp_price IS NOT NULL AND a.mp_price > 0
                   ORDER BY r.started_at ASC""",
                (brand,),
            ).fetchall()
        # One price point per run (take max price seen that run = product, not shipping)
        by_date: dict[str, float] = {}
        for r in rows:
            d = (r["started_at"] or "")[:10]
            p = float(r["mp_price"] or 0)
            if p > 0:
                by_date[d] = max(by_date.get(d, 0), p)
        out = [{"date": d, "price": p} for d, p in sorted(by_date.items())]
    except Exception:
        pass
    return out


@st.cache_data(ttl=300)
def _brand_history(brand: str) -> dict:
    """Aggregate this brand's ads across ALL runs in the DB.

    Returns:
        - per_run: list of {run_id, started_at, active_ads, total_ads, max_days}
        - first_seen: ISO date of earliest run with this brand
        - last_seen: ISO date of most recent run with this brand
        - milestones: list of {date, days_crossed, library_id} for 30/60/90/180-day crossings
        - longest_lived: dict with library_id, days_running, ad_text
    """
    from datetime import datetime as _dth
    out = {
        "per_run": [],
        "first_seen": None,
        "last_seen": None,
        "milestones": [],
        "longest_lived": None,
        "total_runs_seen": 0,
    }
    if not brand:
        return out

    try:
        with db.connect() as conn:
            rows = conn.execute(
                """
                SELECT a.run_id, r.started_at, a.library_id, a.days_running,
                       a.is_active, a.ad_text, a.start_date
                FROM ads a
                JOIN runs r ON r.run_id = a.run_id
                WHERE LOWER(COALESCE(a.brand, a.page_name)) = LOWER(?)
                ORDER BY r.started_at ASC
                """,
                (brand,)
            ).fetchall()
            ads = [dict(r) for r in rows]
    except Exception as e:
        try:
            import logging as _lg
            _lg.getLogger("pra.app").debug("brand_history fail: %s", e)
        except Exception:
            pass
        return out

    if not ads:
        return out

    # Aggregate per run
    by_run: dict[int, dict] = {}
    for a in ads:
        rid = a["run_id"]
        entry = by_run.setdefault(rid, {
            "run_id": rid, "started_at": a["started_at"],
            "active_ads": 0, "total_ads": 0, "max_days": 0,
        })
        entry["total_ads"] += 1
        if a.get("is_active"):
            entry["active_ads"] += 1
        entry["max_days"] = max(entry["max_days"], a.get("days_running") or 0)

    out["per_run"] = sorted(by_run.values(), key=lambda x: x["started_at"])
    out["total_runs_seen"] = len(out["per_run"])
    if out["per_run"]:
        out["first_seen"] = out["per_run"][0]["started_at"][:10]
        out["last_seen"] = out["per_run"][-1]["started_at"][:10]

    # Milestone crossings — first time max_days crossed 30/60/90/180
    _thresholds = [30, 60, 90, 180]
    _hit: set[int] = set()
    for r in out["per_run"]:
        for t in _thresholds:
            if t not in _hit and r["max_days"] >= t:
                _hit.add(t)
                out["milestones"].append({
                    "date": r["started_at"][:10],
                    "days_crossed": t,
                    "run_id": r["run_id"],
                })

    # Longest-lived ad
    longest = max(ads, key=lambda a: a.get("days_running") or 0)
    out["longest_lived"] = {
        "library_id": longest.get("library_id"),
        "days_running": longest.get("days_running") or 0,
        "ad_text": (longest.get("ad_text") or "")[:240],
        "start_date": longest.get("start_date"),
    }

    return out


@st.cache_data(ttl=300)
def _brand_research(brand: str) -> dict:
    """Full cross-run research aggregate for one brand — the backbone of the
    Winner Score and the research hero. Pulls EVERY scrape this brand appears in
    (not just the loaded run), so research works from anywhere in the app.
    """
    from collections import Counter as _Counter
    out = {
        "brand": brand, "found": False,
        "niche": "—", "category": "", "sub_category": "", "location": "",
        "distinct_ads": 0, "active_now": 0, "max_days": 0,
        "runs_seen": 0, "first_seen": None, "last_seen": None,
        "mp_sold": 0, "mp_price": None, "mp_rating": None, "mp_reviews": 0, "mp_source": "",
        "sources": ["FB"], "source_count": 1,
        "per_run_active": [], "momentum": "🆕 New",
        "score_sum": 0.0, "sample_landing": "", "ads_full": [],
    }
    if not brand:
        return out
    recent_ids: list = []
    try:
        with db.connect() as conn:
            _rws = conn.execute(
                """SELECT a.*, r.started_at AS _run_started
                   FROM ads a JOIN runs r ON r.run_id = a.run_id
                   WHERE LOWER(COALESCE(a.brand, a.page_name)) = LOWER(?)
                   ORDER BY r.started_at ASC""",
                (brand,),
            ).fetchall()
            ads_all = [dict(x) for x in _rws]
            # Two most-recent META scrapes overall (runs that produced ads) — used so
            # "active now"/momentum reflect the CURRENT run, not the brand's last appearance.
            recent_ids = [r["run_id"] for r in conn.execute(
                """SELECT r.run_id FROM runs r JOIN ads a ON a.run_id = r.run_id
                   GROUP BY r.run_id ORDER BY r.started_at DESC LIMIT 2"""
            ).fetchall()]
    except Exception:
        return out
    if not ads_all:
        return out
    out["found"] = True

    # Latest observation per library_id (ads_all is oldest→newest, so last write wins)
    _latest: dict[str, dict] = {}
    for a in ads_all:
        _latest[a.get("library_id") or ""] = a
    latest_ads = list(_latest.values())

    out["distinct_ads"] = len(latest_ads)
    out["max_days"] = max((a.get("days_running") or 0) for a in latest_ads)
    out["score_sum"] = round(sum(float(a.get("score_normalized") or 0) for a in latest_ads), 1)
    out["ads_full"] = latest_ads

    def _common(key):
        c = _Counter(a.get(key) for a in latest_ads if a.get(key))
        return c.most_common(1)[0][0] if c else ""
    out["niche"] = _common("niche") or "—"
    out["category"] = _common("category")
    out["sub_category"] = _common("sub_category")
    out["location"] = _common("location")

    # Best marketplace signal seen across all observations
    for a in ads_all:
        s = a.get("mp_sold") or 0
        if s and s > out["mp_sold"]:
            out["mp_sold"] = s
            out["mp_price"] = a.get("mp_price")
            out["mp_rating"] = a.get("mp_rating")
            out["mp_reviews"] = a.get("mp_reviews") or 0
            out["mp_source"] = a.get("mp_source") or ""
    if out["mp_price"] is None:
        for a in latest_ads:
            if a.get("mp_price"):
                out["mp_price"] = a.get("mp_price")
                break

    # Source mix from landing URLs (FB ads always present)
    _srcs = {"FB"}
    for a in latest_ads:
        u = (a.get("landing_url") or "").lower()
        if "shopee" in u or "shp.ee" in u:
            _srcs.add("Shopee")
        elif "lazada" in u:
            _srcs.add("Lazada")
        elif "tiktok" in u:
            _srcs.add("TikTok")
    out["sources"] = ["FB"] + sorted(_srcs - {"FB"})
    out["source_count"] = len(_srcs)

    # Sample landing page (from the longest-running ad)
    for a in sorted(latest_ads, key=lambda x: -(x.get("days_running") or 0)):
        if a.get("landing_url"):
            out["sample_landing"] = a.get("landing_url")
            break

    # Per-run active counts (chronological) → momentum
    _by_run: dict = {}
    for a in ads_all:
        rid = a.get("run_id")
        e = _by_run.setdefault(rid, {"started": a.get("_run_started") or "", "active": 0})
        if a.get("is_active"):
            e["active"] += 1
    _per = [_by_run[k] for k in sorted(_by_run, key=lambda k: _by_run[k]["started"])]
    out["runs_seen"] = len(_per)
    out["per_run_active"] = [p["active"] for p in _per]
    if _per:
        out["first_seen"] = (_per[0]["started"] or "")[:10]
        out["last_seen"] = (_per[-1]["started"] or "")[:10]

    # "Active now" + momentum measured against the GLOBAL latest scrape(s), not the
    # brand's last appearance — a brand that dropped out reads 0 active / Cooling.
    _cur = recent_ids[0] if recent_ids else None
    _prev = recent_ids[1] if len(recent_ids) > 1 else None
    out["active_now"] = _by_run.get(_cur, {}).get("active", 0) if _cur is not None else 0
    if _cur is None:
        out["momentum"] = "🆕 New"
    elif _cur not in _by_run:
        out["momentum"] = "📉 Cooling"
    elif _prev is None or _prev not in _by_run:
        out["momentum"] = "🆕 New"
    else:
        _na, _np = out["active_now"], _by_run[_prev]["active"]
        out["momentum"] = "📈 Scaling" if _na > _np else ("📉 Cooling" if _na < _np else "➡ Steady")
    return out


def _winner_score(res: dict) -> dict:
    """0–100 composite buy signal from cross-run research, with a transparent
    breakdown and a plain verdict. Longevity + velocity + persistence +
    cross-source presence + marketplace demand."""
    days = res.get("max_days") or 0
    active = res.get("active_now") or 0
    runs = res.get("runs_seen") or 0
    sold = res.get("mp_sold") or 0
    srcs = res.get("source_count") or 1
    has_mp = bool(res.get("mp_sold") or res.get("mp_price"))

    if days >= 180:
        long_pts = 35
    elif days >= 90:
        long_pts = 30
    elif days >= 60:
        long_pts = 24
    elif days >= 30:
        long_pts = 16
    elif days >= 14:
        long_pts = 9
    else:
        long_pts = round(days / 14 * 9) if days else 0
    vel_pts = min(20, active * 4)
    pers_pts = min(15, runs * 3)
    src_pts = min(15, 5 + (5 if has_mp else 0) + (5 if srcs >= 3 else 0))
    if sold >= 5000:
        dem_pts = 15
    elif sold >= 1000:
        dem_pts = 11
    elif sold >= 300:
        dem_pts = 7
    elif sold >= 50:
        dem_pts = 4
    elif sold > 0:
        dem_pts = 2
    else:
        dem_pts = 0

    total = max(0, min(100, int(long_pts + vel_pts + pers_pts + src_pts + dem_pts)))
    if total >= 75:
        verdict, vc = "Strong buy", "#7CC4A0"
    elif total >= 55:
        verdict, vc = "Worth testing", "#D4AF37"
    elif total >= 35:
        verdict, vc = "Risky — dig deeper", "#E6CC73"
    else:
        verdict, vc = "Weak signal", "#E0909F"
    return {
        "total": total, "verdict": verdict, "verdict_color": vc,
        "momentum": res.get("momentum", "🆕 New"),
        "breakdown": [
            ("Longevity", long_pts, 35),
            ("Velocity", vel_pts, 20),
            ("Persistence", pers_pts, 15),
            ("Cross-source", src_pts, 15),
            ("Demand", dem_pts, 15),
        ],
    }


def _research_hero_html(res: dict, score: dict) -> str:
    """Premium research-verdict hero: Winner Score donut + verdict + signal grid + breakdown bars."""
    import html as _h
    import math as _m
    vc = score["verdict_color"]
    total = score["total"]
    r = 32
    circ = 2 * _m.pi * r
    dash = circ * total / 100.0
    donut = (
        f"<svg width='86' height='86' viewBox='0 0 86 86' style='flex-shrink:0'>"
        f"<circle cx='43' cy='43' r='{r}' fill='none' stroke='rgba(150,150,150,0.20)' stroke-width='8'/>"
        f"<circle cx='43' cy='43' r='{r}' fill='none' stroke='{vc}' stroke-width='8' stroke-linecap='round' "
        f"stroke-dasharray='{dash:.1f} {circ:.1f}' transform='rotate(-90 43 43)'/>"
        f"<text x='43' y='40' text-anchor='middle' font-size='21' font-weight='700' fill='{vc}' "
        f"font-family='Fraunces,Georgia,serif'>{total}</text>"
        f"<text x='43' y='55' text-anchor='middle' font-size='8' fill='#9a9a9a' letter-spacing='1.5'>SCORE</text>"
        f"</svg>"
    )
    days = res.get("max_days") or 0
    active = res.get("active_now") or 0
    runs = res.get("runs_seen") or 0
    sold = res.get("mp_sold") or 0

    def _sig(label, val):
        return (f"<div><div style='font-size:1.02rem;font-weight:700;color:var(--pra-text-strong);"
                f"font-variant-numeric:tabular-nums'>{val}</div>"
                f"<div style='font-size:0.58rem;letter-spacing:0.1em;text-transform:uppercase;"
                f"color:var(--pra-text-dim);margin-top:1px'>{label}</div></div>")
    _tiles = _sig("Longevity", f"{days}d") + _sig("Active ads", str(active)) + _sig("Scrapes", str(runs))
    _tiles += _sig("Units sold", f"{sold:,}" if sold else "—")
    if res.get("mp_price"):
        _tiles += _sig("Price", f"₱{float(res['mp_price']):,.0f}")
    signals = "<div style='display:flex;gap:18px;flex-wrap:wrap;margin-top:10px'>" + _tiles + "</div>"

    bars = ""
    for label, pts, mx in score["breakdown"]:
        pct = int(pts / mx * 100) if mx else 0
        bars += (
            f"<div style='display:flex;align-items:center;gap:8px;margin-top:4px'>"
            f"<div style='width:84px;font-size:0.62rem;color:var(--pra-text-muted);"
            f"text-transform:uppercase;letter-spacing:0.05em'>{label}</div>"
            f"<div style='flex:1;height:5px;background:var(--pra-subtle-bg);border-radius:3px;overflow:hidden'>"
            f"<div style='width:{pct}%;height:100%;background:{vc}'></div></div>"
            f"<div style='width:36px;text-align:right;font-size:0.62rem;color:var(--pra-text-dim);"
            f"font-variant-numeric:tabular-nums'>{int(pts)}/{mx}</div>"
            f"</div>"
        )
    verdict_pill = (
        f"<span style='display:inline-block;background:rgba(212,175,55,0.10);border:1px solid {vc};"
        f"color:{vc};padding:2px 13px;border-radius:20px;font-size:0.76rem;font-weight:700;"
        f"letter-spacing:0.03em'>{_h.escape(score['verdict'])}</span>"
        f"<span style='margin-left:10px;font-size:0.78rem;color:var(--pra-text-muted)'>"
        f"{_h.escape(str(score['momentum']))}</span>"
    )
    meta_line = ""
    _src_txt = " · ".join(res.get("sources") or ["FB"])
    if res.get("first_seen") and res.get("last_seen"):
        meta_line = (f"<div style='font-size:0.68rem;color:var(--pra-text-dim);margin-top:9px'>"
                     f"Sources: {_h.escape(_src_txt)} &nbsp;·&nbsp; tracked "
                     f"{_h.escape(str(res['first_seen']))} → {_h.escape(str(res['last_seen']))}</div>")
    return (
        "<div style='background:linear-gradient(135deg,rgba(212,175,55,0.05),rgba(212,175,55,0.01));"
        "border:1px solid var(--pra-border);border-radius:14px;padding:16px 20px;margin:4px 0 10px'>"
        "<div style='display:flex;gap:20px;align-items:center'>"
        + donut +
        "<div style='flex:1'>"
        "<div style='font-size:0.6rem;letter-spacing:0.16em;text-transform:uppercase;"
        "font-weight:700;color:var(--pra-accent);margin-bottom:7px'>Research verdict · Winner Score</div>"
        + verdict_pill + signals + meta_line +
        "</div></div>"
        "<div style='margin-top:13px'>" + bars + "</div>"
        "</div>"
    )


def _clear_brand_caches() -> None:
    """Invalidate all cross-run brand caches after the underlying ads/runs data changes
    (Refresh, classify, backfill) so Winner Score / niche / category never go stale."""
    for _fn in (_brand_research, _brand_history, _brand_price_history, _brands_trendlines):
        try:
            _fn.clear()
        except Exception:
            pass


@st.dialog("Welcome to Orbit — quick setup", width="large")
def _render_onboarding_wizard() -> None:
    """Multi-step first-launch wizard. Picks niches, optional Claude key,
    optional Telegram, runs first scrape recommendation.

    Phase 18.6 — replaces the simple onboarding banner with a guided flow."""
    _step = st.session_state.get("_onb_step", 1)
    _total_steps = 4

    # Progress indicator
    st.markdown(
        f"<div style='display:flex;align-items:center;gap:12px;margin-bottom:18px'>"
        f"<div style='color:var(--pra-text-muted);font-size:0.72rem;letter-spacing:0.14em;"
        f"text-transform:uppercase;font-weight:700'>Step {_step} of {_total_steps}</div>"
        f"<div style='flex:1;height:3px;background:var(--pra-border);border-radius:2px;overflow:hidden'>"
        f"<div style='background:var(--pra-accent);width:{(_step/_total_steps)*100}%;height:100%'></div>"
        f"</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    if _step == 1:
        # ---- Step 1: Welcome + niche picker ----
        st.markdown(
            "<div style='font-size:1.45rem;font-weight:600;letter-spacing:-0.015em;"
            "margin-bottom:4px'>Welcome to Orbit 🛰</div>"
            "<div style='color:var(--pra-text-muted);font-size:0.95rem;line-height:1.6;"
            "margin-bottom:18px'>Your AI-powered Philippine ecommerce research desk. "
            "Let's tailor it to your niches.</div>",
            unsafe_allow_html=True,
        )
        st.markdown("**Which niches do you sell or want to research?**")
        _all_niches = list(load_config().get("niches", {}).keys()) or [
            "capsule", "cream", "oil", "coffee",
        ]
        _picked_niches = st.multiselect(
            "Pick at least 1",
            _all_niches, key="_onb_niches",
            default=st.session_state.get("_onb_niches", ["capsule", "cream"]),
        )
        if st.button("Next →", key="onb_step1_next", type="primary",
                     width="stretch", disabled=not _picked_niches):
            st.session_state["filter_niche"] = _picked_niches  # pre-fill sidebar
            st.session_state["_onb_step"] = 2
            st.rerun()

    elif _step == 2:
        # ---- Step 2: Claude API key (optional) ----
        st.markdown(
            "<div style='font-size:1.25rem;font-weight:600;margin-bottom:4px'>"
            "AI features (optional)</div>"
            "<div style='color:var(--pra-text-muted);font-size:0.9rem;line-height:1.6;"
            "margin-bottom:14px'>"
            "Orbit's AI features — Copy Studio, classifier, brand angle reports, "
            "image prompts, receipt OCR — all use Claude (Anthropic). "
            "Costs ~₱0.30-₱2 per request.<br><br>"
            "<strong>You can skip this step</strong> — set the key later in Settings.</div>",
            unsafe_allow_html=True,
        )
        _key_set = bool(os.environ.get("ANTHROPIC_API_KEY"))
        if _key_set:
            st.success("✓ ANTHROPIC_API_KEY is already set. AI features will work.")
        else:
            st.info(
                "Not set. To enable: get a key from "
                "[console.anthropic.com](https://console.anthropic.com/), "
                "then in PowerShell:\n\n"
                "`$env:ANTHROPIC_API_KEY = \"sk-ant-...\"`\n\n"
                "Restart Orbit. You can do this any time."
            )
        _b1, _b2 = st.columns(2)
        if _b1.button("← Back", key="onb_step2_back", width="stretch"):
            st.session_state["_onb_step"] = 1
            st.rerun()
        if _b2.button("Next →", key="onb_step2_next", type="primary", width="stretch"):
            st.session_state["_onb_step"] = 3
            st.rerun()

    elif _step == 3:
        # ---- Step 3: Daily scrape scheduling (optional) ----
        st.markdown(
            "<div style='font-size:1.25rem;font-weight:600;margin-bottom:4px'>"
            "Automate daily research (optional)</div>"
            "<div style='color:var(--pra-text-muted);font-size:0.9rem;line-height:1.6;"
            "margin-bottom:14px'>"
            "Orbit can scrape Meta Ads Library + classify ads every morning via Windows "
            "Task Scheduler. By the time you open the app, fresh data is ready.<br><br>"
            "Skip this if you prefer manual scraping.</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "**Recommended setup:** open **Settings → Daily scrape scheduler** "
            "and click **Enable / Update** with a morning time like 06:00."
        )
        _b1, _b2 = st.columns(2)
        if _b1.button("← Back", key="onb_step3_back", width="stretch"):
            st.session_state["_onb_step"] = 2
            st.rerun()
        if _b2.button("Next →", key="onb_step3_next", type="primary", width="stretch"):
            st.session_state["_onb_step"] = 4
            st.rerun()

    elif _step == 4:
        # ---- Step 4: Done — point to next actions ----
        st.markdown(
            "<div style='font-size:1.45rem;font-weight:600;letter-spacing:-0.015em;"
            "margin-bottom:4px'>You're all set ✓</div>"
            "<div style='color:var(--pra-text-muted);font-size:0.95rem;line-height:1.6;"
            "margin-bottom:14px'>Three things to try next:</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "1. **Click a brand row** on the FB Ads page → opens the detail modal with "
            "timeline, similar brands, FDA compliance check, AI angle report.\n\n"
            "2. **Press `⌘ K` (Ctrl+K)** anywhere → quick navigation between pages and brands.\n\n"
            "3. **Visit Settings** → set up daily scheduler + check API status.\n\n"
            "4. **Press `?`** anywhere → see all keyboard shortcuts."
        )
        if st.button("🚀 Start exploring", key="onb_done", type="primary",
                     width="stretch"):
            mark_onboarding_done()
            st.session_state.pop("_onb_step", None)
            st.session_state["show_onboarding"] = False
            st.toast("✓ Setup complete. Welcome to Orbit!", icon=None)
            st.rerun()


@st.dialog("Keyboard shortcuts", width="medium")
def _render_shortcuts_overlay() -> None:
    """Show all available keyboard shortcuts."""
    st.markdown(
        "<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-bottom:14px'>"
        "Press <kbd style='background:var(--pra-subtle-bg);border:1px solid var(--pra-border);"
        "border-radius:3px;padding:1px 6px;font-family:monospace'>?</kbd> anywhere to "
        "show this panel. <kbd style='background:var(--pra-subtle-bg);border:1px solid var(--pra-border);"
        "border-radius:3px;padding:1px 6px;font-family:monospace'>Esc</kbd> to close."
        "</div>",
        unsafe_allow_html=True,
    )

    _SHORTCUTS = [
        ("Navigation", [
            ("Ctrl/Cmd + K",  "Open command palette · search pages & brands"),
            ("?",             "Show keyboard shortcuts (this panel)"),
            ("Esc",           "Close any open dialog"),
        ]),
        ("Actions", [
            ("Click row",     "Open brand detail modal"),
            ("Click card",    "Open brand detail modal (card view)"),
            ("Click bell",    "Open notification inbox"),
        ]),
        ("Tips", [
            ("Star brands",   "Mark shortlist brands from the brand detail modal"),
            ("Save searches", "Save current filter combination via sidebar"),
            ("Theme toggle",  "Switch light/dark mode from the sidebar footer"),
        ]),
    ]

    for group_name, items in _SHORTCUTS:
        st.markdown(
            f"<div style='color:var(--pra-accent);font-size:0.7rem;letter-spacing:0.14em;"
            f"font-weight:700;text-transform:uppercase;margin:14px 0 8px 0'>{group_name}</div>",
            unsafe_allow_html=True,
        )
        for key, desc in items:
            st.markdown(
                f"<div style='display:flex;justify-content:space-between;align-items:center;"
                f"padding:6px 0;border-bottom:1px solid var(--pra-border);font-size:0.85rem'>"
                f"<span style='color:var(--pra-text)'>{desc}</span>"
                f"<kbd style='background:var(--pra-subtle-bg);border:1px solid var(--pra-border);"
                f"border-radius:3px;padding:2px 8px;font-family:monospace;font-size:0.78rem;"
                f"color:var(--pra-text-muted);font-weight:500'>{key}</kbd>"
                f"</div>",
                unsafe_allow_html=True,
            )


@st.dialog("Notifications", width="large")
def _render_inbox(active_run_id: int, runs_list: list[dict], curr_rows: list[dict]) -> None:
    """Notification inbox modal — lists events since last visit."""
    from datetime import datetime as _dt_ib
    last_seen = inbox_last_seen()
    last_seen_label = (
        _dt_ib.fromisoformat(last_seen).strftime("%b %d, %H:%M")
        if last_seen else "your first visit"
    )

    st.markdown(
        f"<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-bottom:14px'>"
        f"Events since <strong>{last_seen_label}</strong>."
        f"</div>",
        unsafe_allow_html=True,
    )

    # Build events list
    events: list[dict] = []

    # New winners since previous run
    if len(runs_list) >= 2:
        # Find the run that started IMMEDIATELY BEFORE the active one (by timestamp).
        # Without this, users who pick an older run from the dropdown would diff against
        # a NEWER run (whichever was first in the list) — totally wrong baseline.
        _active_meta = next((r for r in runs_list if r["run_id"] == active_run_id), None)
        _active_ts = (_active_meta or {}).get("started_at", "") if _active_meta else ""
        prev_id = None
        if _active_ts:
            _candidates = [r for r in runs_list
                           if r.get("started_at", "") < _active_ts and r["run_id"] != active_run_id]
            _candidates.sort(key=lambda r: r.get("started_at", ""), reverse=True)
            if _candidates:
                prev_id = _candidates[0]["run_id"]
        if prev_id:
            try:
                prev_rows = db.ads_for_run(prev_id)
                if len(prev_rows) >= 50:
                    nw = suggestions.new_winners_since(prev_rows, curr_rows, threshold_days=30)
                    for s in nw[:10]:
                        events.append({
                            "icon": "↑", "type": "positive",
                            "title": s.title, "detail": s.detail,
                        })
                    rt = suggestions.retired_ads(prev_rows, curr_rows)
                    for s in rt[:5]:
                        events.append({
                            "icon": "↓", "type": "warning",
                            "title": s.title, "detail": s.detail,
                        })
                    # Phase 15.7 — anomaly detector
                    try:
                        import anomaly_detector
                        _starred = set(db.list_starred_brands())
                        _anomalies = anomaly_detector.detect(prev_rows, curr_rows, _starred)
                        # Prepend anomalies — they're higher signal than generic winner/retire
                        events = [
                            {"icon": a["icon"], "type": a["type"],
                             "title": a["title"], "detail": a["detail"]}
                            for a in _anomalies[:8]
                        ] + events
                    except Exception:
                        pass
            except Exception:
                pass

    # Latest scrape completion
    if runs_list:
        latest = runs_list[0]
        when = latest.get("started_at", "")
        if when and (not last_seen or when > last_seen):
            events.insert(0, {
                "icon": "✓", "type": "info",
                "title": f"Scrape #{latest['run_id']} completed",
                "detail": f"{latest.get('ads_count', latest.get('total_ads', 0)):,} ads captured · "
                          f"{(when or '')[:16].replace('T', ' ')}",
            })

    # Starred brand status changes
    starred = set(db.list_starred_brands())
    if starred:
        active_count = sum(
            1 for r in curr_rows
            if (r.get("brand") or r.get("page_name") or "") in starred and r.get("is_active")
        )
        events.append({
            "icon": "★", "type": "accent",
            "title": f"Shortlist status",
            "detail": f"{active_count} of {len(starred)} brands still active in latest run",
        })

    # Render events
    if not events:
        st.markdown(
            "<div style='color:var(--pra-text-muted);font-size:0.85rem;padding:24px 0;"
            "text-align:center;font-style:italic'>"
            "No new activity since your last visit."
            "</div>",
            unsafe_allow_html=True,
        )
    else:
        for ev in events[:15]:
            _type_colors = {
                "positive": "var(--pra-success)",
                "warning":  "var(--pra-warning)",
                "info":     "var(--pra-info)",
                "accent":   "var(--pra-accent)",
            }
            _c = _type_colors.get(ev["type"], "var(--pra-info)")
            st.markdown(
                f"<div style='display:flex;gap:12px;padding:10px 12px;"
                f"border-bottom:1px solid var(--pra-border);align-items:flex-start'>"
                f"<div style='flex-shrink:0;width:24px;height:24px;border-radius:4px;"
                f"background:rgba(212,175,55,0.1);display:flex;align-items:center;"
                f"justify-content:center;color:{_c};font-weight:700;font-size:0.85rem'>"
                f"{ev['icon']}</div>"
                f"<div style='flex:1'>"
                f"<div style='color:var(--pra-text);font-size:0.86rem;font-weight:500'>"
                f"{ev['title']}</div>"
                f"<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-top:2px'>"
                f"{ev['detail']}</div>"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    if st.button("Mark all as read", key="inbox_mark_read", width="stretch"):
        mark_inbox_seen()
        st.session_state.show_inbox = False
        st.toast("✓ Inbox cleared", icon=None)
        st.rerun()


@st.dialog("Quick navigation", width="large")
def _render_command_palette(curr_rows: list[dict]) -> None:
    """Command palette modal — search pages + brands by typing.

    Triggered via the ⌘K button in the page header, or by pressing Ctrl+K
    (the JS interceptor finds the trigger button by its test-id and clicks it).
    """
    st.markdown(
        "<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-bottom:10px'>"
        "Search pages or brands. <kbd style='background:var(--pra-subtle-bg);"
        "border:1px solid var(--pra-border);border-radius:3px;padding:1px 6px;"
        "font-size:0.72rem;font-family:monospace'>Esc</kbd> to close."
        "</div>",
        unsafe_allow_html=True,
    )

    query = st.text_input(
        "Search",
        placeholder="Type to filter pages or brand names…",
        key="palette_query",
        label_visibility="collapsed",
    )
    q_low = (query or "").lower().strip()

    # ---- Build candidates: pages + brands ----
    # Pages — all nav items with their icons + labels
    page_results = []
    for _section, _items in _NAV_SECTIONS:
        for _pid, _label, _icon in _items:
            score = 0
            label_low = _label.lower()
            if not q_low:
                score = 1  # show all when query is empty
            elif q_low in label_low:
                score = 100 if label_low.startswith(q_low) else 50
            elif q_low in _pid.lower():
                score = 30
            if score:
                page_results.append({
                    "kind": "page", "pid": _pid, "label": _label, "icon": _icon,
                    "section": _section, "score": score,
                })

    # Brands — from current run's rows
    brand_results = []
    if q_low and len(q_low) >= 2:
        _seen_brands: set = set()
        for r in curr_rows:
            b = (r.get("brand") or r.get("page_name") or "").strip()
            if not b or b in _seen_brands:
                continue
            b_low = b.lower()
            if q_low in b_low:
                score = 100 if b_low.startswith(q_low) else 50
                _seen_brands.add(b)
                brand_results.append({
                    "kind": "brand", "brand": b, "niche": r.get("niche") or "—",
                    "days": r.get("days_running") or 0, "score": score,
                })

    page_results.sort(key=lambda x: -x["score"])
    brand_results.sort(key=lambda x: -x["score"])

    # ---- Render results ----
    if page_results:
        st.markdown(
            "<div style='color:var(--pra-text-dim);font-size:0.62rem;letter-spacing:0.16em;"
            "font-weight:700;text-transform:uppercase;margin:14px 0 6px 0'>Pages</div>",
            unsafe_allow_html=True,
        )
        for p in page_results[:8]:
            if st.button(
                f"{p['icon']}\xa0\xa0{p['label']}\xa0\xa0\xa0\xa0— {p['section']}",
                key=f"palette_p_{p['pid']}",
                width="stretch",
                type="secondary",
            ):
                st.session_state.current_page = p["pid"]
                st.session_state.show_palette = False
                st.rerun()

    if brand_results:
        st.markdown(
            "<div style='color:var(--pra-text-dim);font-size:0.62rem;letter-spacing:0.16em;"
            "font-weight:700;text-transform:uppercase;margin:14px 0 6px 0'>Brands</div>",
            unsafe_allow_html=True,
        )
        for b in brand_results[:10]:
            if st.button(
                f"◉  {b['brand']}\xa0\xa0\xa0\xa0— {b['niche']} · {b['days']:,}d",
                key=f"palette_b_{b['brand']}",
                width="stretch",
                type="secondary",
            ):
                st.session_state["_palette_open_brand"] = b["brand"]
                st.session_state.current_page = "fb_ads"
                st.session_state.show_palette = False
                st.rerun()

    if q_low and not page_results and not brand_results:
        st.markdown(
            f"<div style='color:var(--pra-text-muted);font-size:0.85rem;padding:18px 0;"
            f"text-align:center'>No matches for <strong>{query}</strong></div>",
            unsafe_allow_html=True,
        )


def _daily_brief(active_run_id: int, runs_list: list[dict], curr_rows: list[dict]) -> dict:
    """Compute the rich Daily Brief: narrative intel items + top-action recommendation.

    Returns dict:
        - date_str: 'Wednesday, May 26'
        - greeting: 'Good morning' etc.
        - items: list of {icon, text, type} dicts (positive/warning/info/accent)
        - top_action: {brand, reason} or None
        - first_run: bool — whether this is the only run (no prior to diff against)
    """
    from datetime import datetime as _dt_b
    out = {
        "date_str": _dt_b.now().strftime("%A, %b %d"),
        "greeting": "Good morning" if _dt_b.now().hour < 12
                    else "Good afternoon" if _dt_b.now().hour < 18
                    else "Good evening",
        "items": [],
        "top_action": None,
        "first_run": len(runs_list) < 2,
    }

    # Find the previous run for diffing
    prev_rows: list[dict] = []
    if len(runs_list) >= 2:
        # Find the run that started IMMEDIATELY BEFORE the active one (by timestamp).
        # Without this, users who pick an older run from the dropdown would diff against
        # a NEWER run (whichever was first in the list) — totally wrong baseline.
        _active_meta = next((r for r in runs_list if r["run_id"] == active_run_id), None)
        _active_ts = (_active_meta or {}).get("started_at", "") if _active_meta else ""
        prev_id = None
        if _active_ts:
            _candidates = [r for r in runs_list
                           if r.get("started_at", "") < _active_ts and r["run_id"] != active_run_id]
            _candidates.sort(key=lambda r: r.get("started_at", ""), reverse=True)
            if _candidates:
                prev_id = _candidates[0]["run_id"]
        if prev_id:
            try:
                prev_rows = db.ads_for_run(prev_id)
            except Exception:
                prev_rows = []

    # --- Item 1: New 30-day winners ---
    if prev_rows and len(prev_rows) >= max(50, len(curr_rows) // 2):
        try:
            new_w = suggestions.new_winners_since(prev_rows, curr_rows, threshold_days=30)
            n_new = len(new_w)
            if n_new:
                out["items"].append({
                    "icon": "↑", "type": "positive",
                    "text": f"<strong>{n_new}</strong> new 30-day winner{'s' if n_new != 1 else ''} since last run",
                })
        except Exception:
            pass

        # --- Item 2: Retired ads ---
        try:
            ret = suggestions.retired_ads(prev_rows, curr_rows)
            n_ret = len(ret)
            if n_ret:
                out["items"].append({
                    "icon": "↓", "type": "warning",
                    "text": f"<strong>{n_ret}</strong> ad{'s' if n_ret != 1 else ''} retired (stopped running)",
                })
        except Exception:
            pass

    # --- Item 3: Shortlist status ---
    starred = set(db.list_starred_brands())
    if starred:
        active_starred = sum(
            1 for r in curr_rows
            if (r.get("brand") or r.get("page_name") or "") in starred and r.get("is_active")
        )
        out["items"].append({
            "icon": "★", "type": "accent",
            "text": f"<strong>{active_starred}</strong> of <strong>{len(starred)}</strong> "
                    f"shortlist brand{'s' if len(starred) != 1 else ''} still active",
        })

    # --- Item 4: Hot niche (most active brands) ---
    by_niche_brands: dict[str, set] = defaultdict(set)
    for r in curr_rows:
        if not r.get("is_active"):
            continue
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        n, b = r.get("niche"), (r.get("brand") or r.get("page_name"))
        if n and b:
            by_niche_brands[n].add(b)
    if by_niche_brands:
        hot = max(by_niche_brands.items(), key=lambda x: len(x[1]))
        out["items"].append({
            "icon": "▲", "type": "info",
            "text": f"<strong>{hot[0]}</strong> niche leads with "
                    f"<strong>{len(hot[1])}</strong> active brand{'s' if len(hot[1]) != 1 else ''}",
        })

    # --- Item 5: Total PH-confident in view ---
    n_ph = sum(1 for r in curr_rows if r.get("geo_signal") == "ph-confident")
    n_active = sum(1 for r in curr_rows if r.get("is_active"))
    out["items"].append({
        "icon": "●", "type": "info",
        "text": f"<strong>{n_ph:,}</strong> PH-confident ads · <strong>{n_active:,}</strong> still running",
    })

    # --- Item 6: New brand surge (brand with biggest ad count gain) ---
    if prev_rows:
        prev_brand_counts: dict[str, int] = defaultdict(int)
        curr_brand_counts: dict[str, int] = defaultdict(int)
        for r in prev_rows:
            if r.get("is_active"):
                b = r.get("brand") or r.get("page_name")
                if b:
                    prev_brand_counts[b] += 1
        for r in curr_rows:
            if r.get("is_active"):
                b = r.get("brand") or r.get("page_name")
                if b:
                    curr_brand_counts[b] += 1
        gains = [
            (b, curr_brand_counts[b] - prev_brand_counts.get(b, 0))
            for b in curr_brand_counts
            if (curr_brand_counts[b] - prev_brand_counts.get(b, 0)) >= 3
        ]
        gains.sort(key=lambda x: -x[1])
        if gains:
            top_brand, gain = gains[0]
            out["top_action"] = {
                "brand": top_brand,
                "reason": f"added <strong>{gain}</strong> new active variants since last run",
            }
            out["items"].append({
                "icon": "⚡", "type": "accent",
                "text": f"<strong>{top_brand}</strong> is scaling — +{gain} new variants",
            })

    return out


def _render_daily_brief(brief: dict) -> None:
    """Render the daily brief as a sleek HTML card."""
    if not brief["items"] and brief["first_run"]:
        # First-run state — encourage second scrape
        st.markdown(
            f"<div class='pra-brief'>"
            f"<div class='pra-brief-header'>"
            f"<span class='pra-brief-eyebrow'>{brief['greeting']} · Daily Brief</span>"
            f"<span class='pra-brief-date'>{brief['date_str']}</span>"
            f"</div>"
            f"<div class='pra-brief-empty'>"
            f"This is your first run — run another scrape tomorrow to unlock "
            f"day-over-day intelligence (new winners, retirements, scaling brands)."
            f"</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
        return

    items_html = ""
    for item in brief["items"]:
        items_html += (
            f"<div class='pra-brief-item {item['type']}'>"
            f"<div class='pra-brief-item-icon'>{item['icon']}</div>"
            f"<div class='pra-brief-item-text'>{item['text']}</div>"
            f"</div>"
        )

    # Card body without the CTA (CTA becomes a real st.button rendered right after)
    st.markdown(
        f"<div class='pra-brief'>"
        f"<div class='pra-brief-header'>"
        f"<span class='pra-brief-eyebrow'>{brief['greeting']} · Daily Brief</span>"
        f"<span class='pra-brief-date'>{brief['date_str']}</span>"
        f"</div>"
        f"<div class='pra-brief-items'>{items_html}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )
    # Real clickable CTA (only when there's a top action to recommend)
    if brief.get("top_action"):
        ta = brief["top_action"]
        # Strip HTML tags from the reason for the button text
        import re as _re_clean
        _reason_text = _re_clean.sub(r"<[^>]+>", "", ta["reason"])
        _cta_l, _cta_r = st.columns([4, 1])
        with _cta_l:
            st.markdown(
                f"<div style='color:var(--pra-text-muted);font-size:0.82rem;"
                f"padding:6px 4px 0 4px'>"
                f"Today's focus → check <strong style='color:var(--pra-text)'>"
                f"{ta['brand']}</strong> · <span style='color:var(--pra-text-muted)'>"
                f"{_reason_text}</span></div>",
                unsafe_allow_html=True,
            )
        with _cta_r:
            if st.button("→ Open FB Ads", key="brief_cta_btn",
                         width="stretch", type="secondary",
                         help=f"Jump to FB Ads page to dig into {ta['brand']}"):
                st.session_state.current_page = "fb_ads"
                st.rerun()


# ---- Hunter decision pipeline (Shortlist / Decision Board) ----
# The pure-hunter flow stops at the decision point. Older builds used an
# execution-flavored vocab (investigating/validating/launching/launched/passed);
# those legacy values are migrated to the 4 hunter stages on read so no data is lost.
_HUNT_PIPELINE = ["researching", "validated", "sourced", "decided"]
_STATUS_LABEL = {
    "researching": "Researching",
    "validated":   "Validated",
    "sourced":     "Sourced",
    "decided":     "Decided",
    # legacy values still render if any old rows remain
    "investigating": "Researching",
    "validating":    "Validated",
    "launching":     "Sourced",
    "launched":      "Decided",
    "passed":        "Decided",
}
_STATUS_MIGRATE = {
    "investigating": "researching",
    "validating":    "validated",
    "launching":     "sourced",
    "launched":      "decided",
    "passed":        "decided",
}
_STATUS_COLOR = {
    "researching": "#8FC2B4",   # info teal — just starting
    "validated":   "#E6CC73",   # amber — demand confirmed
    "sourced":     "#D4AF37",   # brass — supplier found
    "decided":     "#7CC4A0",   # green — call made
    # legacy aliases (kept so old rows still colour correctly)
    "investigating": "#8FC2B4",
    "validating":    "#E6CC73",
    "launching":     "#D4AF37",
    "launched":      "#7CC4A0",
    "passed":        "#7CC4A0",
}


def _canon_status(s: str) -> str:
    """Map any stored status (incl. legacy execution vocab) to a canonical hunter stage."""
    s = (s or "").strip().lower()
    return _STATUS_MIGRATE.get(s, s)


def _section_label(text: str) -> None:
    st.markdown(
        f"<div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.12em;"
        f"font-weight:600;text-transform:uppercase;margin:18px 0 8px 0'>{text}</div>",
        unsafe_allow_html=True,
    )


@st.dialog("Brand details", width="large")
def _render_brand_detail_dialog(brand: str, brand_row: dict, ads_for_brand: list[dict]) -> None:
    """Modal showing star/status/notes editor + brand-level data + variant list."""
    meta = db.get_brand_meta(brand) or {}
    is_starred = bool(meta.get("starred"))
    current_status = _canon_status(meta.get("status", "") or "")
    _research = _brand_research(brand)
    _wscore = _winner_score(_research)
    import html as _dh  # escape scraped (advertiser-controlled) values before unsafe_allow_html

    # ---- Header: brand + state badges ----
    badges_html = ""
    if is_starred:
        badges_html += (
            "<span style='display:inline-block;background:rgba(212,175,55,0.15);"
            "border:1px solid #D4AF37;color:#D4AF37;padding:2px 8px;border-radius:3px;"
            "font-size:0.7rem;font-weight:600;letter-spacing:0.08em;text-transform:uppercase;"
            "margin-right:6px'>★ Shortlist</span>"
        )
    if current_status:
        sc = _STATUS_COLOR.get(current_status, "#8FC2B4")
        badges_html += (
            f"<span style='display:inline-block;background:rgba(155,170,200,0.1);"
            f"border:1px solid {sc};color:{sc};padding:2px 8px;border-radius:3px;"
            f"font-size:0.7rem;font-weight:600;letter-spacing:0.08em;text-transform:uppercase'>"
            f"● {_STATUS_LABEL.get(current_status, current_status.capitalize())}</span>"
        )

    # Phase 14: surface category / sub_category / location in the dialog header
    _cat = brand_row.get("category") or ""
    _sub = brand_row.get("sub_category") or ""
    _loc = brand_row.get("location") or ""
    _tag_bits = []
    if _cat:
        _sub_pretty = _sub.replace("-", " ").replace("_", " ").title() if _sub else ""
        _tag_bits.append(
            f"{_dh.escape(_cat)}" + (f" › <strong>{_dh.escape(_sub_pretty)}</strong>" if _sub_pretty else "")
        )
    if _loc:
        _tag_bits.append(f"📍 {_dh.escape(_loc)}")
    _tag_line = (
        f"<div style='color:var(--pra-accent);font-size:0.78rem;margin-bottom:4px;"
        f"letter-spacing:0.01em'>{' · '.join(_tag_bits)}</div>"
        if _tag_bits else ""
    )

    st.markdown(
        f"<div style='font-size:1.5rem;font-weight:600;letter-spacing:-0.015em;margin-bottom:4px'>{_dh.escape(str(brand))}</div>"
        f"{_tag_line}"
        f"<div style='color:var(--pra-text-muted);font-size:0.82rem;margin-bottom:10px'>"
        f"{_dh.escape((brand_row.get('niche') or '—').replace('● ',''))} · {brand_row.get('ad_count',0)} ads · "
        f"max {brand_row.get('max_days_running',0):,} days · "
        f"score {int(brand_row.get('score_normalized') or 0):,}"
        f"</div>"
        f"<div>{badges_html}</div>",
        unsafe_allow_html=True,
    )

    # ---- Research verdict hero (cross-run Winner Score + signals) ----
    if _research.get("found"):
        st.markdown(_research_hero_html(_research, _wscore), unsafe_allow_html=True)

    # ---- Actions row: star + competitor + status ----
    _section_label("Actions")
    _is_competitor = bool(meta.get("competitor"))
    col_star, col_comp, col_status = st.columns([1, 1, 2])
    star_label = "★ Shortlisted" if is_starred else "☆ Shortlist"
    if col_star.button(star_label, width="stretch", type=("secondary" if is_starred else "primary"),
                        key=f"star_btn_{brand}"):
        if is_starred:
            db.upsert_brand_meta(brand, starred=0)
        else:
            # Entering the shortlist enters the pipeline at 'Researching' (matches Radar ★).
            db.upsert_brand_meta(brand, starred=1, status=(current_status or "researching"))
        log_activity("unstarred" if is_starred else "starred", brand)
        st.rerun()
    comp_label = "◎ Competitor ✓" if _is_competitor else "◎ Mark competitor"
    if col_comp.button(comp_label, width="stretch",
                       type=("secondary" if _is_competitor else "secondary"),
                       key=f"comp_btn_{brand}",
                       help="Track this brand on the Competitors watchboard"):
        db.toggle_competitor(brand)
        st.toast(("Removed from" if _is_competitor else "Added to") + " competitors", icon=None)
        st.rerun()

    status_options = [""] + _HUNT_PIPELINE
    if current_status not in status_options:
        current_status = ""
    new_status = col_status.selectbox(
        "Pipeline status",
        status_options,
        index=status_options.index(current_status),
        key=f"status_select_{brand}",
        format_func=lambda x: "— no status —" if x == "" else _STATUS_LABEL.get(x, x.capitalize()),
        label_visibility="collapsed",
    )
    if new_status != current_status:
        db.upsert_brand_meta(brand, status=new_status)
        st.toast(f"Status: {new_status or 'cleared'}", icon=None)
        st.rerun()

    # ---- Notes ----
    _section_label("Notes")
    notes_val = st.text_area(
        "Notes",
        value=meta.get("notes", ""),
        height=100,
        placeholder="Checked Shopee — 5k sold · Test order placed 5/12 · Supplier found on 1688",
        key=f"notes_area_{brand}",
        label_visibility="collapsed",
    )
    if st.button("Save notes", key=f"notes_save_{brand}"):
        db.upsert_brand_meta(brand, notes=notes_val)
        st.toast("Notes saved", icon=None)

    _lp_col, _test_col = st.columns([3, 2])
    if brand_row.get("sample_landing_url"):
        _lp_col.link_button("Open landing page", brand_row["sample_landing_url"],
                            type="primary", width="stretch")
    if _test_col.button("⚙ Find supplier", key=f"modal_sup_{brand}",
                          help="Source this product on 1688 (Supplier page)",
                          width="stretch"):
        st.session_state.current_page = "supplier"
        st.rerun()

    # ---- AI Brand Angle Report (Phase 17.1) ----
    _section_label("AI brand analysis")
    _angle_key = f"angle_report_{brand}"
    _has_anth_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
    _angle_l, _angle_r = st.columns([3, 1])
    with _angle_l:
        st.caption(
            "Claude reads this brand's ads and writes a 1-paragraph 'why they win' "
            "analyst report — winning angle, target demo, hook patterns, "
            "supplier hints, next moves. ~₱2 per report."
        )
    with _angle_r:
        if st.button(
            "🧠 Generate report", key=f"angle_btn_{brand}",
            type="primary", width="stretch",
            disabled=not _has_anth_key,
            help=("Run Claude analysis" if _has_anth_key
                  else "ANTHROPIC_API_KEY env var not set"),
        ):
            with st.spinner("Claude is analyzing this brand..."):
                import brand_angle_report
                st.session_state[_angle_key] = brand_angle_report.generate(
                    brand, ads_for_brand, brand_row,
                )
    _angle = st.session_state.get(_angle_key)
    if _angle:
        if not _angle["ok"]:
            st.error(f"Report failed: {_angle['error']}")
        else:
            st.markdown(
                f"<div style='background:linear-gradient(135deg,rgba(212,175,55,0.04) 0%,"
                f"rgba(212,175,55,0.01) 100%);border:1px solid var(--pra-border);"
                f"border-left:2px solid var(--pra-accent);border-radius:6px;"
                f"padding:14px 18px;margin-bottom:8px'>"
                f"<div style='color:var(--pra-accent);font-size:0.66rem;letter-spacing:0.16em;"
                f"font-weight:700;text-transform:uppercase;margin-bottom:8px'>"
                f"AI analyst report · {_angle['ads_used']} ads analysed</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
            st.markdown(_angle["report_markdown"])

    # ---- Quick research actions ----
    _section_label("Quick research")
    from urllib.parse import quote_plus as _qp_dlg
    qa1, qa2, qa3, qa4 = st.columns(4)
    _q = _qp_dlg(brand)
    qa1.link_button("Shopee PH", f"https://shopee.ph/search?keyword={_q}", width="stretch",
                     help="Search Shopee Philippines for this brand")
    qa2.link_button("Lazada PH", f"https://www.lazada.com.ph/catalog/?q={_q}", width="stretch",
                     help="Search Lazada Philippines")
    qa3.link_button("1688 supplier", f"https://s.1688.com/selloffer/offer_search.htm?keywords={_q}",
                     width="stretch", help="Find similar products from 1688 (Alibaba China wholesale)")
    qa4.link_button("Google review", f"https://www.google.com/search?q={_qp_dlg(brand + ' philippines review')}",
                     width="stretch", help="Search Google for reviews")

    # ---- Google Trends (PH search interest) ----
    _section_label("PH search demand · Google Trends")
    _tr_key = f"trend_{brand}"
    if _tr_key not in st.session_state:
        st.session_state[_tr_key] = None
    if st.button("📈 Fetch 12-month PH trend", key=f"trend_btn_{brand}",
                  help="Calls Google Trends API for search interest in PH market"):
        with st.spinner("Fetching trend..."):
            import trends as _trends
            st.session_state[_tr_key] = _trends.fetch_trend(brand, months=12, geo="PH")
    _tr = st.session_state[_tr_key]
    if _tr:
        if _tr.get("error"):
            st.caption(f"⚠ {_tr['error']}")
        elif _tr.get("values"):
            import pandas as _pd_tr
            import trends as _trends
            _df_tr = _pd_tr.DataFrame({"interest": _tr["values"]}, index=_tr["labels"])
            st.line_chart(_df_tr, height=180)
            st.caption(
                f"Avg interest: **{_tr['avg']}/100** · {_trends.trend_summary(_tr)} · "
                f"(0 = no search; 100 = peak in window)"
            )

    # ---- Similar brands (collapsible — secondary research path) ----
    with st.expander("Similar brands · same niche or hook patterns", expanded=False):
        _sim_list: list[dict] = []
        try:
            import brand_similarity as _bs
            _sim_list = _bs.find_similar(brand, rows, top_n=3)
        except Exception as _e_sim:
            st.caption(f"⚠ Similar-brand engine unavailable: {_e_sim}")
        if not _sim_list:
            st.caption("No similar brands found in this run.")
        else:
            for _s in _sim_list:
                _reasons = " · ".join(_s["reasons"])
                st.markdown(
                    f"<div style='padding:8px 12px;border:1px solid var(--pra-border);"
                    f"border-radius:4px;margin-bottom:6px;background:var(--pra-panel)'>"
                    f"<div style='display:flex;justify-content:space-between;align-items:baseline'>"
                    f"<span style='font-weight:500;color:var(--pra-text)'>{_dh.escape(str(_s['brand']))}</span>"
                    f"<span style='color:var(--pra-accent);font-size:0.75rem;font-weight:600'>"
                    f"score {_s['score']:.0f}</span>"
                    f"</div>"
                    f"<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-top:4px'>"
                    f"{_reasons}"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    # ---- FDA compliance check (Phase 16.1) ----
    try:
        import fda_compliance
        _fda_result = fda_compliance.scan_brand(ads_for_brand)
        if _fda_result["worst_severity"]:
            _sev = _fda_result["worst_severity"]
            _color = fda_compliance.SEVERITY_COLORS.get(_sev, "#8FC2B4")
            _sev_label = {
                "critical": "FDA CRITICAL · do not copy this hook",
                "high":     "FDA risk · review before adopting",
                "medium":   "FDA caution · soften the language",
                "low":      "FDA notice · reframe possible",
            }.get(_sev, _sev)
            _findings_html = ""
            for f in _fda_result["sample_matches"]:
                _fc = fda_compliance.SEVERITY_COLORS.get(f["severity"], "#8FC2B4")
                _findings_html += (
                    f"<div style='padding:6px 0;border-bottom:1px solid var(--pra-border);"
                    f"font-size:0.82rem;color:var(--pra-text)'>"
                    f"<span style='display:inline-block;background:rgba(224,127,159,0.12);"
                    f"border:1px solid {_fc};color:{_fc};padding:1px 6px;border-radius:3px;"
                    f"font-size:0.66rem;font-weight:700;letter-spacing:0.08em;"
                    f"text-transform:uppercase;margin-right:8px'>{f['severity']}</span>"
                    f"<strong>{f['label']}</strong> — "
                    f"<span style='color:var(--pra-text-muted)'>match: \"{f['match']}\"</span>"
                    f"<div style='color:var(--pra-text-dim);font-size:0.74rem;"
                    f"margin-top:2px;margin-left:6px;line-height:1.45'>{f['why']}</div>"
                    f"</div>"
                )
            with st.expander(
                f"⚠ FDA compliance check · {_sev_label} ({_fda_result['total_findings']} finding"
                f"{'s' if _fda_result['total_findings'] != 1 else ''})",
                expanded=(_sev in ("critical", "high")),
            ):
                st.markdown(
                    f"<div style='background:rgba(224,127,159,0.04);"
                    f"border:1px solid var(--pra-border);border-left:2px solid {_color};"
                    f"border-radius:4px;padding:10px 14px;font-size:0.82rem'>"
                    f"<div style='color:var(--pra-text-muted);font-size:0.78rem;"
                    f"margin-bottom:8px'>"
                    f"Their ad copy contains language that could trigger FDA action in PH. "
                    f"<strong>Don't copy hooks blindly</strong> — adopt the format, sanitize the claims.</div>"
                    f"{_findings_html}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
    except Exception as _e_fda:
        pass  # FDA check is enhancement, not critical

    # ---- Brand cluster (same operator detection) ----
    try:
        import brand_clusters as _bc
        # Cluster ALL brands in the current run, then check if this brand is in a cluster
        _all_clusters = _bc.find_clusters(rows, threshold=0.35, min_brand_ads=2)
        _my_cluster = next(
            (c for c in _all_clusters if brand in c["brands"]),
            None,
        )
        if _my_cluster and _my_cluster["size"] >= 2:
            _section_label("Brand cluster · same operator detected")
            _other_brands = [b for b in _my_cluster["brands"] if b != brand]
            _feature_strs = []
            for f in _my_cluster["shared_features"][:5]:
                if f.startswith("dom:"):
                    _feature_strs.append(f"landing {f[4:]}")
                elif f.startswith("niche:"):
                    _feature_strs.append(f"{f[6:]} niche")
                elif f.startswith("bigram:"):
                    _feature_strs.append(f'"{f[7:].replace("_", " ")}"')
                elif f.startswith("path:"):
                    _feature_strs.append(f"store {f[5:]}")
            _other_html = " · ".join(
                f"<span style='color:var(--pra-accent);font-weight:500'>{_dh.escape(str(b))}</span>"
                for b in _other_brands[:5]
            )
            st.markdown(
                f"<div style='background:rgba(230,204,115,0.06);"
                f"border:1px solid var(--pra-border);border-left:2px solid var(--pra-warning);"
                f"border-radius:4px;padding:12px 16px;font-size:0.85rem;color:var(--pra-text)'>"
                f"This brand appears to be run by the same operator as "
                f"<strong>{len(_other_brands)}</strong> other brand"
                f"{'s' if len(_other_brands) != 1 else ''}: {_other_html}"
                + (f"<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-top:6px'>"
                   f"Shared signals: {' · '.join(_feature_strs)}</div>" if _feature_strs else "")
                + "</div>",
                unsafe_allow_html=True,
            )
    except Exception as _e_bc:
        pass  # silent failure — cluster detection is enhancement, not critical

    # ---- Brand timeline (collapsible) ----
    # ---- Price history (Phase 20.4) ----
    _price_hist = _brand_price_history(brand)
    if _price_hist:
        with st.expander(f"Price history · {len(_price_hist)} data point"
                         f"{'s' if len(_price_hist) != 1 else ''}", expanded=False):
            import pandas as _pd_ph
            _df_ph = _pd_ph.DataFrame(_price_hist).set_index("date")
            if len(_price_hist) >= 2:
                st.line_chart(_df_ph[["price"]], height=160, color=["#7CC4A0"])
                _first = _price_hist[0]["price"]
                _last = _price_hist[-1]["price"]
                _delta = _last - _first
                _delta_pct = (_delta / _first * 100) if _first else 0
                _color = "var(--pra-danger)" if _delta > 0 else ("var(--pra-success)" if _delta < 0 else "var(--pra-text-muted)")
                _arrow = "↑" if _delta > 0 else ("↓" if _delta < 0 else "→")
                st.markdown(
                    f"<div style='font-size:0.84rem;color:var(--pra-text)'>"
                    f"Price moved from <strong>₱{_first:,.0f}</strong> to "
                    f"<strong>₱{_last:,.0f}</strong> "
                    f"<span style='color:{_color};font-weight:600'>"
                    f"{_arrow} {abs(_delta_pct):.0f}%</span></div>",
                    unsafe_allow_html=True,
                )
            else:
                st.caption(
                    f"Current price: **₱{_price_hist[0]['price']:,.0f}** "
                    f"(need 2+ enriched runs to show a trend)."
                )

    with st.expander("Timeline · history across all runs", expanded=False):
        _hist = _brand_history(brand)
        if _hist["total_runs_seen"] < 2:
            st.caption(
                f"Brand seen in {_hist['total_runs_seen']} run{'s' if _hist['total_runs_seen'] != 1 else ''} so far. "
                "Run more scrapes over time to build a richer timeline."
            )
        else:
            # Header row: first/last seen + total run appearances
            _tl_c1, _tl_c2, _tl_c3 = st.columns(3)
            _tl_c1.metric("First seen", _hist["first_seen"] or "—")
            _tl_c2.metric("Last seen", _hist["last_seen"] or "—")
            _tl_c3.metric("In runs", f"{_hist['total_runs_seen']}")

            # Chart: active ads per run
            import pandas as _pd_tl
            _df_tl = _pd_tl.DataFrame([
                {"run": r["started_at"][:10], "active": r["active_ads"], "max_days": r["max_days"]}
                for r in _hist["per_run"]
            ])
            if not _df_tl.empty:
                _df_tl = _df_tl.set_index("run")
                st.line_chart(_df_tl[["active"]], height=160,
                              color=["#D4AF37"])
                st.caption(
                    "Active ad count per run over time. Rising = brand scaling; "
                    "falling = brand losing momentum or pivoting."
                )

            # Milestones
            if _hist["milestones"]:
                _ms_html = (
                    "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                    "border-radius:4px;padding:12px 16px;margin-top:8px'>"
                    "<div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.12em;"
                    "font-weight:600;text-transform:uppercase;margin-bottom:8px'>Longevity milestones</div>"
                )
                for ms in _hist["milestones"]:
                    _ms_html += (
                        f"<div style='display:flex;justify-content:space-between;align-items:center;"
                        f"padding:4px 0;font-size:0.85rem'>"
                        f"<span style='color:var(--pra-text)'>"
                        f"<span style='color:var(--pra-accent);font-weight:600'>{ms['days_crossed']}d</span>"
                        f" longevity crossed</span>"
                        f"<span style='color:var(--pra-text-muted);font-variant-numeric:tabular-nums'>"
                        f"{ms['date']}</span>"
                        f"</div>"
                    )
                _ms_html += "</div>"
                st.markdown(_ms_html, unsafe_allow_html=True)

            # Longest-lived ad callout
            if _hist["longest_lived"] and _hist["longest_lived"]["days_running"]:
                _ll = _hist["longest_lived"]
                st.markdown(
                    f"<div style='background:rgba(212,175,55,0.06);"
                    f"border:1px solid var(--pra-border);border-left:2px solid var(--pra-accent);"
                    f"border-radius:4px;padding:10px 14px;margin-top:8px;font-size:0.82rem;"
                    f"color:var(--pra-text)'>"
                    f"<strong style='color:var(--pra-accent)'>Hero ad</strong> — "
                    f"{_ll['days_running']:,} days running"
                    + (f" (since {_dh.escape(str(_ll['start_date']))})" if _ll.get('start_date') else "")
                    + (f"<div style='color:var(--pra-text-muted);font-size:0.78rem;"
                       f"margin-top:4px;line-height:1.55'>{_dh.escape(_ll['ad_text'])}…</div>" if _ll.get('ad_text') else "")
                    + "</div>",
                    unsafe_allow_html=True,
                )

    # ---- Variant list (collapsible — the ads themselves, secondary) ----
    with st.expander(f"Ad variants ({len(ads_for_brand)})", expanded=False):
     if not ads_for_brand:
        st.caption("No variants in current filter.")
     else:
        ads_sorted = sorted(ads_for_brand, key=lambda a: -(a.get("days_running") or 0))
        for ad in ads_sorted[:8]:
            is_active = ad.get("is_active")
            color = "#7CC4A0" if is_active else "#8B6B6B"
            label = "Active" if is_active else "Inactive"
            with st.container(border=True):
                st.markdown(
                    f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px'>"
                    f"<div>"
                    f"<span style='color:{color};font-size:0.7rem;letter-spacing:0.1em;"
                    f"font-weight:600;text-transform:uppercase'>● {label}</span>"
                    f"<span style='color:var(--pra-text-muted);font-size:0.78rem;margin-left:12px'>"
                    f"{ad.get('days_running', 0):,} days · keyword: {_dh.escape(str(ad.get('keyword', '—')))}</span>"
                    f"</div>"
                    f"<div style='color:var(--pra-text-dim);font-size:0.7rem;font-variant-numeric:tabular-nums'>"
                    f"ID {ad.get('library_id', '')[:16]}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                if ad.get("ad_text"):
                    st.markdown(
                        f"<div style='color:var(--pra-text);font-size:0.85rem;line-height:1.55'>"
                        f"{_dh.escape((ad.get('ad_text') or '')[:400])}</div>",
                        unsafe_allow_html=True,
                    )
                if ad.get("landing_url"):
                    _lu = str(ad["landing_url"])
                    _lu_disp = _dh.escape(_lu[:80])
                    if _lu.lower().startswith(("http://", "https://")):
                        st.markdown(
                            f"<a href=\"{_dh.escape(_lu, quote=True)}\" target='_blank' rel='noopener noreferrer' "
                            f"style='color:var(--pra-accent);font-size:0.78rem;text-decoration:none;"
                            f"display:inline-block;margin-top:6px'>→ {_lu_disp}</a>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            f"<span style='color:var(--pra-text-dim);font-size:0.78rem;"
                            f"display:inline-block;margin-top:6px'>→ {_lu_disp}</span>",
                            unsafe_allow_html=True,
                        )
        if len(ads_sorted) > 8:
            st.caption(f"+ {len(ads_sorted) - 8} more variants in this brand cluster")


def _filter_rows(
    rows: list[dict], niches: list[str], min_days: int, active_only: bool,
    ph_only: bool, on_niche_only: bool, has_sales_only: bool, q: str,
    date_from=None, date_to=None,
    llm_hook: list[str] | None = None, llm_demo: list[str] | None = None,
    category: list[str] | None = None, sub_category: list[str] | None = None,
    location: list[str] | None = None,
    platforms: list[str] | None = None,
) -> list[dict]:
    out = []
    q_lower = q.lower().strip()
    _from_iso = date_from.isoformat() if date_from else None
    _to_iso = date_to.isoformat() if date_to else None
    for r in rows:
        if niches and r.get("niche") not in niches:
            continue
        if (r.get("days_running") or 0) < min_days:
            continue
        if active_only and not r.get("is_active"):
            continue
        if ph_only and r.get("geo_signal") != "ph-confident":
            continue
        if on_niche_only and r.get("niche_relevance") == "no_match":
            continue
        if has_sales_only and not (r.get("mp_sold") or r.get("mp_price")):
            continue
        if llm_hook and r.get("llm_hook_angle") not in llm_hook:
            continue
        if llm_demo and r.get("llm_target_demo") not in llm_demo:
            continue
        if category and r.get("category") not in category:
            continue
        if sub_category and r.get("sub_category") not in sub_category:
            continue
        if location and r.get("location") not in location:
            continue
        if platforms:
            row_plats = (r.get("platforms") or "").lower()
            if not any(p.lower() in row_plats for p in platforms):
                continue
        if _from_iso or _to_iso:
            sd = r.get("start_date") or ""
            if not sd:
                continue
            if _from_iso and sd < _from_iso:
                continue
            if _to_iso and sd > _to_iso:
                continue
        if q_lower:
            hay = " ".join(str(r.get(k, "") or "") for k in ("brand", "page_name", "ad_text", "landing_url")).lower()
            if q_lower not in hay:
                continue
        out.append(r)
    return out


def _df_from_rows(rows: list[dict], cols: list[str]) -> pd.DataFrame:
    return pd.DataFrame([{c: r.get(c) for c in cols} for r in rows])


# ---------- Sidebar ----------

with st.sidebar:
    # ---- Orbit brand header — actual brass logo image ----
    _sb_logo = ROOT / "static" / "orbit_logo.png"
    if _sb_logo.exists():
        _sbl1, _sbl2, _sbl3 = st.columns([1, 5, 1])
        with _sbl2:
            st.image(str(_sb_logo), width="stretch")
    st.markdown(
        "<div style='text-align:center;font-size:0.56rem;letter-spacing:0.18em;text-transform:uppercase;"
        "color:var(--pra-text-dim);margin-top:-4px;margin-bottom:6px'>Product Research Hunter</div>"
        "<div style='height:1px;background:var(--pra-border);margin:8px 4px 2px'></div>",
        unsafe_allow_html=True,
    )

    # ---- Premium nav (top-level Dashboard + collapsible section dropdowns) ----
    _current = st.session_state.get("current_page", "dashboard")
    st.markdown(
        "<style>"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_']{margin-top:-0.5rem !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button{"
        "justify-content:flex-start !important;text-align:left !important;background:transparent !important;"
        "border:none !important;border-left:2px solid transparent !important;border-radius:7px !important;"
        "padding:4px 12px !important;min-height:0 !important;height:auto !important;box-shadow:none !important;"
        "color:var(--pra-text-muted) !important;gap:10px !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button>div{width:100% !important;justify-content:flex-start !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button>div>span{justify-content:flex-start !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button *{color:inherit !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button p{text-align:left !important;"
        "font-size:0.85rem !important;font-weight:500 !important;margin:0 !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button:hover{"
        "background:rgba(212,175,55,0.07) !important;color:var(--pra-text) !important;}"
        ".pra-nav-sec{font-size:0.6rem;letter-spacing:0.14em;text-transform:uppercase;color:var(--pra-text-dim);"
        "font-weight:600;padding:12px 14px 2px;margin:8px 0 0 0;}"
        "section[data-testid='stSidebar'] .st-key-nav_btn_" + _current + " button{"
        "background:linear-gradient(90deg,var(--pra-accent-dim),transparent) !important;"
        "border-left:2px solid var(--pra-accent) !important;color:var(--pra-text-strong) !important;}</style>",
        unsafe_allow_html=True,
    )
    # ---- Sidebar polish: section spacing + brass markers + smooth hover ----
    st.markdown(
        "<style>"
        ".pra-nav-sec{color:var(--pra-text-muted) !important;font-weight:700 !important;"
        "letter-spacing:0.16em !important;padding:10px 14px 0 14px !important;margin:4px 0 15px 0 !important;"
        "display:flex !important;align-items:center !important;gap:8px !important;}"
        ".pra-nav-sec::before{content:'';width:5px;height:5px;border-radius:1px;"
        "background:var(--pra-accent);opacity:0.75;flex-shrink:0;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button{"
        "transition:background-color 0.13s ease,border-color 0.13s ease !important;}"
        "section[data-testid='stSidebar'] [class*='st-key-nav_btn_'] button:hover{"
        "border-left-color:rgba(212,175,55,0.4) !important;}"
        "</style>",
        unsafe_allow_html=True,
    )
    for _sec_name, _sec_items in _NAV_SECTIONS:
        # Single-item sections render as flat top-level buttons (no expander wrapper).
        # Always use type=secondary so we never collide with real-CTA styling (Scrape etc.).
        # Active state is signalled via a marker class that CSS :has() picks up.
        if _sec_name == "Overview":
            # Every Overview item renders as a flat top-level button (no section label).
            for _pid, _label, _icon in _sec_items:
                _is_active = (_pid == _current)
                _marker_cls = "pra-flat-nav-marker pra-flat-nav-active" if _is_active else "pra-flat-nav-marker"
                st.markdown(
                    f"<div class='{_marker_cls}' style='display:none'></div>",
                    unsafe_allow_html=True,
                )
                if st.button(
                    _label,
                    key=f"nav_btn_{_pid}",
                    icon=_NAV_MAT.get(_pid),
                    width="stretch",
                    help=_label,
                    type="secondary",
                ):
                    st.session_state.current_page = _pid
                    st.rerun()
            continue

        # Multi-item sections: visible label + every item always shown (Open Nav)
        st.markdown(f"<div class='pra-nav-sec'>{_sec_name}</div>", unsafe_allow_html=True)
        for _pid, _label, _icon in _sec_items:
            if st.button(
                _label,
                key=f"nav_btn_{_pid}",
                icon=_NAV_MAT.get(_pid),
                width="stretch",
                help=_label,
                type="secondary",
            ):
                st.session_state.current_page = _pid
                st.rerun()

    st.divider()

    # ---- Premium filters: clean borderless expanders + brass toggles ----
    st.markdown(
        "<style>"
        "section[data-testid='stSidebar'] [data-testid='stExpander']{border:none !important;"
        "background:transparent !important;box-shadow:none !important;}"
        "section[data-testid='stSidebar'] [data-testid='stExpander'] details{border:none !important;background:transparent !important;}"
        "section[data-testid='stSidebar'] [data-testid='stExpander'] summary{padding:7px 4px !important;"
        "background:transparent !important;border:none !important;}"
        "section[data-testid='stSidebar'] [data-testid='stExpander'] summary:hover{background:rgba(212,175,55,0.04) !important;border-radius:6px;}"
        "section[data-testid='stSidebar'] [data-testid='stExpander'] summary p{font-size:0.6rem !important;"
        "letter-spacing:0.16em !important;text-transform:uppercase !important;color:var(--pra-text-muted) !important;font-weight:700 !important;}"
        "section[data-testid='stSidebar'] [data-testid='stExpander'] summary::before{content:'';display:inline-block;"
        "width:5px;height:5px;border-radius:1px;background:var(--pra-accent);opacity:0.75;margin-right:9px;vertical-align:middle;}"
        "section[data-testid='stSidebar'] [data-testid='stExpanderDetails']{padding:6px 2px 10px !important;border:none !important;background:transparent !important;}"
        "section[data-testid='stSidebar'] [data-testid='stExpanderDetails']>div{border:none !important;}"
        "section[data-testid='stSidebar'] [data-baseweb='checkbox']{align-items:center !important;}"
        "</style>",
        unsafe_allow_html=True,
    )
    # ---- Run selector ----
    db.init_db()
    try:
        _cleaned = db.cleanup_empty_runs(min_age_minutes=30)
        if _cleaned:
            get_runs.clear()
    except Exception:
        pass
    runs = get_runs()
    if runs:
        from datetime import datetime as _dt_run
        def _fmt_run_label(r: dict) -> str:
            """Pretty short label: 'Run #3 · May 12 · 952 ads'"""
            rid = r["run_id"]
            ads = r.get("ads_count", r.get("total_ads", 0))
            when = r.get("started_at", "")
            try:
                d = _dt_run.fromisoformat(when)
                when_short = d.strftime("%b %d")
            except Exception:
                when_short = (when or "")[:10]
            return f"Run #{rid}  ·  {when_short}  ·  {ads:,} ads"

        labels = {_fmt_run_label(r): r["run_id"] for r in runs}

        # Wrap in expander (open by default) so it matches Filters / Saved searches dropdowns
        with st.expander("Active run", expanded=True):
            chosen = st.selectbox(
                "Active run", list(labels.keys()), index=0,
                label_visibility="collapsed",
            )
            active_run_id = labels[chosen]
    else:
        st.warning("No data yet. Run a scrape or `python main.py --import-xlsx`.")
        active_run_id = None

    # Apply any pending saved search BEFORE widgets render
    if "_apply_saved_search" in st.session_state:
        for _k, _v in st.session_state.pop("_apply_saved_search").items():
            st.session_state[_k] = _v

    # ---- Filters (collapsed by default — only Dashboard/FB Ads use them heavily) ----
    _filters_default_open = _current in ("dashboard", "fb_ads")
    with st.expander("Filters", expanded=_filters_default_open):
        all_niches = list(load_config().get("niches", {}).keys())
        if not all_niches:
            all_niches = ["capsule", "cream", "oil", "coffee"]
        niche_filter = st.multiselect(
            "Niche", all_niches, default=[], key="filter_niche",
            help="Restrict to specific niches. Empty = show all niches.",
        )
        min_days_filter = st.slider(
            "Min days running", 0, 1000, 30, step=5, key="filter_min_days",
            help="Drop ads younger than this many days — filters out untested launches",
        )
        active_only_filter = st.toggle(
            "Active only", value=True, key="filter_active_only",
            help="Hide ads that have stopped running",
        )
        ph_only_filter = st.toggle(
            "PH-confident only", value=True, key="filter_ph_only",
            help="Hide ads without strong PH-targeting signals.",
        )
        on_niche_only_filter = st.toggle(
            "In-niche text only", value=True, key="filter_on_niche",
            help="Hide ads where the text doesn't contain niche-relevant keywords.",
        )
        has_sales_filter = st.toggle(
            "Has marketplace data only", value=False, key="filter_has_sales",
            help="Show only ads enriched with Shopee/Lazada sales data.",
        )
        search_q = st.text_input(
            "Search brand / ad text", value="", key="filter_search",
            help="Substring match across brand name, page name, ad text, and landing URL",
        )

        # ---- Tag filters (Category / Sub-category / Location / LLM) ----
        # Grouped in a nested expander to keep the Filters expander uncluttered.
        # All filters render regardless of expander state — state still flows.
        import categorization as _cat_mod
        import location_detector as _loc_mod

        # Pretty-format helper for sub-category labels: "anti-aging" -> "Anti Aging"
        def _pretty_subcat(s: str) -> str:
            if not s:
                return ""
            return s.replace("-", " ").replace("_", " ").title()

        with st.expander("More tags · Category · Location · LLM", expanded=False):
            _cat_filter = st.multiselect(
                "Category",
                _cat_mod.all_categories(),
                default=[], key="filter_category",
                help="Top-level intent: Health & Wellness, Beauty & Personal Care, "
                     "Food & Beverage, Wellness & Lifestyle.",
            )
            if _cat_filter:
                _subcat_options = sorted({
                    s for c in _cat_filter
                    for s in _cat_mod.all_sub_categories(c)
                })
            else:
                _subcat_options = _cat_mod.all_sub_categories()
            _subcat_filter = st.multiselect(
                "Sub-category",
                _subcat_options,
                default=[], key="filter_subcategory",
                format_func=_pretty_subcat,
                help="Specific angle within the chosen category(s). "
                     "Picks narrow when you select a Category above.",
            )
            _location_filter = st.multiselect(
                "Location",
                _loc_mod.ALL_LOCATIONS,
                default=[], key="filter_location",
                help="PH city/region detected from ad text + landing URL. "
                     "'PH-wide' = nationwide ad with no specific locale.",
            )
            # Phase 17.7 — Platform filter (FB / IG / Audience Network / Messenger)
            _platform_filter = st.multiselect(
                "Platform",
                ["Facebook", "Instagram", "Audience Network", "Messenger"],
                default=[], key="filter_platform",
                help="Filter by where the ad runs. Meta Library captures cross-platform ads; "
                     "many ads run on FB + IG together. (Requires a fresh scrape — older "
                     "ads may have empty platform data.)",
            )

            st.divider()
            # ---- LLM tags filter (Claude classifier required) ----
            _llm_hook_filter = st.multiselect(
                "LLM hook angle",
                ["scarcity", "social-proof", "before-after", "problem-solution",
                 "claim", "testimonial", "demo", "curiosity", "other"],
                default=[], key="filter_llm_hook",
                help="Filter by Claude-classified hook angle. Run classifier first.",
            )
            _llm_demo_filter = st.multiselect(
                "LLM target demo",
                ["mom", "gym", "boomer", "gen-z", "working-pro", "beauty",
                 "men", "women", "other"],
                default=[], key="filter_llm_demo",
                help="Filter by Claude-classified target demographic.",
            )

        # ---- Date range filter (ad start_date) ----
        from datetime import date as _date_t, timedelta as _td_t
        _use_date_range = st.toggle(
            "Filter by ad start date", value=False,
            help="Restrict to ads whose Meta 'Started running on' date falls in a window",
        )
        if _use_date_range:
            _date_from = st.date_input(
                "From", value=_date_t.today() - _td_t(days=180),
                help="Show ads that started running on or after this date", key="date_from",
            )
            _date_to = st.date_input(
                "To", value=_date_t.today(),
                help="Show ads that started running on or before this date", key="date_to",
            )
        else:
            _date_from = None
            _date_to = None

    # ---- Saved searches (collapsed by default) ----
    _saved = load_saved_searches()
    with st.expander("Saved searches", expanded=False):
        if not _saved:
            st.caption("No saved searches yet. Configure Filters above, then save.")
        for _name in list(_saved.keys()):
            _del_confirm_key = f"_confirm_del_search_{_name}"
            _is_pending_del = st.session_state.get(_del_confirm_key, False)

            # Row: load button (stays stable) + ✕ delete trigger
            _sscol1, _sscol2 = st.columns([5, 1])
            if _sscol1.button(
                _name, key=f"load_search_{_name}",
                help=f"Apply '{_name}'", width="stretch",
            ):
                st.session_state["_apply_saved_search"] = _saved[_name]
                st.rerun()
            if _sscol2.button(
                "✕", key=f"del_search_{_name}",
                help=f"Delete '{_name}'",
            ):
                st.session_state[_del_confirm_key] = True
                st.rerun()

            # Inline confirmation prompt below the row (only when armed)
            if _is_pending_del:
                st.markdown(
                    f"<div style='color:var(--pra-danger);font-size:0.75rem;"
                    f"margin:2px 0 4px 4px;letter-spacing:0.02em'>"
                    f"Delete '<strong>{_name}</strong>'?</div>",
                    unsafe_allow_html=True,
                )
                _ccol1, _ccol2 = st.columns(2)
                if _ccol1.button("Cancel", key=f"cancel_del_search_{_name}",
                                 width="stretch", type="secondary"):
                    st.session_state.pop(_del_confirm_key, None)
                    st.rerun()
                if _ccol2.button("Yes, delete", key=f"confirm_del_search_{_name}",
                                 width="stretch", type="primary"):
                    del _saved[_name]
                    save_saved_searches(_saved)
                    st.session_state.pop(_del_confirm_key, None)
                    st.toast(f"✕ Deleted '{_name}'", icon=None)
                    st.rerun()
        st.divider()
        _new_search_name = st.text_input(
            "Save current as…", placeholder="e.g. Cream 90+ days",
            key="save_search_name", label_visibility="collapsed",
        )
        if st.button("Save current filters", key="save_search_btn",
                     width="stretch", type="secondary",
                     disabled=not _new_search_name.strip()):
            current = {
                "filter_niche": niche_filter,
                "filter_min_days": min_days_filter,
                "filter_active_only": active_only_filter,
                "filter_ph_only": ph_only_filter,
                "filter_on_niche": on_niche_only_filter,
                "filter_has_sales": has_sales_filter,
                "filter_search": search_q,
            }
            saved_dict = load_saved_searches()
            saved_dict[_new_search_name.strip()] = current
            save_saved_searches(saved_dict)
            st.toast(f"✓ Saved as '{_new_search_name.strip()}'", icon=None)
            st.rerun()

    st.divider()

    # ---- Actions ----
    # In cloud mode, the Scrape button is HIDDEN because Playwright can't run
    # there. Only Refresh remains. Local-only scrape stays as primary CTA on
    # the desktop install.
    if IS_CLOUD:
        if st.button("↻ Refresh", width="stretch", type="primary",
                     help="Reload data from DB (cloud view-only mode)"):
            get_runs.clear()
            get_ads.clear()
            get_stats.clear()
            _clear_brand_caches()
            st.rerun()
        st.caption(
            "📡 Cloud view-only · Scrape from your local desktop install, then "
            "commit the updated `orbit.db` to GitHub to refresh this view."
        )
    else:
        _act1, _act2 = st.columns(2)
        if _act1.button("⟳ Scrape", width="stretch", type="primary",
                        help="Launch a fresh scrape in background"):
            subprocess.Popen(
                [sys.executable, "-u", str(ROOT / "main.py")],
                cwd=str(ROOT),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
            )
            with st.status("Scrape launched", state="running", expanded=True):
                st.write("Subprocess started — Chromium headless")
                st.write("ETA ~15 min · tail `logs/agent.log`")
                st.write("Click Refresh below when done")

        if _act2.button("↻ Refresh", width="stretch", help="Reload data from DB"):
            get_runs.clear()
            get_ads.clear()
            get_stats.clear()
            _clear_brand_caches()
            st.rerun()

    st.divider()

    # ---- Recent activity (today) — collapsed by default ----
    _today_activity = get_activity_today()
    if _today_activity:
        with st.expander(f"Today's activity ({len(_today_activity)})", expanded=False):
            _ACTION_LABELS = {
                "starred":            ("★", "starred"),
                "unstarred":          ("☆", "unstarred"),
                "added_to_testing":   ("◈", "→ testing"),
                "deleted_testing":    ("✕", "deleted testing"),
                "generated_ad_copy":  ("✎", "generated copy"),
                "saved_search":       ("⊞", "saved search"),
            }
            for _act in _today_activity[:8]:
                _icon, _label = _ACTION_LABELS.get(
                    _act.get("action", ""),
                    ("●", _act.get("action", "")),
                )
                _ts = (_act.get("ts") or "")[11:16]
                _target = _act.get("target", "")
                # Full text in title attr so hover reveals long names
                _full_title = f"{_label} {_target} at {_ts}".strip()
                st.markdown(
                    f"<div title='{_full_title}' style='display:flex;align-items:center;"
                    f"gap:6px;padding:4px 0;font-size:0.76rem;"
                    f"border-bottom:1px solid var(--pra-border);line-height:1.4'>"
                    f"<span style='color:var(--pra-accent);flex-shrink:0'>{_icon}</span>"
                    f"<div style='color:var(--pra-text);overflow:hidden;text-overflow:ellipsis;"
                    f"white-space:nowrap;flex:1;min-width:0'>"
                    f"<span style='color:var(--pra-text-muted)'>{_label}</span> "
                    f"<span style='font-weight:500'>{_target}</span>"
                    f"</div>"
                    f"<span style='color:var(--pra-text-dim);font-variant-numeric:tabular-nums;"
                    f"font-size:0.72rem;flex-shrink:0'>{_ts}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    s = get_stats()

    # ---- Footer: DB stats (left) + theme toggle (right) ----
    _is_dark = st.session_state.theme_mode == "dark"
    _theme_icon = "☀" if _is_dark else "☾"
    _theme_help = "Switch to light mode" if _is_dark else "Switch to dark mode"
    # Layout: stats on the left, two equal-width icon buttons on the right.
    # The two icon buttons match the theme-toggle style — circular, no label wrap.
    _foot_l, _foot_sh, _foot_r = st.columns([5, 1, 1])
    with _foot_l:
        st.markdown(
            f"<div style='color:var(--pra-text-muted);font-size:0.7rem;line-height:1.5;"
            f"padding-top:6px'>"
            f"{s['total_runs']} runs · {s['unique_library_ids']:,} ads · "
            f"{s['unique_brands']:,} brands"
            f"</div>",
            unsafe_allow_html=True,
        )
    with _foot_sh:
        # Icon-only "?" — also clickable target for the JS ? key interceptor
        if st.button("?", key="shortcuts_trigger_btn",
                     help="Keyboard shortcuts (press ?)",
                     width="stretch"):
            st.session_state.show_shortcuts = True
            st.rerun()
    with _foot_r:
        if st.button(_theme_icon, key="theme_toggle_btn", help=_theme_help, width="stretch"):
            st.session_state.theme_mode = "light" if _is_dark else "dark"
            st.rerun()


# ---------- Main ----------

if not active_run_id:
    st.title("Product Research Agent")
    st.info("No runs in the database yet.")
    st.code("python main.py --import-xlsx   # backfill from existing xlsx files\npython main.py                # run a fresh scrape")
    st.stop()


rows = get_ads(active_run_id)
filtered = _filter_rows(
    rows, niche_filter, min_days_filter, active_only_filter,
    ph_only_filter, on_niche_only_filter, has_sales_filter, search_q,
    date_from=_date_from, date_to=_date_to,
    llm_hook=st.session_state.get("filter_llm_hook") or None,
    llm_demo=st.session_state.get("filter_llm_demo") or None,
    category=st.session_state.get("filter_category") or None,
    sub_category=st.session_state.get("filter_subcategory") or None,
    location=st.session_state.get("filter_location") or None,
    platforms=st.session_state.get("filter_platform") or None,
)

# --- Header bar ---
_run_meta = next((r for r in runs if r["run_id"] == active_run_id), None)
_run_when = (_run_meta or {}).get("started_at", "")[:16].replace("T", " ")
_total_db = s.get("unique_library_ids", 0)
_total_brands = s.get("unique_brands", 0)
# Time-aware greeting + page-aware header title (UX polish)
_hdr_hour = datetime.now().hour
if IS_CLOUD:
    _hdr_hour = (_hdr_hour + 8) % 24  # cloud runs UTC; PH is UTC+8
_hdr_greet = ("Good morning" if _hdr_hour < 12 else "Good afternoon" if _hdr_hour < 18 else "Good evening")
_hdr_cp = st.session_state.get("current_page", "dashboard")
if _hdr_cp not in _PAGE_TITLES:
    _hdr_cp = "dashboard"
_hdr_main = _hdr_greet if _hdr_cp == "dashboard" else _PAGE_TITLES.get(_hdr_cp, "Orbit")
_header_l, _header_inbox = st.columns([9, 1])
with _header_l:
    st.markdown(
        f"""
        <div class="pra-header">
            <div>
                <div style="font-family:Fraunces,Georgia,serif;font-size:1.3rem;font-weight:600;color:var(--pra-text-strong);line-height:1.1">{_hdr_main}</div>
                <div style="color:var(--pra-text-muted);font-size:0.76rem;margin-top:3px;font-variant-numeric:tabular-nums">Run <strong style="color:var(--pra-text)">#{active_run_id}</strong> · {_run_when} · <strong style="color:var(--pra-text)">{_total_db:,}</strong> ads · <strong style="color:var(--pra-text)">{_total_brands:,}</strong> brands</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
with _header_inbox:
    # Inbox bell with unread badge — count actual unread runs since last visit.
    _last_seen = inbox_last_seen()
    if not _last_seen:
        _unread_count = len(runs)  # never visited inbox → all runs are unread
    else:
        _unread_count = sum(
            1 for r in runs
            if r.get("started_at", "") > _last_seen
        )
    _bell_label = f"🔔 {_unread_count}" if _unread_count else "🔔"
    if st.button(_bell_label, key="inbox_trigger_btn",
                 help="Notification inbox",
                 width="stretch"):
        st.session_state.show_inbox = True
        st.rerun()
# Hidden command-palette trigger — Ctrl/Cmd+K still works, no visible button
if st.button("⌘ K  Search", key="palette_trigger_btn",
             help="Quick navigation (Ctrl/Cmd+K)"):
    st.session_state.show_palette = True
    st.rerun()
st.markdown(
    "<style>.st-key-palette_trigger_btn{display:none !important;}</style>",
    unsafe_allow_html=True,
)

# Show at most ONE dialog at a time (Streamlit only supports one @st.dialog instance).
# Priority order: palette > inbox > shortcuts. If multiple flags are set, the others
# are cleared so they don't fire on the next rerun.
_dialog_priority = [
    ("show_palette",   lambda: _render_command_palette(rows)),
    ("show_inbox",     lambda: _render_inbox(active_run_id, runs, rows)),
    ("show_shortcuts", lambda: _render_shortcuts_overlay()),
]
_active_dialog = None
for _flag, _renderer in _dialog_priority:
    if st.session_state.get(_flag):
        if _active_dialog is None:
            _active_dialog = (_flag, _renderer)
        else:
            # Higher-priority dialog already won — clear this one's flag
            st.session_state[_flag] = False
if _active_dialog is not None:
    _active_dialog[1]()

# JS interceptor: Ctrl/Cmd+K opens palette · ? opens shortcuts overlay
st.markdown(
    """
    <script>
    (function() {
        if (window.__pra_palette_kbd_installed) return;
        window.__pra_palette_kbd_installed = true;
        const findAndClick = (matcher) => {
            const btns = window.parent.document.querySelectorAll('button');
            for (const btn of btns) {
                if (matcher(btn)) { btn.click(); return true; }
            }
            return false;
        };
        const handler = (e) => {
            const target = e.target;
            const isInput = target && (target.tagName === 'INPUT' || target.tagName === 'TEXTAREA' || target.isContentEditable);
            // Ctrl/Cmd+K — palette
            if ((e.ctrlKey || e.metaKey) && (e.key === 'k' || e.key === 'K')) {
                e.preventDefault();
                findAndClick(b => b.textContent && b.textContent.includes('⌘ K'));
                return;
            }
            // ? — shortcuts overlay (only when not typing in an input).
            // The trigger is the small "?" icon button in the sidebar footer.
            // We match buttons whose text is exactly "?" AND whose aria-label or
            // help text contains "shortcuts" to avoid false positives.
            if (!isInput && e.key === '?' && !e.ctrlKey && !e.metaKey) {
                e.preventDefault();
                findAndClick(b => {
                    const txt = b.textContent && b.textContent.trim();
                    if (txt !== '?') return false;
                    const title = (b.getAttribute('title') || '').toLowerCase();
                    const aria  = (b.getAttribute('aria-label') || '').toLowerCase();
                    return title.includes('shortcut') || aria.includes('shortcut');
                });
                return;
            }
        };
        window.addEventListener('keydown', handler);
        window.parent.addEventListener('keydown', handler);
    })();
    </script>
    """,
    unsafe_allow_html=True,
)

# (Command palette + inbox + shortcuts dialogs are rendered together below, mutually
#  exclusive — Streamlit only supports one @st.dialog at a time.)

# --- Onboarding banner (first launch only) ---
# Phase 18.6 — First-launch onboarding wizard (multi-step st.dialog)
if not onboarding_done() and not st.session_state.get("_onb_dismissed_banner") and not _total_db:
    _onb_l, _onb_r1, _onb_r2 = st.columns([4, 1, 1])
    with _onb_l:
        st.markdown(
            "<div style='background:linear-gradient(135deg,rgba(212,175,55,0.12) 0%,"
            "rgba(212,175,55,0.04) 100%);border:1px solid var(--pra-accent);"
            "border-radius:6px;padding:12px 18px;margin-bottom:12px'>"
            "<div style='color:var(--pra-accent);font-size:0.72rem;letter-spacing:0.14em;"
            "font-weight:600;text-transform:uppercase;margin-bottom:4px'>"
            "👋 Welcome to Orbit</div>"
            "<div style='color:var(--pra-text);font-size:0.88rem;line-height:1.55'>"
            "First time here? Run the quick setup wizard — pick your niches, "
            "configure AI, set up the daily scheduler."
            "</div>"
            "</div>",
            unsafe_allow_html=True,
        )
    with _onb_r1:
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        if st.button("Start setup", key="onb_open_wizard", type="primary",
                     width="stretch"):
            st.session_state["show_onboarding"] = True
            st.session_state["_onb_step"] = 1
            st.rerun()
    with _onb_r2:
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        if st.button("Skip", key="onb_skip_banner", width="stretch"):
            st.session_state["_onb_dismissed_banner"] = True
            mark_onboarding_done()
            st.rerun()

if st.session_state.get("show_onboarding"):
    _render_onboarding_wizard()

# --- Resolve current page (drives sidebar nav + conditional content below) ---
current_page = st.session_state.get("current_page", "dashboard")
if current_page not in _PAGE_TITLES:
    current_page = "dashboard"
    st.session_state["current_page"] = "dashboard"

# Status line under header (always visible — quick context)
brands_in_view = {r.get("brand") or r.get("page_name") for r in filtered}
active_count = sum(1 for r in filtered if r.get("is_active"))
_page_title = _PAGE_TITLES.get(current_page, "Dashboard")
st.markdown(
    f"<div style='display:flex;align-items:center;justify-content:space-between;"
    f"margin:0 4px 18px 4px'>"
    f"<div style='font-size:1.25rem;font-weight:600;color:var(--pra-text);"
    f"letter-spacing:-0.01em'>{_page_title}</div>"
    f"<div style='color:var(--pra-text-muted);font-size:0.78rem;letter-spacing:0.01em'>"
    f"<strong style='color:var(--pra-text)'>{len(filtered):,}</strong> of "
    f"{len(rows):,} ads · "
    f"<strong style='color:var(--pra-text)'>{len(brands_in_view):,}</strong> brands · "
    f"<strong style='color:var(--pra-text)'>{active_count:,}</strong> active"
    f"</div>"
    f"</div>",
    unsafe_allow_html=True,
)

# Phase 12.5 — first-visit hint card (skip dashboard — greeting + hunt brief cover it)
if current_page != "dashboard":
    render_page_hint(current_page)

# Dashboard-only: compact metrics row
if current_page == "dashboard":
    import html as _hl
    _brief = _research_brief(active_run_id, runs, rows)

    # ---- Top winner for the hero (lightweight aggregate) ----
    _hero_agg: dict[str, dict] = {}
    for r in rows:
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        if not r.get("is_active"):
            continue
        _b = (r.get("brand") or r.get("page_name") or "").strip()
        if not _b:
            continue
        _e = _hero_agg.setdefault(_b, {"brand": _b, "niche": r.get("niche") or "—",
                                       "score": 0.0, "days": 0, "ads": 0})
        _e["score"] += float(r.get("score_normalized") or 0)
        _e["days"] = max(_e["days"], r.get("days_running") or 0)
        _e["ads"] += 1
    _winner = max(_hero_agg.values(), key=lambda x: x["score"]) if _hero_agg else None
    _ph_conf = sum(1 for r in rows if r.get("geo_signal") == "ph-confident")

    # ---- Hunt brief (one-line, guided) ----
    _hb_parts = [f"🎯 <strong style='color:var(--pra-accent)'>{len(_hero_agg)} winning brands</strong> ngayon"]
    if _brief.get("hot_niche"):
        _hb_parts.append(f"<strong style='color:var(--pra-accent)'>{_hl.escape(str(_brief['hot_niche']).title())}</strong> niche leading")
    if _winner:
        _hb_parts.append(f"top pick: <strong style='color:var(--pra-accent)'>{_hl.escape(_winner['brand'])}</strong>")
    _hb = " &middot; ".join(_hb_parts)
    st.markdown(
        "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
        "border-left:2px solid var(--pra-accent);border-radius:9px;padding:12px 16px;"
        "margin-bottom:14px;font-size:0.88rem;color:var(--pra-text)'>" + _hb + "</div>",
        unsafe_allow_html=True,
    )

    # ---- Quick actions row ----
    _qa_items = [
        ("🔍 Browse winners", "fb_ads"),
        ("🏆 Bestsellers", "bestsellers"),
        ("◎ Competitors", "competitors"),
        ("⚙ Find suppliers", "supplier"),
    ]
    _qa_cols = st.columns(len(_qa_items))
    for _qi, (_qlabel, _qpage) in enumerate(_qa_items):
        with _qa_cols[_qi]:
            if st.button(_qlabel, key=f"qa_{_qpage}", width="stretch"):
                st.session_state.current_page = _qpage
                st.rerun()
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ===== BENTO ROW 1: hero winner (2/3) + stat stack (1/3) =====
    _h_left, _h_right = st.columns([2, 1], gap="medium")
    with _h_left:
        if _winner:
            st.markdown(
                "<div style='background:linear-gradient(135deg,rgba(212,175,55,0.13) 0%,"
                "rgba(212,175,55,0.02) 100%);border:1px solid rgba(212,175,55,0.35);"
                "border-radius:12px;padding:24px 28px;height:100%'>"
                "<div style='color:var(--pra-accent);font-size:0.68rem;letter-spacing:0.16em;"
                "font-weight:700;text-transform:uppercase;margin-bottom:12px'>◆ Today's top winner</div>"
                f"<div style='font-family:Fraunces,Georgia,serif;font-size:2.2rem;font-weight:600;"
                f"color:var(--pra-text-strong);line-height:1.05'>{_hl.escape(_winner['brand'])}</div>"
                f"<div style='color:var(--pra-text-muted);font-size:0.9rem;margin-top:9px'>"
                f"{_hl.escape(str(_winner['niche']).title())} &middot; {_winner['days']} days live "
                f"&middot; {_winner['ads']} active ads &middot; score {int(_winner['score'])}</div>"
                "<div style='margin-top:18px;height:6px;background:var(--pra-subtle-bg);border-radius:4px;"
                "overflow:hidden'><div style='width:100%;height:100%;background:linear-gradient(90deg,"
                "#D4AF37,#E6CC73)'></div></div>"
                "<div style='color:var(--pra-text-dim);font-size:0.74rem;margin-top:11px'>"
                "Longest-running PH winner this run — proven demand, modelong-mauna.</div>"
                "</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                "border-radius:12px;padding:24px 28px;height:100%'>"
                "<div style='color:var(--pra-accent);font-size:0.68rem;letter-spacing:0.16em;"
                "font-weight:700;text-transform:uppercase;margin-bottom:10px'>◆ Today's top winner</div>"
                "<div style='color:var(--pra-text-muted);font-size:0.92rem'>Walang qualifying winner pa "
                "— mag-scrape para mapuno ang top brand ngayon.</div></div>",
                unsafe_allow_html=True,
            )
    with _h_right:
        _stat_cards = [
            ("Active winners", f"{len(_hero_agg):,}", "qualifying PH brands"),
            ("PH-confident ads", f"{_ph_conf:,}", "geo-verified"),
            ("Top niche", str(_brief["hot_niche"]).title(),
             f"{_brief['hot_niche_brands']} brands" if _brief["hot_niche_brands"] else "—"),
        ]
        _stat_html = ""
        for _lbl, _val, _sub in _stat_cards:
            _stat_html += (
                "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                "border-radius:10px;padding:13px 17px;margin-bottom:11px'>"
                f"<div style='color:var(--pra-text-muted);font-size:0.65rem;letter-spacing:0.1em;"
                f"text-transform:uppercase'>{_lbl}</div>"
                f"<div style='font-family:Fraunces,Georgia,serif;font-size:1.55rem;font-weight:600;"
                f"color:var(--pra-text-strong);line-height:1.1;margin-top:3px'>{_hl.escape(str(_val))}</div>"
                f"<div style='color:var(--pra-accent);font-size:0.7rem;margin-top:2px'>{_hl.escape(str(_sub))}</div>"
                "</div>"
            )
        st.markdown(_stat_html, unsafe_allow_html=True)
    st.markdown("<div style='margin-bottom:14px'></div>", unsafe_allow_html=True)


# --- Dashboard page: cross-source overview ---
if current_page == "dashboard":
    # ---- Slim section label (bento) ----
    st.markdown(
        "<div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.16em;"
        "font-weight:700;text-transform:uppercase;margin:8px 0 12px 2px'>All-sources overview "
        "<span style='color:var(--pra-text-dim);font-weight:400;letter-spacing:0;text-transform:none'>"
        "&middot; counts from every feed in one place</span></div>",
        unsafe_allow_html=True,
    )

    # ---- Source-count cards ----
    _all_n_fb = len(rows)
    _all_n_active = sum(1 for r in rows if r.get("is_active"))
    _all_n_shopee_lp = sum(
        1 for r in rows
        if any(kw in (r.get("landing_url") or "").lower() for kw in ("shopee.ph", "shp.ee"))
    )
    _all_n_lazada_lp = sum(
        1 for r in rows
        if "lazada." in (r.get("landing_url") or "").lower()
    )
    _all_n_shopee_enriched = sum(
        1 for r in rows
        if (r.get("mp_source") or "").lower() == "shopee" and r.get("mp_enriched_at")
    )
    _all_n_lazada_enriched = sum(
        1 for r in rows
        if (r.get("mp_source") or "").lower() == "lazada" and r.get("mp_enriched_at")
    )
    _all_n_tiktok = db.stats().get("tiktok_ads", 0)

    ac1, ac2, ac3, ac4 = st.columns(4)
    ac1.metric("FB Ads Library", f"{_all_n_fb:,}", delta=f"{_all_n_active:,} active")
    ac2.metric("Shopee ads", f"{_all_n_shopee_lp:,}", delta=f"{_all_n_shopee_enriched:,} enriched")
    ac3.metric("Lazada ads", f"{_all_n_lazada_lp:,}", delta=f"{_all_n_lazada_enriched:,} enriched")
    ac4.metric("TikTok ads", f"{_all_n_tiktok:,}", delta="Creative Center")

    st.divider()

    # ---- Source distribution mini-viz ----
    _src_counts = [
        ("FB Ads Library", _all_n_fb, "#D4AF37"),
        ("→ Shopee", _all_n_shopee_lp, "#E6CC73"),
        ("→ Lazada", _all_n_lazada_lp, "#8FC2B4"),
        ("TikTok", _all_n_tiktok, "#E0909F"),
    ]
    _all_max = max((c for _, c, _ in _src_counts), default=1) or 1

    col_dist, col_insights = st.columns([1, 1])
    with col_dist:
        _dist_html = (
            "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);border-radius:6px;"
            "padding:18px 22px'>"
            "<div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.16em;"
            "font-weight:700;text-transform:uppercase;margin-bottom:14px'>Source distribution</div>"
        )
        for label, cnt, color in _src_counts:
            pct = int((cnt / _all_max) * 100) if _all_max else 0
            _dist_html += (
                f"<div style='display:flex;align-items:center;margin-bottom:10px'>"
                f"<div style='width:130px;color:var(--pra-text);font-size:0.82rem'>{label}</div>"
                f"<div style='flex:1;background:var(--pra-border);border-radius:2px;height:6px;"
                f"margin:0 12px;overflow:hidden'>"
                f"<div style='background:{color};width:{pct}%;height:100%'></div></div>"
                f"<div style='width:50px;text-align:right;color:var(--pra-text);font-size:0.82rem;"
                f"font-variant-numeric:tabular-nums'>{cnt:,}</div>"
                f"</div>"
            )
        _dist_html += "</div>"
        st.markdown(_dist_html, unsafe_allow_html=True)

    # ---- Cross-source insights ----
    with col_insights:
        _fb_brands = {
            (r.get("brand") or r.get("page_name") or "").strip()
            for r in rows if r.get("is_active")
        }
        _fb_brands.discard("")
        _tt_advertisers = set()
        try:
            _tt_run = db.latest_tiktok_run_id()
            if _tt_run:
                for r in db.tiktok_ads_for_run(_tt_run):
                    a = (r.get("advertiser") or "").strip()
                    if a:
                        _tt_advertisers.add(a)
        except Exception:
            pass

        # Fuzzy overlap (case-insensitive substring match)
        overlap_fb_tt = 0
        for fb in _fb_brands:
            fb_low = fb.lower()
            for tt in _tt_advertisers:
                tt_low = tt.lower()
                if fb_low in tt_low or tt_low in fb_low:
                    overlap_fb_tt += 1
                    break

        _ph_pct = round(100 * sum(
            1 for r in rows if r.get("geo_signal") == "ph-confident"
        ) / max(_all_n_fb, 1))
        _shopee_pct = round(100 * _all_n_shopee_lp / max(_all_n_fb, 1))
        _lazada_pct = round(100 * _all_n_lazada_lp / max(_all_n_fb, 1))

        st.markdown(
            "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);border-radius:6px;"
            "padding:18px 22px;height:100%'>"
            "<div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.16em;"
            "font-weight:700;text-transform:uppercase;margin-bottom:14px'>Cross-source intelligence</div>"
            "<ul style='color:var(--pra-text);font-size:0.85rem;line-height:1.75;"
            "padding-left:18px;margin:0'>"
            f"<li><strong style='color:#D4AF37'>{_ph_pct}%</strong> of FB ads are PH-confident</li>"
            f"<li><strong style='color:#E6CC73'>{_shopee_pct}%</strong> of FB ads link to Shopee</li>"
            f"<li><strong style='color:#8FC2B4'>{_lazada_pct}%</strong> of FB ads link to Lazada</li>"
            f"<li><strong style='color:#E0909F'>{overlap_fb_tt}</strong> brands appear on both FB &amp; TikTok</li>"
            f"<li><strong style='color:#7CC4A0'>{len(_fb_brands):,}</strong> unique active brands across FB feed</li>"
            "</ul>"
            "</div>",
            unsafe_allow_html=True,
        )

    st.divider()

    # ---- Combined top brands table (all sources) ----
    st.markdown(
        "<div style='display:flex;align-items:baseline;justify-content:space-between;"
        "flex-wrap:wrap;gap:8px;margin:8px 0 12px 2px'>"
        "<div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.16em;"
        "font-weight:700;text-transform:uppercase'>Top brands across all sources</div>"
        f"<div style='color:var(--pra-text-dim);font-size:0.74rem'>"
        f"<strong style='color:var(--pra-success)'>{_brief['new_winners_count']}</strong> new 30-day winners "
        f"&middot; <strong style='color:var(--pra-danger)'>{_brief['retired_count']}</strong> stopped "
        f"&middot; updated {_brief['scrape_freshness']}</div>"
        "</div>",
        unsafe_allow_html=True,
    )

    _all_brand_agg: dict[str, dict] = {}
    for r in rows:
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        if not r.get("is_active"):
            continue
        b = (r.get("brand") or r.get("page_name") or "").strip()
        if not b:
            continue
        entry = _all_brand_agg.setdefault(b, {
            "brand": b, "niche": r.get("niche") or "—",
            "fb_ads": 0, "shopee_ads": 0, "lazada_ads": 0, "tiktok_ads": 0,
            "score": 0.0, "max_days": 0, "sample_landing": "",
        })
        entry["fb_ads"] += 1
        entry["score"] += float(r.get("score_normalized") or 0)
        entry["max_days"] = max(entry["max_days"], r.get("days_running") or 0)
        url_low = (r.get("landing_url") or "").lower()
        if "shopee.ph" in url_low or "shp.ee" in url_low:
            entry["shopee_ads"] += 1
        elif "lazada." in url_low:
            entry["lazada_ads"] += 1
        if not entry["sample_landing"] and r.get("landing_url"):
            entry["sample_landing"] = r["landing_url"]

    # Layer in TikTok presence (fuzzy match)
    try:
        for ad in (db.tiktok_ads_for_run(db.latest_tiktok_run_id() or 0) or []):
            adv = (ad.get("advertiser") or "").strip().lower()
            if not adv:
                continue
            for b, entry in _all_brand_agg.items():
                if adv in b.lower() or b.lower() in adv:
                    entry["tiktok_ads"] += 1
                    break
    except Exception:
        pass

    if not _all_brand_agg:
        st.info("No qualifying brands in this run to aggregate yet.")
    else:
        _ab_rows = sorted(_all_brand_agg.values(), key=lambda x: -x["score"])[:30]
        for b in _ab_rows:
            b["score"] = round(b["score"], 1)
        df_all = pd.DataFrame(_ab_rows)
        max_score = float(df_all["score"].max() or 1)

        df_all["sources"] = df_all.apply(
            lambda r: " ".join(s for s, c in (
                ("FB", r["fb_ads"]), ("SH", r["shopee_ads"]),
                ("LZ", r["lazada_ads"]), ("TT", r["tiktok_ads"]),
            ) if c > 0),
            axis=1,
        )

        # ---- Custom HTML table (theme-aware) ----
        # Streamlit's st.dataframe uses a canvas renderer that ignores CSS vars; this
        # HTML version properly switches between light and dark mode.
        import html as _html
        import re as _re_lp
        _rows_html = []
        for _row in _ab_rows:
            _score_pct = int((_row["score"] / max_score) * 100) if max_score else 0
            _landing = _row.get("sample_landing", "") or ""
            _dom_m = _re_lp.match(r"^https?://(?:www\.)?(.{1,55})", _landing)
            _landing_display = _dom_m.group(1) if _dom_m else "—"
            _landing_html = (
                f'<a class="pra-landing" href="{_html.escape(_landing)}" target="_blank">'
                f'{_html.escape(_landing_display)}</a>'
                if _landing else "—"
            )
            _rows_html.append(
                "<tr>"
                f'<td class="brand-cell">{_html.escape(_row["brand"])}</td>'
                f'<td>{_html.escape((_row.get("niche") or "—"))}</td>'
                f'<td class="sources-cell">{_html.escape(_row.get("sources","—"))}</td>'
                f'<td class="num">{_row["fb_ads"]}</td>'
                f'<td class="num">{_row["shopee_ads"]}</td>'
                f'<td class="num">{_row["lazada_ads"]}</td>'
                f'<td class="num">{_row["tiktok_ads"]}</td>'
                f'<td class="num">{_row["max_days"]:,} d</td>'
                f'<td>'
                f'<div class="pra-score-cell">'
                f'<div class="pra-bar-track">'
                f'<div class="pra-bar-fill" style="width:{_score_pct}%"></div>'
                f'</div>'
                f'<span class="pra-score-num">{int(_row["score"]):,}</span>'
                f'</div>'
                f'</td>'
                f'<td>{_landing_html}</td>'
                "</tr>"
            )

        _table_html = (
            "<div style='max-height:560px;overflow-y:auto'>"
            "<table class='pra-table'>"
            "<thead><tr>"
            "<th>Brand</th><th>Niche</th><th>Sources</th>"
            "<th class='num'>FB</th><th class='num'>SH</th>"
            "<th class='num'>LZ</th><th class='num'>TT</th>"
            "<th class='num'>Days</th><th>Score</th><th>Landing</th>"
            "</tr></thead>"
            f"<tbody>{''.join(_rows_html)}</tbody>"
            "</table>"
            "</div>"
        )
        st.markdown(_table_html, unsafe_allow_html=True)
        st.caption(
            "Brands appearing in multiple source columns are running coordinated campaigns — "
            "those are usually the most-scaled operators. Switch to dedicated tabs for deep-dive analysis."
        )

# ---- Pages render below as `if current_page == "X":` blocks ----
# Each top-level page is a conditional; content renders in natural order.
# Sub-sections (e.g. expanders inside FB Ads) are declared inline where needed.

if current_page == "shortlist":
    import html as _slh
    import hashlib as _slhash
    st.caption("Ang mga na-shortlist mong products — i-organize sa pipeline: Researching → Validated → Sourced → Decided. Ito ang decision hub mo.")

    _meta_all = db.all_brand_meta()
    _starred_meta = {b: m for b, m in _meta_all.items() if m.get("starred")}

    if not _starred_meta:
        # ---- Empty state: nudge toward the Radar where the ★ lives ----
        st.markdown(
            "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
            "border-radius:14px;padding:38px 30px;text-align:center;margin-top:8px'>"
            "<div style='font-size:2.4rem;line-height:1;margin-bottom:10px;color:var(--pra-accent)'>★</div>"
            "<div style='font-family:Fraunces,Georgia,serif;font-size:1.4rem;font-weight:600;"
            "color:var(--pra-text-strong);margin-bottom:8px'>Wala ka pang na-shortlist</div>"
            "<div style='color:var(--pra-text-muted);font-size:0.9rem;max-width:460px;margin:0 auto'>"
            "Pumunta sa <strong>Product Radar</strong> at i-tap ang <strong>★ Shortlist</strong> sa mga "
            "winning products na gusto mong i-research. Lalabas sila dito para ma-track mo mula "
            "research hanggang decision.</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        _cta1, _cta2, _cta3 = st.columns([1, 1, 1])
        with _cta2:
            if st.button("🛰️  Go to Product Radar", width="stretch", type="primary", key="sl_empty_radar"):
                st.session_state.current_page = "radar"
                st.rerun()
    else:
        # ---- Cross-run research per shortlisted brand (all scrapes, not just the
        #      loaded run). idx is fixed from the stable DB row order so per-card
        #      widget keys stay bound to their brand. ----
        _stage_order = {s: i for i, s in enumerate(_HUNT_PIPELINE)}
        _items = []
        for _i, (_b, _m) in enumerate(_starred_meta.items()):
            _stage = _canon_status(_m.get("status", ""))
            if _stage not in _HUNT_PIPELINE:
                _stage = ""
            _res = _brand_research(_b)
            _ws = _winner_score(_res)
            _items.append({
                "idx": _i, "brand": _b, "stage": _stage,
                "notes": (_m.get("notes") or ""),
                "res": _res, "ws": _ws, "score": _ws["total"],
                "saved": (_m.get("created_at") or "")[:10],
            })
        _items.sort(key=lambda x: (_stage_order.get(x["stage"], 99), -x["score"], x["brand"].lower()))

        # ---- Stage counts ----
        _counts = {s: 0 for s in _HUNT_PIPELINE}
        _unstaged = 0
        for _it in _items:
            if _it["stage"] in _counts:
                _counts[_it["stage"]] += 1
            else:
                _unstaged += 1
        _total = len(_items)

        # ---- Pipeline funnel (hero overview) ----
        _funnel = "<div style='display:flex;gap:10px;margin:12px 0 6px;flex-wrap:wrap'>"
        for _st in _HUNT_PIPELINE:
            _fc = _STATUS_COLOR[_st]
            _funnel += (
                "<div style='flex:1 1 130px;background:var(--pra-panel);border:1px solid var(--pra-border);"
                "border-top:3px solid " + _fc + ";border-radius:11px;padding:14px 16px'>"
                "<div style='font-family:Fraunces,Georgia,serif;font-size:1.7rem;font-weight:700;"
                "line-height:1;color:" + _fc + "'>" + str(_counts[_st]) + "</div>"
                "<div style='font-size:0.64rem;letter-spacing:0.12em;text-transform:uppercase;"
                "font-weight:700;color:var(--pra-text-muted);margin-top:7px'>" + _STATUS_LABEL[_st] + "</div>"
                "</div>"
            )
        _funnel += "</div>"
        st.markdown(_funnel, unsafe_allow_html=True)

        _avg_ws = round(sum(it["score"] for it in _items) / len(_items)) if _items else 0
        _summary_bits = [str(_total) + (" product" if _total == 1 else " products") + " shortlisted",
                         "avg Winner Score " + str(_avg_ws)]
        if _counts["decided"]:
            _summary_bits.append(str(_counts["decided"]) + " decided")
        if _unstaged:
            _summary_bits.append(str(_unstaged) + " unstaged")
        st.markdown(
            "<div style='color:var(--pra-text-dim);font-size:0.78rem;margin:0 0 12px'>"
            + "  ·  ".join(_summary_bits) + "</div>", unsafe_allow_html=True)

        # ---- Stage filter + sort ----
        _fcol, _sortcol = st.columns([3, 1])
        with _fcol:
            _filter_opts = ["All", "🔬 Researching", "✅ Validated", "🏭 Sourced", "🎯 Decided"]
            if _unstaged:
                _filter_opts.append("☆ Unstaged")
            _flt = st.segmented_control("Stage filter", _filter_opts, default="All",
                                        key="sl_stage_filter", label_visibility="collapsed")
        with _sortcol:
            _sort = st.selectbox("Sort", ["Pipeline", "Winner Score", "Longevity", "Recent"],
                                 key="sl_sort", label_visibility="collapsed")
        _flt_map = {"🔬 Researching": "researching", "✅ Validated": "validated",
                    "🏭 Sourced": "sourced", "🎯 Decided": "decided", "☆ Unstaged": ""}
        _shown = [it for it in _items if it["stage"] == _flt_map[_flt]] if _flt in _flt_map else list(_items)
        if _sort == "Winner Score":
            _shown.sort(key=lambda x: -x["score"])
        elif _sort == "Longevity":
            _shown.sort(key=lambda x: -(x["res"].get("max_days") or 0))
        elif _sort == "Recent":
            _shown.sort(key=lambda x: x["saved"], reverse=True)

        _sl_open = None
        if not _shown:
            st.info("Walang item sa stage na ito.")
        else:
            _stage_select_opts = [""] + _HUNT_PIPELINE
            _scols = st.columns(3, gap="medium")
            for _pos, _it in enumerate(_shown):
                _brand = _it["brand"]
                # Brand-stable widget key (NOT list position) — list order shifts when a
                # brand is starred/removed, and position keys would make Streamlit reuse a
                # neighbour's session_state and silently rewrite the wrong brand's stage/notes.
                _bk = _slhash.md5(_brand.encode("utf-8")).hexdigest()[:10]
                _stage = _it["stage"]
                _res = _it["res"]
                _ws = _it["ws"]
                _bcol = _STATUS_COLOR.get(_stage, "var(--pra-text-dim)")
                _badge_lbl = _STATUS_LABEL.get(_stage, "Unstaged") if _stage else "Unstaged"
                _wsc = _ws["verdict_color"]
                if _res.get("found"):
                    _stat = (str(_res.get("niche") or "—").title() + " · "
                             + str(_res.get("distinct_ads") or 0) + " ads · "
                             + str(_res.get("max_days") or 0) + "d · " + str(_res.get("momentum") or ""))
                    if (_res.get("active_now") or 0) > 0:
                        _dot = "<span style='color:var(--pra-success)'>● " + str(_res.get("active_now")) + " active</span>"
                    else:
                        _dot = "<span style='color:var(--pra-text-dim)'>○ inactive</span>"
                else:
                    _stat = "Bagong shortlist — mag-scan para sa research data"
                    _dot = ("<span style='color:var(--pra-text-dim)'>saved " + _it["saved"] + "</span>"
                            if _it["saved"] else "")
                _notes_prev = ""
                if _it["notes"]:
                    _nt = _it["notes"].strip().replace("\n", " ")
                    if len(_nt) > 88:
                        _nt = _nt[:88] + "…"
                    _notes_prev = ("<div style='color:var(--pra-text-muted);font-size:0.74rem;font-style:italic;"
                                   "margin-top:8px;border-left:2px solid var(--pra-border);padding-left:9px'>"
                                   + _slh.escape(_nt) + "</div>")
                with _scols[_pos % 3]:
                    st.markdown(
                        "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                        "border-radius:12px;padding:15px 17px;margin-bottom:4px;min-height:138px'>"
                        "<div style='display:flex;align-items:center;justify-content:space-between;margin-bottom:7px'>"
                        "<span style='display:inline-block;background:rgba(212,175,55,0.09);border:1px solid " + _bcol + ";"
                        "color:" + _bcol + ";padding:1px 9px;border-radius:20px;font-size:0.6rem;font-weight:700;"
                        "letter-spacing:0.08em;text-transform:uppercase'>" + _badge_lbl + "</span>"
                        "<span style='font-size:0.66rem;font-weight:700;color:" + _wsc + "'>WS " + str(_ws["total"]) + "</span>"
                        "</div>"
                        "<div style='font-family:Fraunces,Georgia,serif;font-size:1.2rem;font-weight:600;"
                        "color:var(--pra-text-strong);line-height:1.12'>" + _slh.escape(_brand) + "</div>"
                        "<div style='color:var(--pra-text-muted);font-size:0.73rem;margin-top:5px'>" + _slh.escape(_stat) + "</div>"
                        "<div style='font-size:0.66rem;margin-top:6px'>" + _dot + "</div>"
                        + _notes_prev +
                        "</div>",
                        unsafe_allow_html=True,
                    )
                    if st.button("🔬 Research deep-dive", key="sl_research_" + _bk, width="stretch"):
                        _sl_open = _it
                    _cc1, _cc2, _cc3 = st.columns([3, 1, 1])
                    _new_stage = _cc1.selectbox(
                        "Stage", _stage_select_opts,
                        index=_stage_select_opts.index(_stage),
                        key="sl_stage_" + _bk,
                        format_func=lambda x: "☆ Unstaged" if x == "" else _STATUS_LABEL.get(x, x.title()),
                        label_visibility="collapsed",
                    )
                    if _new_stage != _stage:
                        db.upsert_brand_meta(_brand, status=_new_stage)
                        st.toast(_brand + " → " + (_STATUS_LABEL.get(_new_stage, "Unstaged") if _new_stage else "Unstaged"))
                        st.rerun()
                    with _cc2.popover("📝", help="Notes"):
                        _nv = st.text_area(
                            "Notes", value=_it["notes"], height=120,
                            key="sl_notes_" + _bk,
                            placeholder="Checked Shopee — 5k sold · test order 6/01 · supplier on 1688…",
                            label_visibility="collapsed",
                        )
                        if st.button("Save notes", key="sl_notes_save_" + _bk, type="primary", width="stretch"):
                            db.upsert_brand_meta(_brand, notes=_nv)
                            st.toast("Notes saved")
                            st.rerun()
                    if _cc3.button("☆", key="sl_rm_" + _bk, help="Remove from shortlist", width="stretch"):
                        db.upsert_brand_meta(_brand, starred=0)
                        log_activity("unstarred", _brand)
                        st.toast(_brand + " removed from shortlist")
                        st.rerun()

        # ---- Open the full research deep-dive dialog (deferred to top level) ----
        if _sl_open:
            _ores = _sl_open["res"]
            _obrow = {
                "brand": _sl_open["brand"], "niche": _ores.get("niche") or "—",
                "category": _ores.get("category") or "", "sub_category": _ores.get("sub_category") or "",
                "location": _ores.get("location") or "", "ad_count": _ores.get("distinct_ads") or 0,
                "max_days_running": _ores.get("max_days") or 0,
                "score_normalized": _ores.get("score_sum") or 0,
                "mp_sold": _ores.get("mp_sold") or 0, "mp_price": _ores.get("mp_price"),
                "sample_landing_url": _ores.get("sample_landing") or "",
            }
            _render_brand_detail_dialog(_sl_open["brand"], _obrow, _ores.get("ads_full") or [])

if current_page == "radar":
    import html as _rhl
    st.caption("Lahat ng winning products across sources — ranked by score, tagged by momentum. Ito ang core hunt view.")

    _radar_agg = {}
    for r in rows:
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        if not r.get("is_active"):
            continue
        _b = (r.get("brand") or r.get("page_name") or "").strip()
        if not _b:
            continue
        _e = _radar_agg.setdefault(_b, {"brand": _b, "niche": r.get("niche") or "—",
                                        "category": (r.get("category") or ""), "location": (r.get("location") or ""),
                                        "score": 0.0, "days": 0, "ads": 0, "sh": 0, "lz": 0})
        _e["score"] += float(r.get("score_normalized") or 0)
        _e["days"] = max(_e["days"], r.get("days_running") or 0)
        _e["ads"] += 1
        _u = (r.get("landing_url") or "").lower()
        if "shopee" in _u or "shp.ee" in _u:
            _e["sh"] += 1
        elif "lazada" in _u:
            _e["lz"] += 1

    _radar = sorted(_radar_agg.values(), key=lambda x: -x["score"])
    _fresh_n = sum(1 for d in _radar if d["days"] <= 21)
    _proven_n = sum(1 for d in _radar if d["days"] > 60)

    if not _radar:
        st.info("Walang qualifying winners pa sa run na ito — i-loosen ang filters o mag-scrape.")
    else:
        _rs1, _rs2, _rs3 = st.columns(3)
        _rs1.metric("On radar", f"{len(_radar):,}", delta="qualifying products", delta_color="off")
        _rs2.metric("🔥 Fresh", f"{_fresh_n:,}", delta="≤21 days — early movers", delta_color="off")
        _rs3.metric("🏆 Proven", f"{_proven_n:,}", delta=">60 days — safe bets", delta_color="off")
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

        # ---- Saturation: competing brands per niche (Open = low competition) ----
        _niche_sat = Counter(str(d["niche"]).title() for d in _radar if d.get("niche") and d["niche"] != "—")
        def _sat_class(n):
            if n <= 4:
                return ("Open", "var(--pra-success)")
            if n <= 9:
                return ("Medium", "var(--pra-warning)")
            return ("Crowded", "var(--pra-danger)")
        if _niche_sat:
            _sat_chips = ""
            for _snm, _scnt in sorted(_niche_sat.items(), key=lambda x: x[1]):
                _sl0, _scol0 = _sat_class(_scnt)
                _sat_chips += ("<span style='display:inline-block;background:var(--pra-subtle-bg);"
                               "border:1px solid var(--pra-border);border-radius:14px;padding:4px 11px;"
                               "margin:0 6px 6px 0;font-size:0.72rem;color:var(--pra-text)'>"
                               + _snm + " <strong style='color:" + _scol0 + "'>" + str(_scnt) + " · " + _sl0 + "</strong></span>")
            st.markdown("<div style='color:var(--pra-text-muted);font-size:0.6rem;letter-spacing:0.12em;"
                        "text-transform:uppercase;font-weight:700;margin:2px 0 8px'>Niche saturation "
                        "<span style='color:var(--pra-text-dim);font-weight:400;letter-spacing:0;text-transform:none'>"
                        "· competing brands per niche (Open = mas madaling pasukin)</span></div>"
                        "<div style='margin-bottom:14px'>" + _sat_chips + "</div>", unsafe_allow_html=True)
        _mf = st.segmented_control("Momentum", ["All", "🔥 Fresh", "📈 Scaling", "🏆 Proven"],
                                   default="All", key="radar_momentum", label_visibility="collapsed")
        if _mf == "🔥 Fresh":
            _shown = [d for d in _radar if d["days"] <= 21]
        elif _mf == "📈 Scaling":
            _shown = [d for d in _radar if 21 < d["days"] <= 60]
        elif _mf == "🏆 Proven":
            _shown = [d for d in _radar if d["days"] > 60]
        else:
            _shown = _radar
        _shown = _shown[:30]

        _maxsc = max((d["score"] for d in _radar), default=1) or 1
        _rcols = st.columns(3, gap="medium")
        for _ri, _d in enumerate(_shown):
            if _d["days"] <= 21:
                _ml, _mc = "🔥 Fresh", "var(--pra-warning)"
            elif _d["days"] <= 60:
                _ml, _mc = "📈 Scaling", "var(--pra-info)"
            else:
                _ml, _mc = "🏆 Proven", "var(--pra-success)"
            _srcs = ["FB"]
            if _d["sh"]:
                _srcs.append("SH")
            if _d["lz"]:
                _srcs.append("LZ")
            _srcbadge = " · ".join(_srcs)
            _pct = int((_d["score"] / _maxsc) * 100) if _maxsc else 0
            _sub = str(_d["niche"]).title()
            if _d["category"]:
                _sub += " · " + str(_d["category"]).replace("_", " ").title()
            if _d["location"]:
                _sub += " · " + str(_d["location"])
            _csat_n = _niche_sat.get(str(_d["niche"]).title(), 0)
            _csat_lbl, _csat_col = _sat_class(_csat_n)
            with _rcols[_ri % 3]:
                _card = (
                    "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                    "border-radius:12px;padding:16px 18px;margin-bottom:4px;height:176px;"
                    "display:flex;flex-direction:column;justify-content:space-between'>"
                    "<div>"
                    "<div style='display:flex;align-items:center;justify-content:space-between;margin-bottom:8px'>"
                    "<span style='font-size:0.64rem;font-weight:700;letter-spacing:0.04em;color:" + _mc + "'>" + _ml + "</span>"
                    "<span style='font-size:0.62rem;color:var(--pra-text-dim);letter-spacing:0.06em'>" + _srcbadge + "</span>"
                    "</div>"
                    "<div style='font-family:Fraunces,Georgia,serif;font-size:1.25rem;font-weight:600;"
                    "color:var(--pra-text-strong);line-height:1.1'>" + _rhl.escape(_d["brand"]) + "</div>"
                    "<div style='color:var(--pra-text-muted);font-size:0.76rem;margin-top:5px'>" + _rhl.escape(_sub) + "</div>"
                    "<div style='color:var(--pra-text-dim);font-size:0.74rem;margin-top:6px'>"
                    + str(_d["days"]) + " days · " + str(_d["ads"]) + " ads · "
                    + "<span style='color:" + _csat_col + "'>" + _csat_lbl + " niche</span></div>"
                    "</div>"
                    "<div style='margin-top:10px;height:5px;background:var(--pra-subtle-bg);border-radius:3px;overflow:hidden'>"
                    "<div style='width:" + str(_pct) + "%;height:100%;background:linear-gradient(90deg,#D4AF37,#E6CC73)'></div></div>"
                    "</div>"
                )
                st.markdown(_card, unsafe_allow_html=True)
                if st.button("★ Shortlist", key="radar_star_" + str(_ri), width="stretch"):
                    _exm = db.get_brand_meta(_d["brand"]) or {}
                    db.upsert_brand_meta(
                        _d["brand"], starred=1,
                        status=(_canon_status(_exm.get("status", "")) or "researching"),
                    )
                    log_activity("starred", _d["brand"])
                    st.toast("★ " + _d["brand"] + " added to shortlist")

if current_page == "supplier":
    st.subheader("1688 supplier search")
    st.caption(
        "Find Chinese suppliers for any product. Paste a product name or keyword — "
        "we search 1688.com and convert prices to PHP. Closes the research-to-launch loop."
    )

    _sup_col1, _sup_col2 = st.columns([4, 1])
    _sup_query = _sup_col1.text_input(
        "Product / keyword (English or Chinese)",
        placeholder="e.g. slimming capsule, whitening cream, ginger oil",
        key="supplier_query",
        label_visibility="collapsed",
    )
    _sup_run = _sup_col2.button("Search 1688", key="supplier_search_btn",
                                  type="primary", width="stretch",
                                  disabled=not _sup_query.strip())

    if IS_CLOUD and _sup_run:
        st.info("1688 supplier search runs on the local desktop app only (needs Playwright).")
    elif _sup_run and _sup_query.strip():
        with st.status(f"Searching 1688 for {_sup_query!r}...", state="running",
                        expanded=True) as _sup_status:
            st.write("Launching Playwright (headless Chromium)…")
            st.write("Visiting 1688 search page…")
            try:
                import supplier_matcher
                _sup_results = supplier_matcher.search_suppliers(
                    _sup_query.strip(), max_results=10, headless=True,
                )
                st.session_state["supplier_results"] = _sup_results
                st.session_state["supplier_last_query"] = _sup_query.strip()
                _sup_status.update(
                    label=f"Found {len(_sup_results)} suppliers", state="complete",
                )
            except Exception as e:
                st.session_state["supplier_results"] = []
                _sup_status.update(label=f"Search failed: {e}", state="error")

    _sup_results = st.session_state.get("supplier_results")
    if _sup_results:
        try:
            import supplier_matcher as _sm_ref
            _rate = _sm_ref.RMB_TO_PHP
        except Exception:
            _rate = 7.95
        st.caption(
            f"Showing results for **{st.session_state.get('supplier_last_query', '')!r}** · "
            f"prices converted at ₱{_rate:.2f}/¥"
        )
        if not _sup_results:
            st.info("No suppliers found. Try a different keyword (English product names sometimes work better).")
        else:
            import html as _h_sup
            _rows_html = []
            for r in _sup_results:
                _price_rmb = f"¥{r['price_rmb']:.2f}" if r.get("price_rmb") else "—"
                _price_php = f"₱{r['price_php']:.0f}" if r.get("price_php") else "—"
                _moq = f"{r['moq']} pcs" if r.get("moq") else "—"
                _supplier = _h_sup.escape(r.get("supplier", "") or "—")
                _title = _h_sup.escape(r.get("title", "—"))
                _url = _h_sup.escape(r.get("url", ""))
                _url_html = (
                    f'<a class="pra-landing" href="{_url}" target="_blank">Open on 1688 →</a>'
                    if _url else "—"
                )
                _rows_html.append(
                    "<tr>"
                    f'<td class="brand-cell">{_title}</td>'
                    f'<td class="num">{_price_rmb}</td>'
                    f'<td class="num"><strong>{_price_php}</strong></td>'
                    f'<td class="num">{_moq}</td>'
                    f'<td>{_supplier}</td>'
                    f'<td>{_url_html}</td>'
                    "</tr>"
                )
            _table_html = (
                "<div style='max-height:600px;overflow-y:auto'>"
                "<table class='pra-table'>"
                "<thead><tr>"
                "<th>Product title</th>"
                "<th class='num'>Price (¥)</th>"
                "<th class='num'>Price (₱)</th>"
                "<th class='num'>MOQ</th>"
                "<th>Supplier</th>"
                "<th>Link</th>"
                "</tr></thead>"
                f"<tbody>{''.join(_rows_html)}</tbody>"
                "</table>"
                "</div>"
            )
            st.markdown(_table_html, unsafe_allow_html=True)
            st.caption(
                "**Workflow:** click 'Open on 1688' to see the listing · contact supplier via 1688 chat · "
                "negotiate MOQ + sample · ship to PH agent (e.g. Aliship, EZShip) · sell on Shopee/Lazada."
            )
    else:
        st.info(
            "Enter a product keyword above and click **Search 1688** to find suppliers. "
            "Tip: try the brand name from any winning ad — like *'TriHealth probiotic'* or "
            "*'slimming capsule l-carnitine'*."
        )

    # ============================================================
    # PROFIT MARGIN CALCULATOR — works standalone (no supplier search needed)
    # ============================================================
    st.markdown("<div style='margin-top:36px'></div>", unsafe_allow_html=True)
    # ---- Courier rate calculator (Phase 16.5) ----
    st.subheader("Courier rate calculator")
    st.caption(
        "Quote J&T / LBC / Ninja Van / JRS for your COGS budgeting. "
        "**Rates are approximate** (late-2025) — verify with the courier before quoting customers."
    )
    import courier_calc as _cc
    _cr_a, _cr_b, _cr_c = st.columns([2, 2, 1])
    with _cr_a:
        _cr_weight = st.number_input(
            "Package weight (kg)", min_value=0.05, max_value=30.0,
            value=0.5, step=0.05, key="cr_weight",
            help="Total weight including box + filler. 1 capsule bottle ≈ 0.3–0.5 kg.",
        )
    with _cr_b:
        _cr_dest = st.selectbox(
            "Destination",
            _cc.all_destinations(),
            index=0, key="cr_dest",
            help="PH location of buyer. PH-wide ads typically dominate Metro Manila + Luzon.",
        )
    with _cr_c:
        _cr_order_val = st.number_input(
            "Order value (₱)", min_value=0.0, value=0.0, step=100.0,
            key="cr_order_val",
            help="For COD fee calc. Leave 0 to skip COD.",
        )

    _cr_quotes = _cc.quote(_cr_weight, _cr_dest)
    if _cr_quotes:
        _cr_html = (
            "<table class='pra-table' style='margin-top:6px'>"
            "<thead><tr>"
            "<th>Courier</th><th>Zone</th><th class='num'>Base ₱</th>"
            "<th class='num'>COD fee</th><th class='num'>Total ₱</th>"
            "</tr></thead><tbody>"
        )
        for _q in _cr_quotes:
            _cod = _cc.cod_fee(_q["courier"], _cr_order_val) if _cr_order_val else 0
            _total = _q["price_php"] + _cod
            _cr_html += (
                f"<tr><td class='brand-cell'>{_q['courier']}</td>"
                f"<td>{_q['zone']}</td>"
                f"<td class='num'>₱{_q['price_php']:.2f}</td>"
                f"<td class='num'>₱{_cod:.2f}</td>"
                f"<td class='num'><strong>₱{_total:.2f}</strong></td></tr>"
            )
        _cr_html += "</tbody></table>"
        st.markdown(_cr_html, unsafe_allow_html=True)
        st.caption(
            "Cheapest courier shown first. Add the **Total** to your COGS when "
            "calculating margin below."
        )
    else:
        st.warning("Couldn't map destination to a zone. Try a major city.")

    st.divider()

    st.subheader("Profit margin calculator")
    st.caption(
        "Plug in costs to see if a product is actually profitable before launching. "
        "All values in PHP unless noted."
    )

    with st.form("margin_calc_form"):
        mc_a, mc_b = st.columns(2)
        with mc_a:
            mc_sell_price = st.number_input(
                "Sell price (₱) per unit *", min_value=0.0, value=899.0, step=10.0,
                help="What you'll charge on Shopee / Lazada / your funnel",
            )
            mc_supplier_cost = st.number_input(
                "Supplier cost (₱) per unit *", min_value=0.0, value=120.0, step=5.0,
                help="From 1688/AliExpress, converted to PHP",
            )
            mc_shipping_cost = st.number_input(
                "Shipping CN→PH per unit (₱)", min_value=0.0, value=80.0, step=5.0,
                help="Sea freight per piece ~₱50-100, air ~₱150-300",
            )
        with mc_b:
            mc_ad_cost = st.number_input(
                "Ad cost per sale / CPA (₱)", min_value=0.0, value=150.0, step=10.0,
                help="Your cost per acquisition from FB/TikTok ads (estimate ₱100-300 for PH health/beauty)",
            )
            mc_marketplace_fee_pct = st.number_input(
                "Marketplace fee (%)", min_value=0.0, max_value=100.0, value=8.0, step=0.5,
                help="Shopee ~5-8% · Lazada ~5-10% · own funnel ~3-4% (payment gateway only)",
            )
            mc_other_cost = st.number_input(
                "Other cost per unit (₱)", min_value=0.0, value=20.0, step=5.0,
                help="Packaging, COD fees, returns provision, etc.",
            )
        mc_submitted = st.form_submit_button(
            "Calculate margin", type="primary", width="stretch",
        )

    if mc_submitted or st.session_state.get("margin_calc_seen", False):
        st.session_state["margin_calc_seen"] = True
        mp_fee = mc_sell_price * (mc_marketplace_fee_pct / 100)
        total_cost = (mc_supplier_cost + mc_shipping_cost + mc_ad_cost
                      + mp_fee + mc_other_cost)
        profit_per_unit = mc_sell_price - total_cost
        margin_pct = (profit_per_unit / mc_sell_price * 100) if mc_sell_price else 0
        breakeven_roas = (mc_sell_price / mc_ad_cost) if mc_ad_cost else 0
        # COGS-based markup ratio (helps gauge "x markup" mental model)
        markup_x = (mc_sell_price / max(mc_supplier_cost, 0.01))

        # Status tag
        if margin_pct >= 35:
            _tag = ("PROFITABLE", "#7CC4A0")
            _verdict = "Healthy margin. Launch confidently."
        elif margin_pct >= 15:
            _tag = ("MARGINAL", "#E6CC73")
            _verdict = "Tight margin. Volume + low return rate needed to make it work."
        else:
            _tag = ("UNPROFITABLE", "#E0909F")
            _verdict = "Losing money or barely breaking even. Renegotiate supplier or raise price."

        st.markdown(
            f"<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
            f"border-left:3px solid {_tag[1]};border-radius:6px;padding:18px 22px'>"
            f"<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:10px'>"
            f"<span style='color:{_tag[1]};font-size:0.72rem;letter-spacing:0.14em;"
            f"font-weight:700;text-transform:uppercase'>● {_tag[0]}</span>"
            f"<span style='color:var(--pra-text-muted);font-size:0.78rem'>{_verdict}</span>"
            f"</div>"
            f"<div style='display:flex;gap:24px;margin-top:14px;flex-wrap:wrap'>"
            f"<div><div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.08em;text-transform:uppercase'>Profit/unit</div>"
            f"<div style='font-size:1.5rem;font-weight:600;color:var(--pra-text);font-variant-numeric:tabular-nums'>₱{profit_per_unit:,.0f}</div></div>"
            f"<div><div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.08em;text-transform:uppercase'>Margin</div>"
            f"<div style='font-size:1.5rem;font-weight:600;color:{_tag[1]};font-variant-numeric:tabular-nums'>{margin_pct:.1f}%</div></div>"
            f"<div><div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.08em;text-transform:uppercase'>Breakeven ROAS</div>"
            f"<div style='font-size:1.5rem;font-weight:600;color:var(--pra-text);font-variant-numeric:tabular-nums'>{breakeven_roas:.2f}×</div></div>"
            f"<div><div style='color:var(--pra-text-muted);font-size:0.7rem;letter-spacing:0.08em;text-transform:uppercase'>Markup</div>"
            f"<div style='font-size:1.5rem;font-weight:600;color:var(--pra-text);font-variant-numeric:tabular-nums'>{markup_x:.1f}×</div></div>"
            f"</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
        st.caption(
            f"**Breakdown:** Sell ₱{mc_sell_price:,.0f} − Supplier ₱{mc_supplier_cost:,.0f} "
            f"− Shipping ₱{mc_shipping_cost:,.0f} − Ads ₱{mc_ad_cost:,.0f} "
            f"− Marketplace fee ₱{mp_fee:,.0f} − Other ₱{mc_other_cost:,.0f} "
            f"= **₱{profit_per_unit:,.0f} profit per sale**."
        )


# ============================================================================
# Copy Studio — AI ad copy generator (Taglish, uses winning hooks from DB)
# ============================================================================

if current_page == "copy_studio":
    st.subheader("AI Ad Copy Studio")
    st.caption(
        "Generate Taglish ad variations that weave in **winning hook phrases** "
        "automatically pulled from your scraped corpus. Powered by Claude. "
        "Cost ~₱0.30 per generation."
    )

    # API key status banner
    _has_anthropic_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
    if not _has_anthropic_key:
        st.warning(
            "⚠ `ANTHROPIC_API_KEY` env var not set. To enable AI generation:\n\n"
            "1. Get a key from https://console.anthropic.com/\n"
            "2. In PowerShell: `$env:ANTHROPIC_API_KEY = \"sk-ant-...\"`\n"
            "3. Restart the dashboard"
        )

    with st.form("copy_studio_form"):
        cs_a, cs_b = st.columns([3, 2])
        with cs_a:
            cs_product = st.text_input(
                "Product / brand to write for *",
                placeholder="e.g. Glow Lean Fit Slimming Capsule",
                help="The product you're going to launch / advertise.",
            )
            cs_audience = st.text_input(
                "Target audience (optional)",
                placeholder="e.g. moms 25-45, weight-conscious, active on Shopee",
                help="Helps the AI tailor the voice and angles",
            )
        with cs_b:
            _cs_niches = ["all"] + list(load_config().get("niches", {}).keys())
            cs_niche = st.selectbox(
                "Niche", _cs_niches, index=0,
                format_func=lambda x: "Use all niches' hooks" if x == "all" else x.replace("_", " "),
                help="Pulls winning hook phrases from this niche to ground the AI generation",
            )
            cs_language = st.selectbox(
                "Language",
                [
                    "Taglish", "Tagalog", "English",
                    "Bisaya (Cebuano)", "Bisalish (Bisaya-English mix)",
                    "Ilocano", "Ilocanglish (Ilocano-English mix)",
                    "Hiligaynon (Ilonggo)",
                ],
                index=0,
                help=(
                    "Match the language to your target region:\n"
                    "• Taglish/Tagalog — Metro Manila + Luzon\n"
                    "• Bisaya / Bisalish — Cebu, Davao, Visayas, Mindanao\n"
                    "• Ilocano / Ilocanglish — Ilocos, Pangasinan, Northern Luzon\n"
                    "• Hiligaynon — Iloilo, Bacolod, Western Visayas"
                ),
            )

        cs_submit = st.form_submit_button(
            "✨ Generate 6 ad variations", type="primary", width="stretch",
            disabled=not _has_anthropic_key,
        )

    if cs_submit and cs_product.strip():
        with st.spinner("Pulling winning hooks + generating ad copy..."):
            import ad_copy_generator
            _gen_result = ad_copy_generator.generate(
                product_name=cs_product.strip(),
                niche=cs_niche,
                language=cs_language,
                audience_note=cs_audience.strip(),
            )
            st.session_state["last_copy_gen"] = _gen_result

    _gen = st.session_state.get("last_copy_gen")
    if _gen:
        if not _gen["ok"]:
            st.error(f"Generation failed: {_gen['error']}")
        else:
            # Show which hooks the AI was given
            if _gen.get("hooks_used"):
                with st.expander(f"🧠 Hooks the AI used ({len(_gen['hooks_used'])} phrases)", expanded=False):
                    st.markdown(
                        "<div style='color:var(--pra-text-muted);font-size:0.82rem;line-height:1.7'>"
                        + " · ".join(f'<span style="color:var(--pra-accent)">{h}</span>'
                                     for h in _gen["hooks_used"])
                        + "</div>",
                        unsafe_allow_html=True,
                    )

            # Display the ad copy in a nice container
            st.markdown(
                f"<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                f"border-left:3px solid var(--pra-accent);border-radius:6px;"
                f"padding:20px 24px;font-size:0.92rem;line-height:1.7;white-space:pre-wrap;"
                f"color:var(--pra-text);font-family:Inter,sans-serif'>"
                f"{_gen['text']}"
                f"</div>",
                unsafe_allow_html=True,
            )

            # ---- Phase 17.2: Hook performance prediction ----
            # Train on PH-confident in-niche active ads from THIS run + score the generated copy
            try:
                import hook_predictor
                _hp_model = hook_predictor.train_from_rows(rows, min_count=2)
                if _hp_model:
                    _hp_result = hook_predictor.predict(_gen["text"], _hp_model)
                    _tier_color = {
                        "Strong": "var(--pra-success)",
                        "Decent": "var(--pra-accent)",
                        "Weak":   "var(--pra-warning)",
                    }.get(_hp_result["tier"], "var(--pra-text-muted)")
                    _conf_label = _hp_result["confidence"].capitalize()
                    st.markdown(
                        f"<div style='background:rgba(124,196,160,0.05);"
                        f"border:1px solid var(--pra-border);"
                        f"border-left:2px solid {_tier_color};border-radius:5px;"
                        f"padding:10px 14px;margin-top:10px;font-size:0.84rem'>"
                        f"<span style='color:{_tier_color};font-weight:700;letter-spacing:0.14em;"
                        f"font-size:0.66rem;text-transform:uppercase;margin-right:10px'>"
                        f"Predicted: {_hp_result['tier']}</span>"
                        f"<span style='color:var(--pra-text)'>"
                        f"Estimated longevity <strong>{_hp_result['predicted_days']:.0f} days</strong> "
                        f"based on {_hp_result['n_matches']} matching hook patterns "
                        f"from your DB. Confidence: <strong>{_conf_label}</strong>.</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                    if _hp_result["matched_ngrams"]:
                        with st.expander("Show matched hook patterns", expanded=False):
                            for ng, avg_days, count in _hp_result["matched_ngrams"]:
                                st.markdown(
                                    f"<div style='display:flex;justify-content:space-between;"
                                    f"padding:4px 0;font-size:0.82rem;"
                                    f"border-bottom:1px solid var(--pra-border)'>"
                                    f"<span style='color:var(--pra-text)'>{ng}</span>"
                                    f"<span style='color:var(--pra-text-muted);"
                                    f"font-variant-numeric:tabular-nums'>"
                                    f"{avg_days:.0f}d avg · {count}× seen</span>"
                                    f"</div>",
                                    unsafe_allow_html=True,
                                )
            except Exception:
                pass  # predictor is enhancement, not critical

            # Copy + regenerate buttons
            _cp_c1, _cp_c2, _cp_c3 = st.columns(3)
            _cp_c1.download_button(
                "📥 Download as .txt",
                data=_gen["text"],
                file_name=f"ad_copy_{cs_product.strip().replace(' ', '_')[:30]}.txt",
                mime="text/plain",
                width="stretch",
            )
            if _cp_c2.button("🔄 Regenerate (different angles)", width="stretch"):
                import ad_copy_generator
                with st.spinner("Generating new variations..."):
                    st.session_state["last_copy_gen"] = ad_copy_generator.generate(
                        product_name=cs_product.strip(),
                        niche=cs_niche,
                        language=cs_language,
                        audience_note=cs_audience.strip(),
                    )
                st.rerun()
            if _cp_c3.button("Clear", width="stretch"):
                st.session_state.pop("last_copy_gen", None)
                st.rerun()
    elif _has_anthropic_key:
        st.info(
            "Fill in the product name + niche above, then click **Generate**. "
            "The AI will pull your top hook phrases from the DB and weave them into "
            "6 different ad angles (problem-solution, social proof, scarcity, etc.)."
        )

    # ---- Phase 17.3 — Image prompt generator ----
    st.divider()
    st.subheader("AI image prompt generator")
    st.caption(
        "Generate 6 ready-to-paste prompts for Midjourney / Stable Diffusion / Flux to "
        "create scroll-stopping ad images. Cost ~₱0.30 per generation."
    )
    with st.form("image_prompt_form"):
        _ip_a, _ip_b = st.columns([3, 2])
        with _ip_a:
            ip_product = st.text_input(
                "Product to visualize *",
                placeholder="e.g. Glow Lean Fit Slimming Capsule",
                key="ip_product",
            )
            ip_copy = st.text_area(
                "Sample ad copy (optional — guides the vibe)",
                placeholder="Paste a winning ad copy to match its energy",
                height=80, key="ip_copy",
            )
        with _ip_b:
            _ip_niches = ["all"] + list(load_config().get("niches", {}).keys())
            ip_niche = st.selectbox(
                "Niche", _ip_niches, index=0, key="ip_niche",
                format_func=lambda x: "All" if x == "all" else x.replace("_", " "),
            )
            ip_audience = st.text_input(
                "Target audience",
                placeholder="e.g. Filipina moms 25-45, urban, weight-conscious",
                key="ip_audience",
            )
        ip_submit = st.form_submit_button(
            "✨ Generate 6 image prompts", type="primary", width="stretch",
            disabled=not _has_anthropic_key,
        )
    if ip_submit and ip_product.strip():
        with st.spinner("Claude is writing your image prompts..."):
            import image_prompt_gen
            st.session_state["last_image_prompts"] = image_prompt_gen.generate(
                product_name=ip_product.strip(),
                ad_copy=ip_copy.strip(),
                niche=ip_niche if ip_niche != "all" else "",
                target_audience=ip_audience.strip(),
            )
    _ip_gen = st.session_state.get("last_image_prompts")
    if _ip_gen:
        if not _ip_gen["ok"]:
            st.error(f"Generation failed: {_ip_gen['error']}")
        else:
            st.markdown(
                f"<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                f"border-left:3px solid var(--pra-info);border-radius:6px;"
                f"padding:18px 22px;font-size:0.88rem;line-height:1.7;"
                f"color:var(--pra-text);font-family:Inter,sans-serif'>"
                f"{_ip_gen['prompts_markdown']}</div>",
                unsafe_allow_html=True,
            )
            _ipd_a, _ipd_b = st.columns([3, 1])
            _ipd_a.download_button(
                "📥 Download as .txt",
                data=_ip_gen["prompts_markdown"],
                file_name=f"image_prompts_{ip_product.strip().replace(' ', '_')[:30]}.txt",
                mime="text/plain", width="stretch",
            )
            if _ipd_b.button("Clear", key="ip_clear", width="stretch"):
                st.session_state.pop("last_image_prompts", None)
                st.rerun()

    # ---- Phase 17.4 — Landing page copy generator ----
    st.divider()
    st.subheader("AI landing page generator")
    st.caption(
        "Generate a full Markdown landing page (headline, USPs, testimonials, FAQ, CTA) "
        "ready to paste into Wix / Shopify / Shopee description. Cost ~₱1-2 per page."
    )
    with st.form("lp_gen_form"):
        _lp_a, _lp_b = st.columns([3, 2])
        with _lp_a:
            lp_product = st.text_input(
                "Product name *", key="lp_product",
                placeholder="e.g. Glow Lean Fit Slimming Capsule",
            )
            lp_pain = st.text_area(
                "Pain point / problem to address",
                placeholder="e.g. Sumusobra na ang timbang at nahihiya na lumabas",
                height=70, key="lp_pain",
            )
            lp_ingredients = st.text_area(
                "Key ingredients / features",
                placeholder="e.g. Garcinia Cambogia, Green Tea Extract, CLA",
                height=60, key="lp_ingredients",
            )
        with _lp_b:
            _lp_niches = ["all"] + list(load_config().get("niches", {}).keys())
            lp_niche = st.selectbox(
                "Niche", _lp_niches, index=0, key="lp_niche",
                format_func=lambda x: "All" if x == "all" else x.replace("_", " "),
            )
            lp_price = st.number_input(
                "Sell price (₱)", min_value=0.0, value=899.0, step=10.0,
                key="lp_price",
            )
            lp_audience = st.text_input(
                "Target audience", key="lp_audience",
                placeholder="e.g. Moms 25-45, urban",
            )
            lp_language = st.selectbox(
                "Language",
                ["Taglish", "Tagalog", "English", "Bisaya (Cebuano)",
                 "Ilocano", "Hiligaynon (Ilonggo)"],
                index=0, key="lp_language",
            )
        lp_submit = st.form_submit_button(
            "📝 Generate landing page", type="primary", width="stretch",
            disabled=not _has_anthropic_key,
        )
    if lp_submit and lp_product.strip():
        with st.spinner("Claude is writing your landing page..."):
            import landing_page_gen
            st.session_state["last_landing_page"] = landing_page_gen.generate(
                product_name=lp_product.strip(),
                niche=lp_niche if lp_niche != "all" else "",
                price_php=float(lp_price),
                pain_point=lp_pain.strip(),
                target_audience=lp_audience.strip(),
                ingredients=lp_ingredients.strip(),
                language=lp_language,
            )
    _lp_gen = st.session_state.get("last_landing_page")
    if _lp_gen:
        if not _lp_gen["ok"]:
            st.error(f"Generation failed: {_lp_gen['error']}")
        else:
            with st.container(border=True):
                st.markdown(_lp_gen["markdown"])
            _lpd_a, _lpd_b = st.columns([3, 1])
            _lpd_a.download_button(
                "📥 Download as .md",
                data=_lp_gen["markdown"],
                file_name=f"landing_{lp_product.strip().replace(' ', '_')[:30]}.md",
                mime="text/markdown", width="stretch",
            )
            if _lpd_b.button("Clear", key="lp_clear", width="stretch"):
                st.session_state.pop("last_landing_page", None)
                st.rerun()


# ============================================================================
# Testing tab — product testing lineup with PDF export
# (Content is rendered at the bottom of this file inside `if current_page == "testing":`)
# ============================================================================


# ---- FB Ads Library page (Top Brands → Suggestions → Expanders) ----

if current_page == "fb_ads":
    # ============================================================
    #  RESEARCH AGENT — analyst brief + hero card + niche heat
    # ============================================================
    from datetime import datetime as _dt
    _hour = _dt.now().hour
    if _hour < 12:
        _greet = "Good morning"
    elif _hour < 18:
        _greet = "Good afternoon"
    else:
        _greet = "Good evening"
    _date_today = _dt.now().strftime("%A, %B %d")

    # ---- LLM auto-classify trigger (Phase 9.4) — premium mini-card style ----
    _n_unclassified = sum(
        1 for r in rows
        if (r.get("ad_text") or "").strip() and not r.get("llm_classified_at")
    )
    _n_classified = sum(1 for r in rows if r.get("llm_classified_at"))
    if _n_unclassified > 0 or _n_classified > 0:
        _has_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
        _llm_l, _llm_r = st.columns([4, 1])
        with _llm_l:
            if _n_unclassified > 0:
                _msg_body = (
                    f"<strong>{_n_classified:,}</strong> tagged · "
                    f"<strong>{_n_unclassified:,}</strong> pending"
                )
                if not _has_key:
                    _msg_body += (
                        " · <span style='color:var(--pra-warning)'>"
                        "Set ANTHROPIC_API_KEY to enable</span>"
                    )
                _accent_color = "var(--pra-accent)"
            else:
                _msg_body = (
                    f"All <strong>{_n_classified:,}</strong> ads classified — "
                    f"use sidebar filters to slice by hook angle or demo."
                )
                _accent_color = "var(--pra-success)"
            st.markdown(
                f"<div style='background:linear-gradient(135deg,rgba(212,175,55,0.04) 0%,"
                f"rgba(212,175,55,0.01) 100%);border:1px solid var(--pra-border);"
                f"border-left:2px solid {_accent_color};border-radius:5px;"
                f"padding:10px 14px;font-size:0.84rem;color:var(--pra-text)'>"
                f"<span style='color:{_accent_color};font-weight:700;letter-spacing:0.14em;"
                f"font-size:0.66rem;text-transform:uppercase;margin-right:10px'>AI tags</span>"
                f"<span style='color:var(--pra-text-muted)'>{_msg_body}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with _llm_r:
            if _n_unclassified > 0:
                if st.button(
                    "✨ Classify ~$1", key="llm_classify_btn",
                    type="primary", width="stretch",
                    disabled=not _has_key,
                    help=("Classify up to 200 ads via Claude API (~₱60/$1). "
                          "Adds hook angle, claim type, target demo tags." if _has_key
                          else "ANTHROPIC_API_KEY env var not set"),
                ):
                    with st.spinner(f"Classifying up to 200 ads via Claude…"):
                        import llm_classifier
                        _attempt, _ok = llm_classifier.classify_run(
                            active_run_id, max_ads=200,
                        )
                    st.toast(f"✓ Classified {_ok}/{_attempt} ads", icon=None)
                    get_ads.clear()
                    _clear_brand_caches()
                    st.rerun()

    _narrative = _agent_narrative(active_run_id, runs, rows)
    if _narrative:
        # Compact analyst note (no heavy gradient panel)
        st.markdown(
            f"<div style='margin:4px 4px 22px 4px;line-height:1.7;color:var(--pra-text);"
            f"font-size:0.92rem'>"
            f"<span style='color:var(--pra-text-muted);font-size:0.72rem;letter-spacing:0.1em;"
            f"font-weight:600;text-transform:uppercase;margin-right:10px'>"
            f"{_greet} · {_date_today}</span>"
            f"{_narrative}"
            f"</div>",
            unsafe_allow_html=True,
        )

    # ---- Pre-compute brand_rows so we can render hero + table both ----
    _all_meta_pre = db.all_brand_meta()
    _by_brand_pre: dict[str, dict] = {}
    for r in filtered:
        b = (r.get("brand") or r.get("page_name") or "?").strip()
        if not b:
            continue
        # Apply same default ranking criteria as table
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        if not r.get("is_active"):
            continue
        entry = _by_brand_pre.setdefault(b, {
            "brand": b, "niche": r.get("niche"), "score": 0.0,
            "max_days": 0, "ad_count": 0, "landing": "",
            "ad_text": "", "creative": None,
        })
        entry["score"] += float(r.get("score_normalized") or 0)
        entry["max_days"] = max(entry["max_days"], r.get("days_running") or 0)
        entry["ad_count"] += 1
        if not entry["landing"] and r.get("landing_url"):
            entry["landing"] = r["landing_url"]
        if not entry["ad_text"] and r.get("ad_text"):
            entry["ad_text"] = (r.get("ad_text") or "")[:280]
        if not entry["creative"] and r.get("creative_path"):
            entry["creative"] = _creative_path_url(r.get("creative_path"))

    _hero = None
    if _by_brand_pre:
        _hero_candidates = sorted(_by_brand_pre.values(), key=lambda x: -x["score"])
        _hero = _hero_candidates[0] if _hero_candidates else None

    # ---- Hero card: today's top winner ----
    if _hero:
        from urllib.parse import urlparse as _up2, quote_plus as _qp
        try:
            _hero_dom = _up2(_hero["landing"]).netloc.replace("www.", "")
        except Exception:
            _hero_dom = ""
        _hero_meta = _all_meta_pre.get(_hero["brand"], {})
        _hero_starred = bool(_hero_meta.get("starred"))
        _hero_niche = (_hero.get("niche") or "—").replace("● ", "")
        _hero_score = int(_hero.get("score") or 0)

        col_hero_l, col_hero_r = st.columns([3, 2])
        with col_hero_l:
            _shortlist_chip = (
                "<span style='display:inline-block;background:rgba(212,175,55,0.12);"
                "color:var(--pra-accent);padding:2px 8px;border-radius:3px;"
                "font-size:0.62rem;font-weight:600;letter-spacing:0.1em;"
                "text-transform:uppercase;margin-left:10px;vertical-align:middle'>"
                "★ Shortlist</span>" if _hero_starred else ""
            )
            st.markdown(
                f"<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                f"border-radius:6px;padding:20px 22px'>"
                f"<div style='color:var(--pra-accent);font-size:0.65rem;letter-spacing:0.16em;"
                f"font-weight:700;text-transform:uppercase;margin-bottom:6px'>Today's top winner</div>"
                f"<div style='font-size:1.5rem;font-weight:600;letter-spacing:-0.015em;"
                f"color:var(--pra-text);line-height:1.15;margin-bottom:4px'>"
                f"{_hero['brand']}{_shortlist_chip}</div>"
                f"<div style='color:var(--pra-text-muted);font-size:0.8rem;margin-bottom:12px;"
                f"font-variant-numeric:tabular-nums'>"
                f"{_hero_niche} · <strong style='color:var(--pra-text)'>{_hero['max_days']:,}</strong>d · "
                f"<strong style='color:var(--pra-text)'>{_hero['ad_count']}</strong> variants · "
                f"score <strong style='color:var(--pra-accent)'>{_hero_score:,}</strong>"
                f"{(' · ' + _hero_dom) if _hero_dom else ''}"
                f"</div>"
                f"<div style='color:var(--pra-text-muted);font-size:0.84rem;line-height:1.55;"
                f"font-style:italic;border-left:2px solid var(--pra-accent);padding-left:10px'>"
                f"&ldquo;{(_hero.get('ad_text') or '—').replace(chr(10),' ')[:200]}&rdquo;"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
            hac1, hac2, hac3 = st.columns(3)
            if _hero["landing"]:
                hac1.link_button("Open landing", _hero["landing"], width="stretch", type="primary")
            hac2.link_button(
                "Search Shopee",
                f"https://shopee.ph/search?keyword={_qp(_hero['brand'])}",
                width="stretch",
            )
            hac3.link_button(
                "Search 1688",
                f"https://s.1688.com/selloffer/offer_search.htm?keywords={_qp(_hero['brand'])}",
                width="stretch",
            )

        # ---- Niche heat mini-viz on the right ----
        with col_hero_r:
            _by_niche_count: dict[str, int] = defaultdict(int)
            for b in _by_brand_pre.values():
                n = (b.get("niche") or "—").replace("● ", "")
                _by_niche_count[n] += 1
            if _by_niche_count:
                _max_n = max(_by_niche_count.values()) or 1
                _heat_html = (
                    "<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                    "border-radius:6px;padding:20px 22px;height:100%'>"
                    "<div style='color:var(--pra-text-muted);font-size:0.65rem;letter-spacing:0.16em;"
                    "font-weight:700;text-transform:uppercase;margin-bottom:14px'>Niche heat</div>"
                )
                for niche_name, cnt in sorted(_by_niche_count.items(), key=lambda x: -x[1]):
                    pct = int((cnt / _max_n) * 100)
                    _heat_html += (
                        f"<div style='display:flex;align-items:center;margin-bottom:8px'>"
                        f"<div style='width:62px;color:var(--pra-text);font-size:0.8rem'>{niche_name}</div>"
                        f"<div style='flex:1;background:var(--pra-border);border-radius:2px;height:5px;"
                        f"margin:0 10px;overflow:hidden'>"
                        f"<div style='background:var(--pra-accent);"
                        f"width:{pct}%;height:100%'></div></div>"
                        f"<div style='width:28px;text-align:right;color:var(--pra-text);font-size:0.78rem;"
                        f"font-variant-numeric:tabular-nums'>{cnt}</div>"
                        f"</div>"
                    )
                _heat_html += "</div>"
                st.markdown(_heat_html, unsafe_allow_html=True)

    st.markdown("<div style='margin-top:32px'></div>", unsafe_allow_html=True)

    # --- Quick-filter chips ---
    _chip_opts = ["All", "Proven 90+ days", "New winners (30+ days)", "Has sales data", "My shortlist"]
    chip = st.pills("Quick filter", _chip_opts, default="All", label_visibility="collapsed")

    _starred_set: set[str] = set()
    if chip == "My shortlist":
        _starred_set = set(db.list_starred_brands())

    _chip_filtered = filtered
    if chip == "Proven 90+ days":
        _chip_filtered = [r for r in filtered if (r.get("days_running") or 0) >= 90]
    elif chip == "New winners (30+ days)":
        _chip_filtered = [
            r for r in filtered
            if (r.get("days_running") or 0) >= 30 and (r.get("days_running") or 0) < 90
            and r.get("is_active")
        ]
    elif chip == "Has sales data":
        _chip_filtered = [r for r in filtered if r.get("mp_sold") or r.get("mp_price")]
    elif chip == "My shortlist":
        _chip_filtered = [
            r for r in filtered
            if (r.get("brand") or r.get("page_name") or "") in _starred_set
        ]

    if chip != "All":
        st.caption(f"Chip filter **{chip}** active · {len(_chip_filtered):,} of {len(filtered):,} ads shown.")

    by_brand: dict[str, dict] = {}
    # Phase 14: track tag frequencies per brand so we surface the MOST-COMMON
    # category/sub_category/location, not just the first seen.
    _brand_tag_counts: dict[str, dict[str, Counter]] = defaultdict(
        lambda: {"category": Counter(), "sub_category": Counter(), "location": Counter()}
    )
    for r in _chip_filtered:
        b = (r.get("brand") or r.get("page_name") or "?").strip()
        entry = by_brand.setdefault(b, {
            "brand": b, "niche": r.get("niche"), "ad_count": 0,
            "total_score": 0.0, "score_normalized": 0.0,
            "max_days_running": 0, "any_active": False,
            "sample_ad_text": "", "sample_landing_url": "",
            "page_names": set(), "sample_creative": None,
            "mp_sold": 0, "mp_price": None, "mp_rating": None, "mp_reviews": 0,
            "mp_source": "",
            # Phase 14 — placeholders; resolved AFTER the loop via _brand_tag_counts
            "category": "", "sub_category": "", "location": "",
        })
        # Tally tag values across all this brand's ads
        if r.get("category"):
            _brand_tag_counts[b]["category"][r["category"]] += 1
        if r.get("sub_category"):
            _brand_tag_counts[b]["sub_category"][r["sub_category"]] += 1
        if r.get("location"):
            _brand_tag_counts[b]["location"][r["location"]] += 1
        entry["ad_count"] += 1
        entry["total_score"] += float(r.get("score") or 0)
        entry["score_normalized"] += float(r.get("score_normalized") or 0)
        entry["max_days_running"] = max(entry["max_days_running"], r.get("days_running") or 0)
        entry["any_active"] = entry["any_active"] or bool(r.get("is_active"))
        entry["page_names"].add(r.get("page_name") or "")
        if not entry["sample_ad_text"] and r.get("ad_text"):
            entry["sample_ad_text"] = (r.get("ad_text") or "")[:240]
        if not entry["sample_landing_url"] and r.get("landing_url"):
            entry["sample_landing_url"] = r.get("landing_url")
        if not entry["sample_creative"] and r.get("creative_path"):
            entry["sample_creative"] = _creative_path_url(r.get("creative_path"))

        sold = r.get("mp_sold") or 0
        if sold and sold > (entry["mp_sold"] or 0):
            entry["mp_sold"] = sold
            entry["mp_price"] = r.get("mp_price")
            entry["mp_rating"] = r.get("mp_rating")
            entry["mp_reviews"] = r.get("mp_reviews") or 0
            entry["mp_source"] = r.get("mp_source") or ""

    brand_rows = sorted(by_brand.values(), key=lambda x: -x["score_normalized"])
    for b in brand_rows:
        b["page_names"] = ", ".join(sorted(n for n in b["page_names"] if n))
        b["total_score"] = round(b["total_score"], 1)
        b["score_normalized"] = round(b["score_normalized"], 1)
        # Phase 14: resolve cat/sub/loc to most-common value across this brand's ads
        _bname = b["brand"]
        _tags = _brand_tag_counts.get(_bname, {})
        for _k in ("category", "sub_category", "location"):
            counter = _tags.get(_k)
            if counter:
                b[_k] = counter.most_common(1)[0][0]

    for b in brand_rows:
        b["has_creative"] = bool(b.get("sample_creative"))

    has_any_creative = any(b.get("has_creative") for b in brand_rows)
    has_any_marketplace = any(b.get("mp_sold") or b.get("mp_price") for b in brand_rows)

    df = pd.DataFrame(brand_rows)
    if df.empty:
        if not rows:
            st.warning(
                "This run has **0 ads** — likely a failed or cancelled scrape. "
                "Pick a different run from the sidebar (history is in **Settings → Run history**), "
                "or click **Run scrape** in the sidebar to start a fresh one."
            )
        else:
            ph_count = sum(1 for r in rows if r.get("geo_signal") == "ph-confident")
            on_niche_count = sum(1 for r in rows if r.get("niche_relevance") == "match")
            st.info(
                f"Filters hide everything in this run — {len(rows):,} ads exist but none match. "
                f"Try: lower **Min days** slider, or untick **PH-confident only** "
                f"({ph_count:,} PH-confident in this run) / **In-niche text only** ({on_niche_count:,} in-niche). "
                f"Sidebar →"
            )
    else:
        # ---- Pre-format columns for premium display ----
        _all_meta = db.all_brand_meta()
        _max_score = max((b.get("score_normalized") or 0) for b in brand_rows) or 1

        # Niche prefix — subtle bullet for visual texture
        _NICHE_PREFIX = {"capsule": "● capsule", "cream": "● cream", "oil": "● oil", "coffee": "● coffee"}
        df["niche"] = df["niche"].apply(lambda n: _NICHE_PREFIX.get(n, n or "—"))

        # Active as text status indicator (more scannable than checkbox)
        df["any_active"] = df["any_active"].apply(lambda x: "● Active" if x else "○ Inactive")

        # Creative column as text marker
        if "has_creative" in df.columns:
            df["has_creative"] = df["has_creative"].apply(lambda x: "●" if x else "")

        # Star column
        df["_starred"] = df["brand"].apply(lambda b: "★" if _all_meta.get(b, {}).get("starred") else "")

        # Pipeline status with bullet
        _STATUS_DISPLAY = {
            "researching": "● Researching", "validated": "● Validated",
            "sourced": "● Sourced", "decided": "● Decided",
        }
        df["_status"] = df["brand"].apply(
            lambda b: _STATUS_DISPLAY.get(_canon_status(_all_meta.get(b, {}).get("status", "")), "")
        )

        # Days running with suffix
        df["max_days_running"] = df["max_days_running"].astype(int)

        # Phase 14: pretty-format sub_category for display ("anti-aging" -> "Anti Aging")
        if "sub_category" in df.columns:
            df["sub_category"] = df["sub_category"].apply(
                lambda s: s.replace("-", " ").replace("_", " ").title() if s else ""
            )

        # ---- Trendline sparkline (Phase 14.3) ----
        # Per-brand active-ad-count time series across last 12 META runs.
        _brands_for_tl = tuple(df["brand"].tolist())
        _tl_map = _brands_trendlines(_brands_for_tl, n_recent_runs=12)
        df["trendline"] = df["brand"].apply(lambda b: _tl_map.get(b.lower(), []))

        # Phase 14: compute a shared y_max so sparklines compare fairly across brands
        _all_tl_values = [v for series in _tl_map.values() for v in series]
        _tl_y_max = max(_all_tl_values) if _all_tl_values else 1

        col_cfg = {
            "_starred":     st.column_config.TextColumn("★", width="small",
                              help="★ = on your shortlist. Star brands from the detail modal (click any row)."),
            "brand":        st.column_config.TextColumn("Brand", width="medium",
                              help="Clustered brand name — multiple sister pages may roll up into one brand"),
            "niche":        st.column_config.TextColumn("Niche", width="small",
                              help="Niche category — what type of product this is"),
            "_status":      st.column_config.TextColumn("Pipeline", width="small",
                              help="Your research pipeline stage (set in detail modal): investigating → validating → launching → launched / passed"),
            "ad_count":     st.column_config.NumberColumn("Ads", format="%d", width="small",
                              help="How many ad variants this brand is running"),
            "score_normalized": st.column_config.ProgressColumn(
                              "Score", format="%d", min_value=0, max_value=float(_max_score),
                              width="medium",
                              help="Niche-normalized score (days × active × variants, capped at niche P95). Bar fill shows relative magnitude."),
            "max_days_running": st.column_config.NumberColumn("Days", format="%d d", width="small",
                              help="Longest-running ad from this brand. >90 days = proven; >180 days = serious."),
            "any_active":   st.column_config.TextColumn("Status", width="small",
                              help="● Active = at least one variant still running. ○ Inactive = all ads stopped."),
            "has_creative": st.column_config.TextColumn("Img", width="small",
                              help="● = creative image saved (browse in Insights → Creative Gallery)"),
            "mp_sold":      st.column_config.NumberColumn("Sold", format="%d", width="small",
                              help="Shopee/Lazada historical units sold (requires marketplace enrichment)"),
            "mp_price":     st.column_config.NumberColumn("Price", format="₱%.0f", width="small",
                              help="Marketplace product price in PHP"),
            "mp_rating":    st.column_config.NumberColumn("Rating", format="%.1f ★", width="small",
                              help="Marketplace customer rating (out of 5)"),
            "mp_reviews":   st.column_config.NumberColumn("Reviews", format="%d", width="small",
                              help="Number of marketplace reviews"),
            "sample_landing_url": st.column_config.LinkColumn(
                              "Landing", width="medium",
                              display_text=r"^https?://(?:www\.)?(.{1,55})",
                              help="Click to open the brand's landing page (Shopee/Lazada/own funnel)"),
            "sample_ad_text": st.column_config.TextColumn("Sample ad text", width="large",
                              help="First chunk of one ad from this brand — gives you a sense of their hook"),
            "page_names":   st.column_config.TextColumn("Pages clustered", width="medium",
                              help="All Facebook page names rolled up into this single brand"),
            # Phase 14 — Category / Sub-category / Location / Trendline
            "category":     st.column_config.TextColumn("Category", width="small",
                              help="Top-level intent — Health, Beauty, F&B, Wellness."),
            "sub_category": st.column_config.TextColumn("Sub-cat", width="small",
                              help="Specific angle (slimming, whitening, anti-aging, etc.)"),
            "location":     st.column_config.TextColumn("Location", width="small",
                              help="PH city/region detected from ad text + landing URL."),
            "trendline":    st.column_config.LineChartColumn(
                              "Trend", width="medium", y_min=0, y_max=float(_tl_y_max),
                              help="Active ad count over last 12 runs (shared scale). "
                                   "Rising = scaling; falling = retiring."),
        }

        # Build column order — keep the table compact. Category replaces niche in default
        # display (niche is the product form like "capsule", category is the intent — the
        # intent reads better). User can still see niche by sorting df externally.
        # Only show columns that have at least some non-empty values to avoid blank cols.
        _has_any_cat = "category" in df.columns and df["category"].astype(str).str.strip().ne("").any()
        _has_any_loc = "location" in df.columns and df["location"].astype(str).str.strip().ne("").any()

        order = ["_starred", "brand"]
        # Category column (replaces niche when classifier has data)
        if _has_any_cat:
            order += ["category", "sub_category"]
        else:
            order.append("niche")
        if _has_any_loc:
            order.append("location")
        order += ["_status", "ad_count", "score_normalized",
                  "max_days_running", "any_active", "trendline"]
        if has_any_creative:
            order.append("has_creative")
        if has_any_marketplace:
            order += ["mp_sold", "mp_price", "mp_rating", "mp_reviews"]
        order += ["sample_landing_url", "sample_ad_text"]
        order = [c for c in order if c in df.columns]

        # ---- View toggle: Table vs Cards (Phase 10.3) ----
        _vt_l, _vt_r = st.columns([5, 1])
        with _vt_r:
            _view_mode = st.segmented_control(
                "View",
                options=["Table", "Cards"],
                default=st.session_state.get("top_brands_view", "Table"),
                key="top_brands_view_picker",
                label_visibility="collapsed",
            )
            if _view_mode and _view_mode != st.session_state.get("top_brands_view"):
                st.session_state["top_brands_view"] = _view_mode

        _active_view = st.session_state.get("top_brands_view", "Table")

        if _active_view == "Cards":
            import html as _h  # escape scraped values before unsafe_allow_html
            # ---- Card grid view ----
            st.caption(
                "Visual-first card view. Click any card to open the brand detail panel."
            )
            _card_brands = brand_rows[:30]
            _ncols = 3
            for _row_start in range(0, len(_card_brands), _ncols):
                _row_brands = _card_brands[_row_start:_row_start + _ncols]
                _ccols = st.columns(_ncols)
                for _ci, _b in enumerate(_row_brands):
                    with _ccols[_ci]:
                        _bname = _b.get("brand") or "—"
                        _bniche = (_b.get("niche") or "—").replace("● ", "")
                        _bscore = int(_b.get("score_normalized") or 0)
                        _bdays = int(_b.get("max_days_running") or 0)
                        _bads = int(_b.get("ad_count") or 0)
                        _bactive = _b.get("any_active")
                        # Phase 14: surface category + sub-cat + location on the card
                        _bcat = _b.get("category") or ""
                        _bsub = _b.get("sub_category") or ""
                        _bsub_pretty = _bsub.replace("-", " ").replace("_", " ").title() if _bsub else ""
                        _bloc = _b.get("location") or ""
                        _bcat_tag_html = ""
                        if _bcat:
                            _bcat_tag_html = (
                                f"<div style='color:var(--pra-accent);font-size:0.7rem;"
                                f"letter-spacing:0.06em;margin-bottom:4px'>"
                                f"{_h.escape(_bcat)}"
                                + (f" › <strong>{_h.escape(_bsub_pretty)}</strong>" if _bsub_pretty else "")
                                + (f" · 📍 {_h.escape(_bloc)}" if _bloc else "")
                                + "</div>"
                            )
                        _is_starred = bool(_all_meta.get(_bname, {}).get("starred"))
                        _star_html = "<span style='color:var(--pra-accent);margin-right:6px'>★</span>" if _is_starred else ""
                        _active_html = (
                            "<span style='color:var(--pra-success);font-size:0.72rem;"
                            "font-weight:600;letter-spacing:0.08em'>● ACTIVE</span>"
                            if _bactive else
                            "<span style='color:var(--pra-text-dim);font-size:0.72rem;"
                            "font-weight:600;letter-spacing:0.08em'>○ INACTIVE</span>"
                        )
                        _score_pct = int((_bscore / _max_score) * 100) if _max_score else 0
                        # The card visual + button live in the same st.container so they
                        # read as ONE interactive unit. The button uses the .pra-card-btn
                        # CSS class (added in stylesheet) for minimal flush styling.
                        st.markdown(
                            f"<div class='pra-brand-card'>"
                            f"<div style='display:flex;justify-content:space-between;align-items:baseline;"
                            f"margin-bottom:6px'>"
                            f"<div style='font-size:0.95rem;font-weight:600;color:var(--pra-text);"
                            f"letter-spacing:-0.005em'>{_star_html}{_h.escape(_bname)}</div>"
                            f"<div style='display:flex;align-items:center;gap:8px'>"
                            f"<span style='color:var(--pra-text-muted);font-size:0.72rem;"
                            f"font-variant-numeric:tabular-nums'>{_bdays:,}d</span>"
                            f"<span style='color:var(--pra-accent);font-size:0.85rem'>→</span>"
                            f"</div>"
                            f"</div>"
                            f"{_bcat_tag_html}"
                            f"<div style='color:var(--pra-text-muted);font-size:0.78rem;"
                            f"margin-bottom:10px'>"
                            f"{_h.escape(_bniche)} · {_bads} variant{'s' if _bads != 1 else ''} · {_active_html}"
                            f"</div>"
                            f"<div style='display:flex;align-items:center;gap:8px'>"
                            f"<div style='flex:1;background:var(--pra-border);border-radius:2px;height:5px;"
                            f"overflow:hidden'>"
                            f"<div style='background:var(--pra-accent);width:{_score_pct}%;height:100%'></div>"
                            f"</div>"
                            f"<span style='color:var(--pra-text-muted);font-size:0.72rem;"
                            f"font-variant-numeric:tabular-nums'>score {_bscore:,}</span>"
                            f"</div>"
                            f"<div class='pra-brand-card-foot-marker'></div>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )
                        # Click target (styled via CSS as a thin transparent attached strip)
                        if st.button(
                            "View brand",
                            key=f"card_open_{_bname}_{_row_start}_{_ci}",
                            width="stretch",
                            type="secondary",
                        ):
                            _ads_for_brand = [r for r in rows if (r.get("brand") or r.get("page_name")) == _bname]
                            _render_brand_detail_dialog(_bname, _b, _ads_for_brand)
            event = None  # no table-event in card mode
        else:
            event = st.dataframe(
                df[order], column_config=col_cfg, width="stretch", hide_index=True, height=620,
                on_select="rerun", selection_mode="single-row", key="top_brands_table",
            )
            _cap_l, _cap_r = st.columns([5, 1])
            with _cap_l:
                st.caption(
                    "Sorted by **Score** (niche-normalized). "
                    "**Click any row** to open the brand detail panel — star, set pipeline status, add notes. "
                    "Click column headers to re-sort."
                )
            with _cap_r:
                # Phase 12.3 — Print-friendly trigger (JS opens browser print dialog)
                st.markdown(
                    """
                    <button onclick="window.parent.print()"
                            style="width:100%;padding:6px 10px;background:transparent;
                                   border:1px solid var(--pra-border);border-radius:4px;
                                   color:var(--pra-text-muted);font-size:0.78rem;
                                   cursor:pointer;font-weight:500"
                            onmouseover="this.style.borderColor='var(--pra-accent)';
                                         this.style.color='var(--pra-text)'"
                            onmouseout="this.style.borderColor='var(--pra-border)';
                                        this.style.color='var(--pra-text-muted)'">
                        ⎙ Print view
                    </button>
                    """,
                    unsafe_allow_html=True,
                )

            if event and getattr(event, "selection", None) and event.selection.get("rows"):
                _sel_idx = event.selection["rows"][0]
                _sel_brand = df.iloc[_sel_idx]["brand"]
                _ads_for_brand = [r for r in rows if (r.get("brand") or r.get("page_name")) == _sel_brand]
                _render_brand_detail_dialog(_sel_brand, df.iloc[_sel_idx].to_dict(), _ads_for_brand)

        # ---- Bulk operations on visible brands ----
        st.divider()
        with st.expander("⚡ Bulk operations", expanded=False):
            _bulk_brands = st.multiselect(
                "Select brands",
                sorted(df["brand"].tolist()),
                key="bulk_brand_select",
                help="Pick multiple brands then apply an action below",
            )
            if _bulk_brands:
                bc1, bc2, bc3 = st.columns(3)
                if bc1.button("⭐ Star all", key="bulk_star", width="stretch",
                                help="Add all selected brands to your shortlist"):
                    for _b in _bulk_brands:
                        db.upsert_brand_meta(_b, starred=1)
                    st.success(f"Starred {len(_bulk_brands)} brand{'s' if len(_bulk_brands) != 1 else ''}.")
                    st.rerun()
                if bc2.button("☆ Unstar all", key="bulk_unstar", width="stretch",
                                help="Remove all selected brands from shortlist"):
                    for _b in _bulk_brands:
                        db.upsert_brand_meta(_b, starred=0)
                    st.success(f"Unstarred {len(_bulk_brands)} brand{'s' if len(_bulk_brands) != 1 else ''}.")
                    st.rerun()
                _bulk_status_options = [""] + _HUNT_PIPELINE
                _bulk_new_status = bc3.selectbox(
                    "Set pipeline status",
                    _bulk_status_options,
                    key="bulk_status_select",
                    format_func=lambda x: "Pick stage…" if x == "" else _STATUS_LABEL.get(x, x.capitalize()),
                    label_visibility="collapsed",
                )
                if _bulk_new_status and st.button(
                    f"Set status: '{_bulk_new_status}'",
                    key="bulk_status_apply",
                    type="primary",
                    help="Apply the selected status to all selected brands",
                ):
                    for _b in _bulk_brands:
                        db.upsert_brand_meta(_b, status=_bulk_new_status)
                    st.success(f"Set status '{_bulk_new_status}' for {len(_bulk_brands)} brand{'s' if len(_bulk_brands) != 1 else ''}.")
                    st.rerun()

        # Export current filtered view to xlsx
        st.divider()
        col_dl1, col_dl2, col_dl3 = st.columns([2, 1, 4])
        col_dl1.markdown("**Export current view**")
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter

        buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Brands"
        # Use the table-as-displayed, but drop columns that hold non-scalar values
        # (e.g. the trendline sparkline is a list of ints — can't write to xlsx cells).
        _export_cols = [c for c in order if c != "trendline"]
        display_df = df[_export_cols].copy()
        ws.append(list(display_df.columns))
        def _safe_cell(v):
            # Lists / arrays / None all become empty string in the export
            if isinstance(v, (list, tuple)):
                return ""
            try:
                if pd.isna(v):
                    return ""
            except (TypeError, ValueError):
                return str(v) if v is not None else ""
            return v
        for _, row in display_df.iterrows():
            ws.append([_safe_cell(v) for v in row.tolist()])
        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(bold=True, color="FFFFFF")
        for i in range(1, len(display_df.columns) + 1):
            c = ws.cell(row=1, column=i)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        last = len(display_df) + 1
        grad = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
        for cname in ("score_normalized", "total_score", "max_days_running", "mp_sold"):
            if cname in display_df.columns:
                L = get_column_letter(list(display_df.columns).index(cname) + 1)
                ws.conditional_formatting.add(f"{L}2:{L}{last}", grad)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(buf)
        buf.seek(0)
        col_dl2.download_button(
            "Download .xlsx",
            data=buf.getvalue(),
            file_name=f"brands_filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )


# ---- Suggestions ----

_SUGG_LABEL = {
    "🏆": ("Proven winner", "#D4AF37"),
    "🎯": ("Heavy tester",  "#8FC2B4"),
    "🔥": ("New winner",    "#E6CC73"),
    "💀": ("Retired",       "#8B6B6B"),
    "📊": ("Niche heat",    "#9F9FE0"),
    "🌱": ("Opportunity",   "#7CC4A0"),
    "🧭": ("Traffic mix",   "#8FC2B4"),
}

if current_page == "fb_ads":
    st.subheader("What to do today")
    st.caption("Rule-based recommendations across the database — auto-prioritised.")
    sugg_list = suggestions.generate(top_n=20)
    if not sugg_list:
        st.info("Not enough data yet. Run more scrapes to enable trend-based suggestions.")
    else:
        for s in sugg_list:
            label, color = _SUGG_LABEL.get(s.icon, ("Insight", "#8FC2B4"))
            with st.container(border=True):
                st.markdown(
                    f"<div style='color:{color};font-size:0.7rem;letter-spacing:0.12em;"
                    f"font-weight:600;text-transform:uppercase;margin-bottom:6px'>{label}</div>",
                    unsafe_allow_html=True,
                )
                st.markdown(f"##### {s.title}")
                st.write(s.detail)
                if s.related:
                    r = s.related[0]
                    link = r.get("landing_url")
                    if link:
                        st.link_button("Open landing page", link)


# ---- New Winners ----

if current_page == "fb_ads":
  with st.expander("New 30-day winners since previous run", expanded=False):
    st.subheader("Ads that just crossed 30-day longevity")
    prev_runs = get_runs(limit=2)
    if len(prev_runs) < 2:
        st.info("Need at least 2 runs to compute. Run another scrape to populate.")
    else:
        prev = get_ads(prev_runs[1]["run_id"])
        new_winners = suggestions.new_winners_since(prev, rows, threshold_days=30)
        if not new_winners:
            st.info("No new 30-day winners since the previous run.")
        for s in new_winners:
            with st.container(border=True):
                st.markdown(f"<div style='color:#D4AF37;font-size:0.72rem;letter-spacing:0.1em;font-weight:600;text-transform:uppercase'>New winner</div>", unsafe_allow_html=True)
                st.markdown(f"##### {s.title}")
                st.write(s.detail)


# ---- Retired ----

if current_page == "fb_ads":
  with st.expander("Recently retired ads", expanded=False):
    st.subheader("Ads that disappeared since the previous run")
    prev_runs = get_runs(limit=2)
    if len(prev_runs) < 2:
        st.info("Need at least 2 runs to compute.")
    else:
        prev = get_ads(prev_runs[1]["run_id"])
        retired = suggestions.retired_ads(prev, rows)
        if not retired:
            st.info("No notable retirements since the previous run.")
        for s in retired:
            with st.container(border=True):
                st.markdown(f"<div style='color:#8B6B6B;font-size:0.72rem;letter-spacing:0.1em;font-weight:600;text-transform:uppercase'>Retired</div>", unsafe_allow_html=True)
                st.markdown(f"##### {s.title}")
                st.write(s.detail)


# ---- Creative Gallery ----

if current_page == "gallery":
    st.subheader("Creative gallery — winners by score")
    st.caption("Saved ad creatives — images shown inline, video files marked with a 🎬 badge (open the file to view).")
    creative_rows = [r for r in filtered if r.get("creative_path") and _creative_path_url(r.get("creative_path"))]
    creative_rows.sort(key=lambda x: -(x.get("score") or 0))
    if not creative_rows:
        st.info("No creatives saved yet. Run `python main.py` to download them (or use the sidebar button).")
    else:
        per_row = 4
        for i in range(0, min(len(creative_rows), 60), per_row):
            cols = st.columns(per_row)
            for j, r in enumerate(creative_rows[i:i + per_row]):
                with cols[j]:
                    media = _creative_path_url(r.get("creative_path"))
                    if media:
                        if media.lower().endswith((".mp4", ".webm", ".mov")):
                            # Video creative — use st.video
                            try:
                                st.video(media)
                            except Exception:
                                st.markdown(
                                    f"<div style='background:var(--pra-panel);"
                                    f"border:1px solid var(--pra-border);border-radius:4px;"
                                    f"padding:16px;text-align:center'>"
                                    f"<div style='font-size:2rem'>🎬</div>"
                                    f"<div style='color:var(--pra-text-muted);font-size:0.78rem'>Video file</div>"
                                    f"</div>",
                                    unsafe_allow_html=True,
                                )
                        else:
                            st.image(media, width="stretch")
                    st.caption(
                        f"**{r.get('brand') or r.get('page_name')}** · {r.get('niche')}\n\n"
                        f"{r.get('days_running')}d · score {round(r.get('score') or 0)} · "
                        f"{'active' if r.get('is_active') else 'inactive'}"
                    )
                    if r.get("landing_url"):
                        st.markdown(f"[Landing →]({r['landing_url']})")
                    # Phase 20.1 — Save to swipe file
                    if st.button("✂ Save to swipe", key=f"swipe_save_{i}_{j}",
                                 width="stretch", type="secondary"):
                        st.session_state["_swipe_pending"] = {
                            "brand": r.get("brand") or r.get("page_name") or "",
                            "niche": r.get("niche") or "",
                            "creative_path": r.get("creative_path") or "",
                            "landing_url": r.get("landing_url") or "",
                            "ad_text": r.get("ad_text") or "",
                        }
                        st.rerun()

    # Swipe-save target picker (appears when a creative is being saved)
    if st.session_state.get("_swipe_pending"):
        _pend = st.session_state["_swipe_pending"]

        @st.dialog("Save to swipe file")
        def _swipe_save_dialog():
            st.markdown(f"**{_pend['brand']}** · {_pend['niche']}")
            _colls = db.list_swipe_collections()
            _coll_names = [c["name"] for c in _colls]
            _choice = st.selectbox(
                "Add to collection",
                ["+ New collection…"] + _coll_names,
            )
            if _choice == "+ New collection…":
                _new_name = st.text_input("New collection name",
                                          placeholder="e.g. Whitening winners")
            _note = st.text_input("Note (optional)",
                                  placeholder="Why you saved this")
            if st.button("Save", type="primary", width="stretch"):
                if _choice == "+ New collection…":
                    if not _new_name.strip():
                        st.error("Enter a collection name.")
                        return
                    _cid = db.create_swipe_collection(_new_name.strip())
                else:
                    _cid = next(c["id"] for c in _colls if c["name"] == _choice)
                db.add_swipe_item(_cid, {**_pend, "notes": _note})
                st.session_state.pop("_swipe_pending", None)
                st.toast("✂ Saved to swipe file", icon=None)
                st.rerun()
            if st.button("Cancel", width="stretch"):
                st.session_state.pop("_swipe_pending", None)
                st.rerun()

        _swipe_save_dialog()


# ---- Swipe File (Phase 20.1) ----

if current_page == "swipe":
    st.subheader("Swipe file · saved winning creatives")
    st.caption(
        "Your curated collections of ad creatives worth referencing. "
        "Save from the Creative Gallery with the ✂ button."
    )

    _sf_collections = db.list_swipe_collections()

    # New collection inline
    _sf_l, _sf_r = st.columns([3, 1])
    with _sf_l:
        _sf_new = st.text_input(
            "New collection name", placeholder="e.g. Slimming hooks, Whitening winners",
            key="sf_new_coll", label_visibility="collapsed",
        )
    with _sf_r:
        if st.button("+ Create collection", width="stretch", type="primary",
                     disabled=not _sf_new.strip()):
            db.create_swipe_collection(_sf_new.strip())
            st.toast(f"✓ Created '{_sf_new.strip()}'", icon=None)
            st.rerun()

    if not _sf_collections:
        st.info(
            "No swipe collections yet. Create one above, then save creatives from the "
            "**Creative Gallery** page using the ✂ button on each."
        )
    else:
        for _coll in _sf_collections:
            with st.expander(
                f"✂ {_coll['name']} · {_coll['item_count']} item"
                f"{'s' if _coll['item_count'] != 1 else ''}",
                expanded=False,
            ):
                _items = db.list_swipe_items(_coll["id"])
                if not _items:
                    st.caption("Empty — save creatives from the Gallery.")
                else:
                    _sf_per_row = 4
                    for _si in range(0, len(_items), _sf_per_row):
                        _scols = st.columns(_sf_per_row)
                        for _sj, _item in enumerate(_items[_si:_si + _sf_per_row]):
                            with _scols[_sj]:
                                _media = _creative_path_url(_item.get("creative_path"))
                                if _media and not _media.lower().endswith((".mp4", ".webm", ".mov")):
                                    st.image(_media, width="stretch")
                                elif _media:
                                    st.markdown(
                                        "<div style='background:var(--pra-panel);"
                                        "border:1px solid var(--pra-border);border-radius:4px;"
                                        "padding:16px;text-align:center'>🎬</div>",
                                        unsafe_allow_html=True,
                                    )
                                st.caption(
                                    f"**{_item.get('brand')}** · {_item.get('niche')}"
                                    + (f"\n\n_{_item['notes']}_" if _item.get("notes") else "")
                                )
                                if _item.get("landing_url"):
                                    st.markdown(f"[Landing →]({_item['landing_url']})")
                                if st.button("✕ Remove", key=f"sf_del_{_item['id']}",
                                             width="stretch"):
                                    db.delete_swipe_item(_item["id"])
                                    st.rerun()
                st.divider()
                if st.button(f"🗑 Delete collection '{_coll['name']}'",
                             key=f"sf_del_coll_{_coll['id']}"):
                    db.delete_swipe_collection(_coll["id"])
                    st.toast(f"✕ Deleted '{_coll['name']}'", icon=None)
                    st.rerun()


# ---- Trends ----

if current_page == "trends":
    st.subheader("Trends across runs")
    all_runs = get_runs(limit=30)
    if len(all_runs) < 2:
        st.info("Need at least 2 runs for trends. Schedule daily runs via `register_daily_task.ps1`.")
    else:
        per_run_niche: list[dict] = []
        for r in all_runs:
            ads_in = db.ads_for_run(r["run_id"])
            niches_c = Counter((a.get("niche") or "?") for a in ads_in if a.get("is_active"))
            for niche, c in niches_c.items():
                per_run_niche.append({
                    "run": f"#{r['run_id']} {r['started_at'][:10]}",
                    "started_at": r["started_at"],
                    "niche": niche,
                    "active_ads": c,
                })
        if per_run_niche:
            df_trend = pd.DataFrame(per_run_niche)
            pivot = df_trend.pivot_table(
                index="run", columns="niche", values="active_ads", aggfunc="sum", fill_value=0,
            ).sort_index()
            st.line_chart(pivot)
            st.caption("Active ads per niche, per run. Rising = competition heating up; falling = market cooling or scraper degrading.")


# ---- Hook patterns (NLP n-gram analysis of winning ad copy) ----

# ============================================================
# Phase 20.3 — Weekly Digest (top winners summary, exportable)
# ============================================================

if current_page == "digest":
    st.subheader("Weekly digest · this run's winners at a glance")
    st.caption(
        "A shareable summary of the top winners, new entrants, retirements, hottest "
        "niche, and FDA-risky brands to avoid. Export as text for your team."
    )

    # Compute the digest from the current filtered rows
    from collections import defaultdict as _dd_dig
    _dig_brand_agg: dict[str, dict] = {}
    for r in filtered:
        if not r.get("is_active"):
            continue
        if r.get("geo_signal") not in ("ph-confident", "ph-likely"):
            continue
        if r.get("niche_relevance") == "no_match":
            continue
        b = (r.get("brand") or r.get("page_name") or "").strip()
        if not b:
            continue
        e = _dig_brand_agg.setdefault(b, {
            "brand": b, "niche": r.get("niche") or "—", "score": 0.0,
            "max_days": 0, "ads": 0, "landing": "",
        })
        e["score"] += float(r.get("score_normalized") or 0)
        e["max_days"] = max(e["max_days"], r.get("days_running") or 0)
        e["ads"] += 1
        if not e["landing"] and r.get("landing_url"):
            e["landing"] = r["landing_url"]

    _top10 = sorted(_dig_brand_agg.values(), key=lambda x: -x["score"])[:10]

    # Hot niche
    _niche_brands = _dd_dig(set)
    for b, e in _dig_brand_agg.items():
        _niche_brands[e["niche"]].add(b)
    _hot_niche = max(_niche_brands.items(), key=lambda x: len(x[1])) if _niche_brands else ("—", set())

    # FDA-risky brands
    _risky = []
    try:
        import fda_compliance
        for b, e in list(_dig_brand_agg.items())[:40]:
            _b_ads = [r for r in filtered if (r.get("brand") or r.get("page_name")) == b]
            _fr = fda_compliance.scan_brand(_b_ads)
            if _fr.get("worst_severity") in ("critical", "high"):
                _risky.append((b, _fr["worst_severity"]))
    except Exception:
        pass

    _run_label = next((f"Run #{r['run_id']} · {r['started_at'][:10]}"
                       for r in runs if r["run_id"] == active_run_id), "current run")

    # --- Render the digest ---
    st.markdown(
        f"<div style='background:linear-gradient(135deg,rgba(212,175,55,0.06) 0%,"
        f"rgba(212,175,55,0.015) 100%);border:1px solid var(--pra-border);"
        f"border-left:2px solid var(--pra-accent);border-radius:6px;padding:18px 22px;"
        f"margin-bottom:18px'>"
        f"<div style='color:var(--pra-accent);font-size:0.7rem;letter-spacing:0.18em;"
        f"font-weight:700;text-transform:uppercase;margin-bottom:6px'>Orbit Digest</div>"
        f"<div style='color:var(--pra-text);font-size:1.05rem;font-weight:600'>{_run_label}</div>"
        f"<div style='color:var(--pra-text-muted);font-size:0.82rem;margin-top:4px'>"
        f"{len(_dig_brand_agg)} active PH brands · hottest niche: "
        f"<strong style='color:var(--pra-text)'>{_hot_niche[0]}</strong> "
        f"({len(_hot_niche[1])} brands)</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # Top 10 table
    st.markdown("##### 🏆 Top 10 winners")
    if _top10:
        _dig_html = (
            "<table class='pra-table'><thead><tr>"
            "<th>#</th><th>Brand</th><th>Niche</th>"
            "<th class='num'>Days</th><th class='num'>Ads</th><th class='num'>Score</th>"
            "</tr></thead><tbody>"
        )
        for _i, _b in enumerate(_top10, 1):
            _dig_html += (
                f"<tr><td class='num'>{_i}</td>"
                f"<td class='brand-cell'>{_b['brand']}</td>"
                f"<td>{_b['niche']}</td>"
                f"<td class='num'>{_b['max_days']:,}d</td>"
                f"<td class='num'>{_b['ads']}</td>"
                f"<td class='num'>{int(_b['score']):,}</td></tr>"
            )
        _dig_html += "</tbody></table>"
        st.markdown(_dig_html, unsafe_allow_html=True)
    else:
        st.caption("No qualifying brands in the current filter.")

    # FDA-risky callout
    if _risky:
        st.markdown("##### ⚠ FDA-risky brands (don't copy their hooks blindly)")
        _risk_html = "<div style='background:rgba(224,127,159,0.05);border:1px solid var(--pra-border);border-left:2px solid var(--pra-danger);border-radius:4px;padding:10px 14px'>"
        for _b, _sev in _risky[:8]:
            _risk_html += (
                f"<div style='font-size:0.85rem;color:var(--pra-text);padding:3px 0'>"
                f"<strong>{_b}</strong> — <span style='color:var(--pra-danger)'>{_sev}</span> claim risk</div>"
            )
        _risk_html += "</div>"
        st.markdown(_risk_html, unsafe_allow_html=True)

    # --- Export as text ---
    st.divider()
    _digest_lines = [
        f"ORBIT WEEKLY DIGEST — {_run_label}",
        "=" * 50,
        f"{len(_dig_brand_agg)} active PH brands · hottest niche: {_hot_niche[0]} ({len(_hot_niche[1])} brands)",
        "",
        "TOP 10 WINNERS:",
    ]
    for _i, _b in enumerate(_top10, 1):
        _digest_lines.append(
            f"  {_i}. {_b['brand']} ({_b['niche']}) — {_b['max_days']}d, {_b['ads']} ads, score {int(_b['score'])}"
        )
    if _risky:
        _digest_lines += ["", "FDA-RISKY BRANDS (avoid copying claims):"]
        for _b, _sev in _risky[:8]:
            _digest_lines.append(f"  - {_b}: {_sev} risk")
    _digest_text = "\n".join(_digest_lines)

    _dg_a, _dg_b = st.columns([3, 1])
    _dg_a.caption("Copy or download this digest to share with your team.")
    _dg_b.download_button(
        "📥 Download digest",
        data=_digest_text,
        file_name=f"orbit_digest_{datetime.now().strftime('%Y%m%d')}.txt",
        mime="text/plain", width="stretch",
    )
    with st.expander("Show plain-text digest", expanded=False):
        st.code(_digest_text, language=None)


if current_page == "hooks":
    st.subheader("Hook patterns · what phrases win in PH")
    st.caption(
        "Recurring 2–4 word phrases across proven (14+ day) PH-confident, in-niche ads. "
        "Phrases used by 2+ different brands — these are battle-tested copy patterns you can adapt."
    )

    import hook_analyzer as _hooks

    _hc1, _hc2, _hc3 = st.columns([2, 1, 1])
    _hook_niche = _hc1.selectbox(
        "Niche filter",
        ["all"] + list(load_config().get("niches", {}).keys()),
        index=0,
        key="hooks_niche",
        format_func=lambda x: "All niches" if x == "all" else x.replace("_", " "),
    )
    _hook_ngram_max = _hc2.selectbox(
        "Phrase length",
        ["2", "3", "4", "2-3", "2-4"],
        index=4,
        key="hooks_ngram",
        help="2 = pairs (e.g. 'kili kili'). Longer = full hooks.",
    )
    _hook_min_count = _hc3.number_input(
        "Min usage", min_value=2, max_value=20, value=3, step=1,
        key="hooks_min_count",
        help="Only show phrases used at least N times.",
    )

    if "-" in _hook_ngram_max:
        _nmin, _nmax = _hook_ngram_max.split("-")
        _n_range = (int(_nmin), int(_nmax))
    else:
        _n = int(_hook_ngram_max)
        _n_range = (_n, _n)

    _hook_phrases = _hooks.extract_phrases(
        rows,
        n_range=_n_range,
        top_n=40,
        min_count=int(_hook_min_count),
        niche=(None if _hook_niche == "all" else _hook_niche),
    )

    if not _hook_phrases:
        st.info(
            "No recurring phrases found with these filters. "
            "Try lowering Min usage or widening the niche."
        )
    else:
        # Custom HTML table — theme-aware
        import html as _h
        _ph_rows = []
        _max_cnt = max(p["count"] for p in _hook_phrases) or 1
        for p in _hook_phrases:
            _pct = int((p["count"] / _max_cnt) * 100)
            _sample = (p.get("sample_text") or "").replace("\n", " ")
            _brands_str = ", ".join(p.get("sample_brands", []))
            _ph_rows.append(
                "<tr>"
                f'<td class="brand-cell">{_h.escape(p["phrase"])}</td>'
                f'<td>'
                f'<div class="pra-score-cell">'
                f'<div class="pra-bar-track">'
                f'<div class="pra-bar-fill" style="width:{_pct}%"></div>'
                f'</div>'
                f'<span class="pra-score-num">{p["count"]}×</span>'
                f'</div>'
                f'</td>'
                f'<td class="num">{p["brands"]}</td>'
                f'<td>{_h.escape(p["top_niche"])}</td>'
                f'<td style="color:var(--pra-text-muted);font-size:0.78rem;font-style:italic">"{_h.escape(_sample[:140])}"</td>'
                f'<td style="color:var(--pra-text-muted);font-size:0.78rem">{_h.escape(_brands_str)}</td>'
                "</tr>"
            )
        _table_html = (
            "<div style='max-height:560px;overflow-y:auto'>"
            "<table class='pra-table'>"
            "<thead><tr>"
            "<th>Phrase</th><th>Frequency</th><th class='num'>Brands</th>"
            "<th>Top niche</th><th>Sample context</th><th>Used by</th>"
            "</tr></thead>"
            f"<tbody>{''.join(_ph_rows)}</tbody>"
            "</table>"
            "</div>"
        )
        st.markdown(_table_html, unsafe_allow_html=True)
        st.caption(
            f"Showing **{len(_hook_phrases)}** phrases · sorted by frequency. "
            "Use these as ad-copy starters — they've already been proven by 2+ PH brands."
        )

    # ---- Hook ↔ longevity correlation ----
    st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
    st.markdown("##### Hook ↔ longevity correlation")
    st.caption(
        "Side-by-side: which phrases appear in **evergreen** (90+ days) winners "
        "vs **losing** (stopped <30 days) ads. The contrast reveals what copy works."
    )

    _bucket_data = _hooks.extract_phrases_by_longevity(
        rows,
        n_range=(2, 3),
        top_n=15,
        min_count=2,
        niche=(None if _hook_niche == "all" else _hook_niche),
    )
    _bs = _bucket_data["_bucket_sizes"]

    bcol1, bcol2, bcol3 = st.columns(3)
    with bcol1:
        st.markdown(
            f"<div style='color:#7CC4A0;font-size:0.7rem;letter-spacing:0.1em;"
            f"font-weight:600;text-transform:uppercase;margin-bottom:6px'>"
            f"Evergreen (90+ days) · {_bs['evergreen']} ads"
            f"</div>",
            unsafe_allow_html=True,
        )
        if not _bucket_data["evergreen"]:
            st.caption("No evergreen winners found with current filters.")
        else:
            for p in _bucket_data["evergreen"][:10]:
                st.markdown(
                    f"<div style='padding:4px 0;border-bottom:1px solid var(--pra-border)'>"
                    f"<span style='font-weight:500'>{p['phrase']}</span> "
                    f"<span style='color:var(--pra-text-muted);font-size:0.78rem'>"
                    f"× {p['count']} · avg {p['avg_days']:.0f}d</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
    with bcol2:
        st.markdown(
            f"<div style='color:#E6CC73;font-size:0.7rem;letter-spacing:0.1em;"
            f"font-weight:600;text-transform:uppercase;margin-bottom:6px'>"
            f"Proven (30-89 days) · {_bs['proven']} ads"
            f"</div>",
            unsafe_allow_html=True,
        )
        if not _bucket_data["proven"]:
            st.caption("No mid-range proven ads with current filters.")
        else:
            for p in _bucket_data["proven"][:10]:
                st.markdown(
                    f"<div style='padding:4px 0;border-bottom:1px solid var(--pra-border)'>"
                    f"<span style='font-weight:500'>{p['phrase']}</span> "
                    f"<span style='color:var(--pra-text-muted);font-size:0.78rem'>"
                    f"× {p['count']} · avg {p['avg_days']:.0f}d</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
    with bcol3:
        st.markdown(
            f"<div style='color:#8B6B6B;font-size:0.7rem;letter-spacing:0.1em;"
            f"font-weight:600;text-transform:uppercase;margin-bottom:6px'>"
            f"Losing (stopped &lt;30d) · {_bs['losing']} ads"
            f"</div>",
            unsafe_allow_html=True,
        )
        if not _bucket_data["losing"]:
            st.caption("No quick-failed ads to analyze yet.")
        else:
            for p in _bucket_data["losing"][:10]:
                st.markdown(
                    f"<div style='padding:4px 0;border-bottom:1px solid var(--pra-border)'>"
                    f"<span style='font-weight:500'>{p['phrase']}</span> "
                    f"<span style='color:var(--pra-text-muted);font-size:0.78rem'>"
                    f"× {p['count']} · avg {p['avg_days']:.0f}d</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    st.caption(
        "**Reading guide:** phrases appearing ONLY in the left column = winning hooks. "
        "Phrases in the right column with no evergreen counterpart = anti-patterns to avoid. "
        "Phrases in both = neutral — they depend on execution."
    )


# ---- Marketplace sections (Shopee + Lazada split) ----

def _render_marketplace_section(container, source: str, label: str,
                                 url_keywords: tuple, all_rows: list) -> None:
    """Render a marketplace tab body for one source. Filters rows + shows enrichment UI."""
    with container:
        st.subheader(f"{label} · ad-to-sales bridge")
        st.caption(
            f"Ads in this run whose landing page goes to **{label}**. "
            f"Once enriched, we capture price, units sold, rating, and review count — "
            f"so you can see which ads actually convert to sales."
        )

        # Filter rows to those whose landing URL points to this marketplace
        def _matches(r):
            url = (r.get("landing_url") or "").lower()
            return any(kw in url for kw in url_keywords)

        scoped = [r for r in all_rows if _matches(r)]
        enriched_here = [r for r in scoped
                          if (r.get("mp_source") or "").lower() == source.lower()
                          and r.get("mp_enriched_at")]
        pending_here = [r for r in scoped if not r.get("mp_enriched_at")]

        mc1, mc2, mc3 = st.columns(3)
        mc1.metric(f"{label} ads", f"{len(scoped):,}")
        mc2.metric("Enriched", f"{len(enriched_here):,}")
        total_sold = sum(int(r.get("mp_sold") or 0) for r in enriched_here)
        mc3.metric("Total units sold", f"{total_sold:,}")

        rc1, rc2 = st.columns([1, 3])
        if IS_CLOUD:
            rc1.info("Enrichment runs on the local desktop app only (needs Playwright).")
        elif rc1.button(
            f"Run enrichment ({label})", type="primary", width="stretch",
            help=f"Launches Playwright in background. Visits {label} product pages. ~10 min.",
            key=f"enrich_btn_{source}",
        ):
            subprocess.Popen(
                [sys.executable, "-u", str(ROOT / "main.py"), "--enrich-browser"],
                cwd=str(ROOT),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
            )
            with st.status(f"{label} enrichment launched", state="running", expanded=True):
                st.write(f"Playwright visiting {label} product pages")
                st.write("ETA ~10 minutes for top 80 products across all marketplaces")
                st.write("Refresh data when done (sidebar button)")
        rc2.caption(
            f"Note: {label} blocks API-based scraping. Browser mode is reliable for prices; "
            f"sold-counts parse correctly ~60% of the time (their DOM changes often)."
        )

        if not enriched_here:
            if scoped:
                st.info(
                    f"Found **{len(scoped)} {label} ads** in this run, but none enriched yet. "
                    f"Click the button above to capture price + sales data."
                )
            else:
                st.info(
                    f"No {label} ads in this run. Try widening filters in the sidebar, "
                    f"or run a fresh scrape."
                )
            return

        df_e = pd.DataFrame(enriched_here)
        keep = ["brand", "niche", "mp_price", "mp_sold", "mp_rating", "mp_reviews",
                "days_running", "is_active", "landing_url"]
        keep = [c for c in keep if c in df_e.columns]
        df_e = df_e.sort_values(by="mp_sold", ascending=False, na_position="last")[keep]
        col_cfg = {
            "brand": st.column_config.TextColumn("Brand", width="medium"),
            "niche": st.column_config.TextColumn("Niche", width="small"),
            "mp_price": st.column_config.NumberColumn("Price", format="₱%.0f", width="small"),
            "mp_sold": st.column_config.NumberColumn("Units sold", format="%d", width="small"),
            "mp_rating": st.column_config.NumberColumn("Rating", format="%.1f ★", width="small"),
            "mp_reviews": st.column_config.NumberColumn("Reviews", format="%d", width="small"),
            "days_running": st.column_config.NumberColumn("Days running", format="%d d", width="small"),
            "is_active": st.column_config.CheckboxColumn("Active", width="small"),
            "landing_url": st.column_config.LinkColumn(
                "Product page", width="large",
                display_text=r"^https?://(?:www\.)?(.{1,55})",
            ),
        }
        st.dataframe(df_e, column_config=col_cfg, width="stretch", hide_index=True, height=520)


if current_page == "shopee":
    _render_marketplace_section(
        st.container(), source="shopee", label="Shopee",
        url_keywords=("shopee.ph", "shp.ee", "shopee."), all_rows=rows,
    )
if current_page == "lazada":
    _render_marketplace_section(
        st.container(), source="lazada", label="Lazada",
        url_keywords=("lazada.com.ph", "lazada.", "s.lazada"), all_rows=rows,
    )


# ---- TikTok Top Ads ----

if current_page == "tiktok":
    st.subheader("TikTok Creative Center — Top Ads (PH)")
    st.caption(
        "Top-performing PH ads as ranked by TikTok's Creative Center. "
        "Engagement signals (likes, plays, CTR) are TikTok's proxy for what's working."
    )

    tt_run = db.latest_tiktok_run_id()
    tc1, tc2, tc3 = st.columns(3)
    tc1.metric("Latest TikTok run", f"#{tt_run}" if tt_run else "none")
    tt_ads = db.tiktok_ads_for_run(tt_run) if tt_run else []
    tc2.metric("TikTok ads in run", len(tt_ads))
    tc3.metric("Avg likes", f"{int(sum(a.get('likes') or 0 for a in tt_ads) / len(tt_ads)):,}" if tt_ads else "—")

    tk1, tk2 = st.columns([1, 3])
    if IS_CLOUD:
        tk1.info("TikTok scraping runs on the local desktop app only (needs Playwright).")
    elif tk1.button("Scrape TikTok now", type="primary", width="stretch",
                     help="~3-5 minutes. Scrapes top ads for each niche."):
        subprocess.Popen(
            [sys.executable, "-u", str(ROOT / "main.py"), "--tiktok"],
            cwd=str(ROOT),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
        )
        with st.status("TikTok scrape launched", state="running", expanded=True):
            st.write("Playwright visiting TikTok Creative Center")
            st.write("ETA ~3-5 minutes")
            st.write("Refresh data when done")
    tk2.caption(
        "**BETA** — scraper runs end-to-end and stores ads, but the cards on TikTok Creative Center "
        "have a non-obvious DOM that needs selector tuning (advertiser names sometimes come back as TikTok's "
        "filter labels like 'Video Views'). Architecture is wired; selectors will improve as we iterate. "
        "Industries are mapped per niche (capsule→Health, cream/oil→Beauty, coffee→F&B)."
    )

    if not tt_ads:
        st.info("No TikTok runs yet. Click the button above to start.")
    else:
        # ---- Clean up stale rows from BEFORE the Phase 11.1 parser fix ----
        # TikTok's Creative Center UI sprinkles campaign-objective labels ("Reach",
        # "Conversions", "Video Views", etc.) into card text. Old scrapes captured
        # these as advertiser names. Filter them out at display time.
        _TIKTOK_BAD_ADVERTISERS = {
            "reach", "conversions", "video views", "app installs", "traffic",
            "leads", "engagement", "brand awareness", "catalog sales",
            "app promotion", "messages", "store visits", "product sales",
            "lead generation", "website conversions",
            # UI strings
            "see more", "view details", "watch now", "all", "filter",
        }
        def _is_real_advertiser(adv: str) -> bool:
            a = (adv or "").strip()
            if not a:
                return False
            if a.lower() in _TIKTOK_BAD_ADVERTISERS:
                return False
            # Pure numeric (e.g. "431", "259") — not a real brand
            if a.replace(",", "").replace(".", "").isdigit():
                return False
            # Must have at least 3 alpha chars
            if sum(c.isalpha() for c in a) < 3:
                return False
            return True

        _clean_ads = [a for a in tt_ads if _is_real_advertiser(a.get("advertiser"))]
        _stale_count = len(tt_ads) - len(_clean_ads)

        if _stale_count:
            _cn_l, _cn_r = st.columns([4, 1])
            with _cn_l:
                st.markdown(
                    f"<div style='background:rgba(230,204,115,0.06);"
                    f"border:1px solid var(--pra-border);border-left:2px solid var(--pra-warning);"
                    f"border-radius:5px;padding:10px 14px;margin-bottom:10px;"
                    f"font-size:0.84rem;color:var(--pra-text)'>"
                    f"<span style='color:var(--pra-warning);font-weight:700;letter-spacing:0.14em;"
                    f"font-size:0.66rem;text-transform:uppercase;margin-right:10px'>Stale rows</span>"
                    f"<strong>{_stale_count}</strong> ad{'s' if _stale_count != 1 else ''} "
                    f"hidden because the advertiser field looks like a TikTok UI label "
                    f"(filter labels were captured by older scrapes). Re-scrape to refresh."
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with _cn_r:
                if st.button(
                    "Purge stale", key="tiktok_purge_btn", type="secondary",
                    width="stretch",
                    help=f"Permanently delete the {_stale_count} bad row(s) from the DB",
                ):
                    import db as _db
                    with _db.connect() as _cn:
                        for _a in tt_ads:
                            if not _is_real_advertiser(_a.get("advertiser")):
                                _cn.execute(
                                    "DELETE FROM tiktok_ads WHERE ad_id = ? AND run_id = ?",
                                    (_a.get("ad_id"), tt_run),
                                )
                    st.toast(f"✓ Purged {_stale_count} stale rows", icon=None)
                    st.rerun()

        if not _clean_ads:
            st.info(
                "All rows in this run look like stale filter-label data. "
                "Click **Scrape TikTok now** to re-fetch with the updated parser."
            )
        else:
            df_t = pd.DataFrame(_clean_ads)
            # Build an absolute, clickable URL from the relative detail_url field.
            # TikTok Creative Center stores paths like "/business/creativecenter/topads/..."
            # which need the domain prefix to actually open.
            def _abs_url(u: str | None) -> str | None:
                if not u:
                    return None
                u = u.strip()
                if not u:
                    return None
                if u.startswith("http"):
                    return u
                if u.startswith("/"):
                    return f"https://ads.tiktok.com{u}"
                return f"https://ads.tiktok.com/{u}"
            if "detail_url" in df_t.columns:
                df_t["detail_url"] = df_t["detail_url"].apply(_abs_url)

            keep = ["advertiser", "industry", "likes", "plays", "ctr",
                    "duration_seconds", "thumbnail_url", "detail_url"]
            keep = [c for c in keep if c in df_t.columns]
            df_t = df_t.sort_values(by="likes", ascending=False)[keep]
            col_cfg = {
                "advertiser": st.column_config.TextColumn("Advertiser", width="medium"),
                "industry":   st.column_config.TextColumn("Niche", width="small"),
                "likes":      st.column_config.NumberColumn("Likes", format="%d", width="small"),
                "plays":      st.column_config.NumberColumn("Plays", format="%d", width="small"),
                "ctr":        st.column_config.NumberColumn("CTR %", format="%.2f", width="small"),
                "duration_seconds": st.column_config.NumberColumn("Length (s)", width="small"),
                "thumbnail_url":    st.column_config.ImageColumn("Thumbnail", width="small"),
                "detail_url":       st.column_config.LinkColumn(
                                       "Open", width="small",
                                       display_text="View →",
                                       help="Open the ad detail page on TikTok Creative Center",
                                    ),
            }
            st.dataframe(df_t, column_config=col_cfg, width="stretch",
                          hide_index=True, height=560)
            st.caption(
                f"Showing **{len(_clean_ads):,}** of {len(tt_ads):,} ads "
                f"(filter-label rows hidden). Click any **View →** link to open the ad on "
                f"TikTok Creative Center in a new tab."
            )


# ============================================================
# Phase 17.5/17.6 — Bestsellers page (Shopee / Lazada / TikTok Shop)
# ============================================================

# ============================================================
# Phase 20.2 — Competitor watchboard
# ============================================================

if current_page == "competitors":
    st.subheader("Competitor watchboard")
    st.caption(
        "Side-by-side view of brands you've flagged as competitors. "
        "Mark any brand as a competitor from its detail modal (◎ button)."
    )

    _comp_brands = set(db.list_competitor_brands())
    if not _comp_brands:
        st.info(
            "No competitors flagged yet. Open any brand (click a row on **FB Ads**), "
            "then click **◎ Mark competitor** in the Actions row."
        )
    else:
        # Build per-competitor metrics from the current run
        _comp_data = []
        for _cb in _comp_brands:
            _cb_ads = [r for r in rows
                       if (r.get("brand") or r.get("page_name") or "").strip().lower() == _cb.lower()]
            if not _cb_ads:
                _comp_data.append({
                    "brand": _cb, "active": 0, "total": 0, "max_days": 0,
                    "niche": "—", "seen": False, "fda": None,
                })
                continue
            _active = sum(1 for a in _cb_ads if a.get("is_active"))
            _maxd = max((a.get("days_running") or 0) for a in _cb_ads)
            _niche = next((a.get("niche") for a in _cb_ads if a.get("niche")), "—")
            # FDA quick-scan
            _fda_sev = None
            try:
                import fda_compliance
                _fr = fda_compliance.scan_brand(_cb_ads)
                _fda_sev = _fr.get("worst_severity")
            except Exception:
                pass
            _comp_data.append({
                "brand": _cb, "active": _active, "total": len(_cb_ads),
                "max_days": _maxd, "niche": _niche, "seen": True, "fda": _fda_sev,
            })

        # Sort: most active first
        _comp_data.sort(key=lambda x: -x["active"])

        st.markdown(
            f"<div style='color:var(--pra-text-muted);font-size:0.82rem;margin-bottom:14px'>"
            f"Tracking <strong>{len(_comp_brands)}</strong> competitor"
            f"{'s' if len(_comp_brands) != 1 else ''} · "
            f"<strong>{sum(1 for c in _comp_data if c['seen'])}</strong> active in this run"
            f"</div>",
            unsafe_allow_html=True,
        )

        # Card grid
        _cmp_per_row = 3
        for _ci in range(0, len(_comp_data), _cmp_per_row):
            _ccols = st.columns(_cmp_per_row)
            for _cj, _c in enumerate(_comp_data[_ci:_ci + _cmp_per_row]):
                with _ccols[_cj]:
                    _fda_badge = ""
                    if _c["fda"]:
                        _fc = {"critical": "var(--pra-danger)", "high": "var(--pra-warning)",
                               "medium": "var(--pra-accent)", "low": "var(--pra-info)"}.get(
                                   _c["fda"], "var(--pra-text-muted)")
                        _fda_badge = (
                            f"<span style='display:inline-block;background:rgba(224,127,159,0.12);"
                            f"border:1px solid {_fc};color:{_fc};padding:1px 6px;border-radius:3px;"
                            f"font-size:0.62rem;font-weight:700;text-transform:uppercase;"
                            f"margin-left:6px'>FDA {_c['fda']}</span>"
                        )
                    _status_html = (
                        f"<span style='color:var(--pra-success)'>● {_c['active']} active</span>"
                        if _c["seen"] and _c["active"]
                        else "<span style='color:var(--pra-text-dim)'>○ not in this run</span>"
                    )
                    st.markdown(
                        f"<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                        f"border-left:2px solid var(--pra-info);border-radius:6px;"
                        f"padding:14px 16px;margin-bottom:10px;min-height:120px'>"
                        f"<div style='font-size:0.95rem;font-weight:600;color:var(--pra-text);"
                        f"margin-bottom:6px'>{_c['brand']}{_fda_badge}</div>"
                        f"<div style='color:var(--pra-text-muted);font-size:0.78rem;line-height:1.7'>"
                        f"{_c['niche']} · {_c['total']} ads · max {_c['max_days']:,}d<br>"
                        f"{_status_html}"
                        f"</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                    if st.button("✕ Remove", key=f"comp_remove_{_c['brand']}",
                                 width="stretch"):
                        db.toggle_competitor(_c["brand"])
                        st.rerun()


if current_page == "bestsellers":
    st.subheader("Marketplace bestsellers")
    st.caption(
        "Top-selling products on Shopee, Lazada, and TikTok Shop — a different signal "
        "than ad library data. **These are products people are actually buying right now.** "
        "Use this to validate that an ad-research winner is also selling."
    )

    # Trigger row
    _bs_run_l, _bs_run_r = st.columns([3, 1])
    with _bs_run_l:
        _bs_keyword = st.text_input(
            "Keyword to scrape", value="slimming capsule",
            help="A product keyword like 'whitening cream', 'pampapayat', 'fish oil'.",
            key="bs_keyword",
        )
        _bs_niche_pick = st.selectbox(
            "Tag with niche (optional)",
            ["", "capsule", "cream", "oil", "coffee", "balm"],
            index=0, key="bs_niche_pick",
        )
        _bs_platforms = st.multiselect(
            "Platforms",
            ["shopee", "lazada", "tiktok_shop"],
            default=["shopee", "lazada"],
            format_func=lambda x: {"shopee": "Shopee", "lazada": "Lazada",
                                   "tiktok_shop": "TikTok Shop"}.get(x, x),
            help="TikTok Shop blocks aggressively — include only if needed.",
        )
    with _bs_run_r:
        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        if IS_CLOUD:
            st.info(
                "Bestsellers scraping runs on the local desktop app only (needs "
                "Playwright). This cloud view shows previously captured data."
            )
        elif st.button("Scrape bestsellers", key="bs_run_btn", type="primary",
                       width="stretch",
                       help="~30-90 seconds per platform per keyword. Headless Playwright."):
            with st.spinner(f"Scraping {len(_bs_platforms)} platform(s)..."):
                import bestsellers_scraper
                _records = bestsellers_scraper.scrape_all(
                    keywords=[_bs_keyword.strip()],
                    niche=_bs_niche_pick or "",
                    platforms=_bs_platforms,
                    max_per_platform=20,
                )
            if not _records:
                st.warning(
                    "0 products scraped — marketplace may have blocked the request or "
                    "DOM changed. Retry in a few minutes, or try a different keyword."
                )
            else:
                db.insert_bestsellers(_records)
                st.toast(f"✓ Captured {len(_records)} bestsellers", icon=None)
                st.rerun()

    st.divider()

    # Display latest bestsellers
    _bs_filter_a, _bs_filter_b = st.columns(2)
    with _bs_filter_a:
        _bs_filter_platform = st.selectbox(
            "Platform filter",
            ["All", "shopee", "lazada", "tiktok_shop"],
            index=0, key="bs_filter_platform",
            format_func=lambda x: {"All": "All platforms", "shopee": "Shopee",
                                   "lazada": "Lazada", "tiktok_shop": "TikTok Shop"}.get(x, x),
        )
    with _bs_filter_b:
        _bs_filter_niche = st.selectbox(
            "Niche filter",
            ["All", "capsule", "cream", "oil", "coffee", "balm"],
            index=0, key="bs_filter_niche",
        )

    _bs_data = db.latest_bestsellers(
        platform=_bs_filter_platform if _bs_filter_platform != "All" else None,
        niche=_bs_filter_niche if _bs_filter_niche != "All" else None,
        limit=80,
    )
    if not _bs_data:
        st.info("No bestseller data captured yet. Click **Scrape bestsellers** above to start.")
    else:
        st.caption(
            f"Showing {len(_bs_data)} top products from the latest snapshot "
            f"({_bs_data[0].get('snapshot_date', '—')})."
        )
        _df_bs = pd.DataFrame(_bs_data)
        _keep_bs = ["rank", "platform", "product_name", "price_php", "units_sold",
                    "shop_name", "thumbnail_url", "product_url"]
        _keep_bs = [c for c in _keep_bs if c in _df_bs.columns]
        _df_bs = _df_bs[_keep_bs]
        _bs_col_cfg = {
            "rank":          st.column_config.NumberColumn("Rank", width="small"),
            "platform":      st.column_config.TextColumn("Platform", width="small"),
            "product_name":  st.column_config.TextColumn("Product", width="large"),
            "price_php":     st.column_config.NumberColumn("Price", format="₱%.0f", width="small"),
            "units_sold":    st.column_config.NumberColumn("Sold", format="%d", width="small"),
            "shop_name":     st.column_config.TextColumn("Shop", width="medium"),
            "thumbnail_url": st.column_config.ImageColumn("Image", width="small"),
            "product_url":   st.column_config.LinkColumn(
                                "Open", width="small", display_text="View →"),
        }
        st.dataframe(_df_bs, column_config=_bs_col_cfg, width="stretch",
                      hide_index=True, height=560)


# ---- Notifications (Telegram) ----

if current_page == "notifications":
    st.subheader("Telegram daily summary")
    st.caption(
        "Every morning, get a 1-message digest: new winners, retirements, niche heat, top sustained brands. "
        "Free Telegram bot — no email setup needed."
    )

    cfg = notifications.load_config()
    configured = notifications.is_configured()

    with st.expander("Setup instructions (first time)", expanded=not configured):
        st.markdown(
            "1. Open Telegram, search **@BotFather**, send `/newbot`. Follow prompts, **save the bot token** it gives you.\n"
            "2. Open a chat with your new bot, send any message (e.g. \"hi\").\n"
            "3. Visit `https://api.telegram.org/bot<YOUR_TOKEN>/getUpdates` in your browser — find `\"chat\":{\"id\":12345...}`. **That number is your chat_id.**\n"
            "4. Paste both below and click Save. Then click Send test to verify."
        )

    with st.form("notify_form"):
        token = st.text_input("Bot token", value=cfg.get("telegram_token", ""), type="password",
                               help="From @BotFather. Looks like `123456789:AAH...`")
        chat_id = st.text_input("Chat ID", value=str(cfg.get("telegram_chat_id", "")),
                                 help="Numeric, e.g. 12345678 (negative for groups)")
        col_s1, col_s2 = st.columns(2)
        save_clicked = col_s1.form_submit_button("Save", width="stretch")
        test_clicked = col_s2.form_submit_button("Send test message", width="stretch")

    if save_clicked:
        notifications.save_config({"telegram_token": token.strip(), "telegram_chat_id": chat_id.strip()})
        st.success("Saved.")
    if test_clicked:
        notifications.save_config({"telegram_token": token.strip(), "telegram_chat_id": chat_id.strip()})
        ok, msg = notifications.send_telegram("Test from Product Research Agent — Telegram setup works.")
        (st.success if ok else st.error)(msg)

    st.divider()
    st.markdown("**Preview of the daily summary** (based on current DB state):")
    st.code(notifications.build_daily_summary(), language="markdown")
    if st.button("Send this summary to Telegram now", type="primary", disabled=not configured):
        ok, msg = notifications.send_daily_summary()
        (st.success if ok else st.error)(msg)


# ---- Run History ----

if current_page == "history":
    st.subheader("All runs in the database")
    st.caption("Complete log including TikTok runs and failed/empty scrapes (which are hidden from the main Run dropdown).")
    all_runs_history = db.list_runs(limit=200, only_meta=False)
    if all_runs_history:
        df_runs = pd.DataFrame(all_runs_history)
        keep = [c for c in ("run_id", "started_at", "finished_at", "niches", "total_ads", "source", "notes")
                if c in df_runs.columns]
        st.dataframe(df_runs[keep], width="stretch", hide_index=True)


# ---- Settings / Profile page (Phase 10.2) ----

if current_page == "settings":
    st.subheader("Settings · Orbit configuration")
    st.caption("Centralized config for API keys, defaults, and data management.")

    # ---- API keys ----
    _section_label("API keys")
    _sa_l, _sa_r = st.columns([3, 1])
    with _sa_l:
        _anth_set = bool(os.environ.get("ANTHROPIC_API_KEY"))
        _ant_status = "✓ Set" if _anth_set else "✗ Not set"
        _ant_color = "var(--pra-success)" if _anth_set else "var(--pra-danger)"
        st.markdown(
            f"<div style='padding:10px 14px;background:var(--pra-panel);"
            f"border:1px solid var(--pra-border);border-radius:4px'>"
            f"<div style='font-size:0.86rem;font-weight:500;color:var(--pra-text)'>"
            f"<strong>ANTHROPIC_API_KEY</strong> "
            f"<span style='color:{_ant_color};font-weight:600;margin-left:8px'>{_ant_status}</span>"
            f"</div>"
            f"<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-top:4px'>"
            f"Needed for: Copy Studio, LLM ad classification. "
            f"Get a key from console.anthropic.com, then set the env var and restart."
            f"</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with _sa_r:
        st.link_button("Get key →", "https://console.anthropic.com/",
                       width="stretch", help="Anthropic console")

    # ---- Telegram config ----
    _section_label("Telegram notifications")
    _tg_cfg = notifications.load_config()
    _tg_set = notifications.is_configured()
    _tg_status = "✓ Configured" if _tg_set else "✗ Not configured"
    _tg_color = "var(--pra-success)" if _tg_set else "var(--pra-text-muted)"
    st.markdown(
        f"<div style='padding:10px 14px;background:var(--pra-panel);"
        f"border:1px solid var(--pra-border);border-radius:4px'>"
        f"<div style='font-size:0.86rem;font-weight:500;color:var(--pra-text)'>"
        f"Daily digest bot "
        f"<span style='color:{_tg_color};font-weight:600;margin-left:8px'>{_tg_status}</span>"
        f"</div>"
        f"<div style='color:var(--pra-text-muted);font-size:0.78rem;margin-top:4px'>"
        f"Configure in the Notifications page. Get morning briefs delivered to Telegram."
        f"</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ---- Defaults ----
    _section_label("Defaults")
    _def_a, _def_b = st.columns(2)
    with _def_a:
        _cfg_niches = list(load_config().get("niches", {}).keys())
        _default_niche = st.selectbox(
            "Default niche on startup",
            ["— all niches —"] + _cfg_niches,
            index=0,
            key="settings_default_niche",
            help="Pre-select this niche filter on app launch",
        )
    with _def_b:
        _default_theme = st.selectbox(
            "Default theme",
            ["dark", "light"],
            index=(0 if st.session_state.get("theme_mode") == "dark" else 1),
            key="settings_default_theme",
        )
        if _default_theme != st.session_state.theme_mode:
            st.session_state.theme_mode = _default_theme
            st.rerun()

    # ---- Daily scheduler (Phase 15.1) ----
    _section_label("Daily scrape scheduler")
    st.caption(
        "Runs `main.py` automatically every day via Windows Task Scheduler. "
        "No admin rights needed (per-user task). Logs to `logs/agent.log`."
    )

    # Detect existing scheduled task
    import subprocess as _sp
    _is_scheduled = False
    _next_run = None
    try:
        _res = _sp.run(
            ["schtasks", "/Query", "/TN", "OrbitDailyScrape", "/FO", "LIST", "/V"],
            capture_output=True, text=True, timeout=4,
        )
        if _res.returncode == 0:
            _is_scheduled = True
            # Parse next-run-time from schtasks output
            for line in _res.stdout.splitlines():
                if "Next Run Time" in line:
                    _next_run = line.split(":", 1)[1].strip()
                    break
    except Exception:
        pass

    _sch_a, _sch_b, _sch_c = st.columns([2, 1, 1])
    with _sch_a:
        _sched_time = st.text_input(
            "Scrape time (HH:MM, 24-hour)",
            value="06:00",
            key="settings_sched_time",
            help="Pick a time before you start work — e.g., 06:00 so data's ready when you open the app.",
        )
        if _is_scheduled:
            _badge = f"<span style='color:var(--pra-success);font-weight:600'>● Active</span>"
            _next = (
                f" · Next run: <strong style='color:var(--pra-text)'>{_next_run}</strong>"
                if _next_run else ""
            )
        else:
            _badge = f"<span style='color:var(--pra-text-muted);font-weight:600'>○ Not scheduled</span>"
            _next = ""
        st.markdown(
            f"<div style='color:var(--pra-text-muted);font-size:0.78rem;"
            f"margin-top:2px'>{_badge}{_next}</div>",
            unsafe_allow_html=True,
        )
    with _sch_b:
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        if st.button(
            "Enable / Update" if not _is_scheduled else "Update time",
            key="settings_sched_register", width="stretch", type="primary",
            help="Registers (or updates) the Windows scheduled task with the time above",
        ):
            _ps_script = ROOT / "register_daily_task.ps1"
            _r = _sp.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
                 "-File", str(_ps_script), "-Time", _sched_time],
                capture_output=True, text=True, timeout=15,
            )
            if _r.returncode == 0:
                st.toast(f"✓ Daily scrape scheduled at {_sched_time}", icon=None)
                st.rerun()
            else:
                st.error(f"Schedule failed: {_r.stderr[:200] or _r.stdout[:200]}")
    with _sch_c:
        st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
        if _is_scheduled and st.button(
            "Disable", key="settings_sched_remove", width="stretch",
            help="Removes the scheduled task",
        ):
            _ps_script = ROOT / "register_daily_task.ps1"
            _r = _sp.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
                 "-File", str(_ps_script), "-Remove"],
                capture_output=True, text=True, timeout=10,
            )
            if _r.returncode == 0:
                st.toast("✓ Daily scrape disabled", icon=None)
                st.rerun()

    # ---- Data management ----
    _section_label("Data management")
    _db_stats = db.stats()
    _dm_a, _dm_b, _dm_c = st.columns(3)
    _dm_a.metric("Total ads in DB", f"{_db_stats['unique_library_ids']:,}")
    _dm_b.metric("Unique brands", f"{_db_stats['unique_brands']:,}")
    _dm_c.metric("Total runs", f"{_db_stats['total_runs']:,}")

    _dm_x, _dm_y, _dm_z = st.columns(3)
    with _dm_x:
        if st.button("Cleanup empty runs", key="settings_cleanup_runs",
                     width="stretch",
                     help="Remove runs >30min old with 0 ads (failed/cancelled scrapes)"):
            _n_cleaned = db.cleanup_empty_runs(min_age_minutes=30)
            get_runs.clear()
            st.toast(f"✓ Removed {_n_cleaned} empty run(s)", icon=None)
            st.rerun()
    with _dm_y:
        if st.button("Backfill category + location", key="settings_backfill_cat",
                     width="stretch",
                     help="Re-run the keyword classifier on all ads to populate "
                          "Category / Sub-category / Location fields"):
            with st.spinner("Running backfill..."):
                _r = db.backfill_categorization()
            st.toast(
                f"✓ {_r['processed']:,} ads scanned · "
                f"+{_r['updated_cat']:,} cat · +{_r['updated_sub']:,} sub · "
                f"+{_r['updated_loc']:,} loc",
                icon=None,
            )
            get_ads.clear()
            _clear_brand_caches()
            st.rerun()
    with _dm_z:
        if st.button("Reset onboarding tour", key="settings_reset_onboarding",
                     width="stretch"):
            try:
                if ONBOARDING_FLAG_PATH.exists():
                    ONBOARDING_FLAG_PATH.unlink()
                st.toast("✓ Onboarding will show next reload", icon=None)
            except Exception as _e:
                st.error(f"Reset failed: {_e}")

    # ---- About ----
    _section_label("About")
    st.markdown(
        "<div style='padding:10px 14px;background:var(--pra-panel);"
        "border:1px solid var(--pra-border);border-radius:4px;font-size:0.85rem;"
        "color:var(--pra-text);line-height:1.6'>"
        "<strong>Orbit</strong> · Product Research Hunter for Philippine ecommerce<br>"
        "<span style='color:var(--pra-text-muted);font-size:0.78rem'>"
        "Local-first · SQLite DB · Free tools only · No cloud sync"
        "</span>"
        "</div>",
        unsafe_allow_html=True,
    )


# ============================================================================
# Testing tab — product testing lineup with PDF export
# ============================================================================

@st.dialog("Edit testing product", width="large")
def _edit_testing_dialog(product_id: int) -> None:
    p = db.get_testing_product(product_id)
    if not p:
        st.error("Product not found.")
        return

    st.markdown(
        f"<div style='font-size:1.35rem;font-weight:600;margin-bottom:4px;letter-spacing:-0.01em'>"
        f"{p.get('product_name','—')}</div>"
        f"<div style='color:var(--pra-text-muted);font-size:0.8rem;margin-bottom:14px'>"
        f"#{p.get('id')} · added {p.get('date_added','—')} · status: {p.get('status','queued')}"
        f"</div>",
        unsafe_allow_html=True,
    )

    with st.form(f"edit_form_{product_id}"):
        # ---- BASICS section ----
        _section_label("Basics")
        col_a, col_b = st.columns(2)
        date_added = col_a.text_input("Date", value=p.get("date_added", ""), key=f"e_date_{product_id}")
        status = col_b.selectbox(
            "Status", db.TESTING_STATUSES,
            index=db.TESTING_STATUSES.index(p.get("status", "queued")) if p.get("status") in db.TESTING_STATUSES else 0,
            key=f"e_status_{product_id}",
        )
        product_name = st.text_input("Product name", value=p.get("product_name", ""), key=f"e_pn_{product_id}")

        col_c, col_d = st.columns(2)
        brand_name = col_c.text_input("Brand name", value=p.get("brand_name", ""), key=f"e_bn_{product_id}")
        niche_options = [""] + list(load_config().get("niches", {}).keys())
        cur_niche = p.get("niche", "")
        if cur_niche not in niche_options:
            niche_options.append(cur_niche)
        niche_val = col_d.selectbox(
            "Niche", niche_options,
            index=niche_options.index(cur_niche),
            key=f"e_niche_{product_id}",
            format_func=lambda x: "— pick a niche —" if x == "" else x.replace("_", " "),
        )
        hunted_by = st.text_input(
            "From (hunted by)",
            value=p.get("hunted_by", "") or "",
            placeholder="Your name / VA / source",
            help="Who found this product?",
            key=f"e_hunted_{product_id}",
        )

        # ---- POSITIONING section (the marketing details) ----
        _section_label("Positioning · why this sells")
        pain_point = st.text_area(
            "Pain point or desire",
            value=p.get("pain_point", ""), height=70, key=f"e_pp_{product_id}",
            placeholder="What problem does this solve? e.g. struggle to lose weight despite diet",
        )
        emotional_benefits = st.text_area(
            "Emotional payoff",
            value=p.get("emotional_benefits", ""), height=70, key=f"e_eb_{product_id}",
            placeholder="How does the customer feel after? e.g. confident, beautiful again",
        )
        physical_effects = st.text_area(
            "Physical effect",
            value=p.get("physical_effects", ""), height=70, key=f"e_pe_{product_id}",
            placeholder="What does it actually do? e.g. reduced bloating in 2 weeks",
        )
        main_ingredients = st.text_area(
            "Main ingredients",
            value=p.get("main_ingredients", ""), height=70, key=f"e_ing_{product_id}",
            placeholder="Key actives — e.g. L-Carnitine, Garcinia Cambogia",
        )

        # ---- TARGET MARKET section ----
        _section_label("Target market")
        tm_a, tm_b = st.columns(2)
        target_age = tm_a.text_input("Age", value=p.get("target_age", ""),
                                       placeholder="25-45", key=f"e_age_{product_id}")
        target_gender = tm_b.text_input("Gender", value=p.get("target_gender", ""),
                                          placeholder="Female", key=f"e_gen_{product_id}")
        target_behavior = st.text_input(
            "Behavior",
            value=p.get("target_behavior", ""),
            placeholder="Shops weekly on Shopee/Lazada, follows beauty influencers",
            key=f"e_beh_{product_id}",
        )
        target_interest = st.text_input(
            "Interest",
            value=p.get("target_interest", ""),
            placeholder="Wellness · beauty · K-pop · fashion",
            key=f"e_int_{product_id}",
        )
        target_demographics = st.text_input(
            "Demographics",
            value=p.get("target_demographics", ""),
            placeholder="NCR/Metro Manila · BPO workers · urban moms",
            key=f"e_dem_{product_id}",
        )

        # ---- PH PERMITS section (Phase 16.2 — DTI / FDA / BIR tracking) ----
        _section_label("PH permits & compliance")
        _pcol_a, _pcol_b = st.columns(2)
        with _pcol_a:
            dti_permit_no = st.text_input(
                "DTI permit no.",
                value=p.get("dti_permit_no", "") or "",
                placeholder="e.g. 12345-XYZ-2026",
                key=f"e_dti_{product_id}",
                help="DTI Business Name Registration number (renew yearly)",
            )
            dti_expiry = st.text_input(
                "DTI expiry (YYYY-MM-DD)",
                value=p.get("dti_expiry", "") or "",
                placeholder="2026-12-31",
                key=f"e_dti_exp_{product_id}",
            )
            bir_or_no = st.text_input(
                "BIR OR / Cert no.",
                value=p.get("bir_or_no", "") or "",
                placeholder="optional",
                key=f"e_bir_{product_id}",
            )
        with _pcol_b:
            fda_cpr_no = st.text_input(
                "FDA CPR / CFRR no.",
                value=p.get("fda_cpr_no", "") or "",
                placeholder="Required for capsules/cream/oil/coffee with claims",
                key=f"e_fda_{product_id}",
                help="FDA Certificate of Product Registration / Certificate of Free "
                     "Registration & Renewal. Without this, ad claims can trigger seizure.",
            )
            fda_expiry = st.text_input(
                "FDA expiry (YYYY-MM-DD)",
                value=p.get("fda_expiry", "") or "",
                placeholder="2027-06-30",
                key=f"e_fda_exp_{product_id}",
            )

        # Expiry warning
        from datetime import date as _date_pe
        _today_iso = _date_pe.today().isoformat()
        _expiring_soon = []
        for _label_pe, _val_pe in (("DTI", dti_expiry), ("FDA", fda_expiry)):
            if _val_pe and _val_pe >= _today_iso:
                try:
                    _exp_date = _date_pe.fromisoformat(_val_pe.strip())
                    _days_left = (_exp_date - _date_pe.today()).days
                    if _days_left <= 60:
                        _expiring_soon.append((_label_pe, _days_left, _val_pe))
                except Exception:
                    pass
        if _expiring_soon:
            for _label_pe, _days_left, _val_pe in _expiring_soon:
                _color_pe = "var(--pra-danger)" if _days_left <= 14 else "var(--pra-warning)"
                st.markdown(
                    f"<div style='background:rgba(230,204,115,0.06);"
                    f"border:1px solid var(--pra-border);border-left:2px solid {_color_pe};"
                    f"border-radius:4px;padding:8px 12px;margin-top:6px;font-size:0.82rem;"
                    f"color:var(--pra-text)'>"
                    f"⚠ <strong>{_label_pe} permit expires in {_days_left} day"
                    f"{'s' if _days_left != 1 else ''}</strong> ({_val_pe}) — renew before "
                    f"ads risk seizure.</div>",
                    unsafe_allow_html=True,
                )

        # ---- RESULTS & ROI section (Phase 9.2 — close the research→test→learn loop) ----
        _section_label("Results & ROI")
        st.caption(
            "Log actual launch performance. The aggregate ROI panel on the Testing page "
            "uses this to surface which research signals predicted winners."
        )

        # ---- Receipt OCR (Phase 16.3) — upload a sales receipt photo, auto-fill revenue ----
        with st.expander("📸 Upload sales receipt to auto-tally revenue (Claude Vision OCR)", expanded=False):
            _has_anth = bool(os.environ.get("ANTHROPIC_API_KEY"))
            if not _has_anth:
                st.caption("Set ANTHROPIC_API_KEY env var to enable receipt OCR (~₱0.30 per receipt).")
            _ocr_file = st.file_uploader(
                "Upload BIR receipt, Shopee/Lazada invoice, or supplier slip",
                type=["png", "jpg", "jpeg", "webp"],
                key=f"e_ocr_upload_{product_id}",
                disabled=not _has_anth,
            )
            if _ocr_file is not None and st.button(
                "Extract receipt data", key=f"e_ocr_run_{product_id}",
                type="primary", disabled=not _has_anth,
            ):
                with st.spinner("Reading receipt via Claude Vision..."):
                    import receipt_ocr
                    _ocr_res = receipt_ocr.extract_from_image(
                        _ocr_file.read(),
                        mime_type=_ocr_file.type or "image/jpeg",
                    )
                if not _ocr_res["ok"]:
                    st.error(f"OCR failed: {_ocr_res['error']}")
                else:
                    _d = _ocr_res["data"]
                    if _d.get("error"):
                        st.warning(f"Not a recognizable receipt: {_d.get('notes', '')}")
                    else:
                        _total = _d.get("total_php")
                        _date = _d.get("date")
                        _line_n = len(_d.get("line_items", []) or [])
                        st.success(
                            f"✓ Extracted: {_d.get('merchant_name','—')} · "
                            f"{_date or '—'} · ₱{_total or 0:,.2f} · "
                            f"{_line_n} line items"
                        )
                        # Stash in session so user can click "Add to revenue"
                        st.session_state[f"_ocr_data_{product_id}"] = _d
                        st.json(_d)
                _stash = st.session_state.get(f"_ocr_data_{product_id}")
                if _stash and _stash.get("total_php"):
                    if st.button(
                        f"➕ Add ₱{_stash['total_php']:,.2f} to revenue field above",
                        key=f"e_ocr_apply_{product_id}",
                    ):
                        st.session_state[f"e_rev_{product_id}"] = (
                            float(p.get("revenue_php") or 0) + float(_stash["total_php"])
                        )
                        if _stash.get("date") and not p.get("launch_date"):
                            st.session_state[f"e_launch_{product_id}"] = _stash["date"]
                        st.toast(
                            f"✓ Added ₱{_stash['total_php']:,.2f} to revenue",
                            icon=None,
                        )
                        st.session_state.pop(f"_ocr_data_{product_id}", None)
                        st.rerun()

        _roi_a, _roi_b, _roi_c = st.columns(3)
        with _roi_a:
            launch_date_v = p.get("launch_date") or ""
            launch_date = st.text_input(
                "Launch date (YYYY-MM-DD)",
                value=launch_date_v,
                placeholder="2026-01-15",
                key=f"e_launch_{product_id}",
            )
            units_sold = st.number_input(
                "Units sold", value=int(p.get("units_sold") or 0),
                min_value=0, step=1, key=f"e_units_{product_id}",
            )
        with _roi_b:
            revenue_php = st.number_input(
                "Revenue (PHP)", value=float(p.get("revenue_php") or 0),
                min_value=0.0, step=100.0, key=f"e_rev_{product_id}",
            )
            ad_spend_php = st.number_input(
                "Ad spend (PHP)", value=float(p.get("ad_spend_php") or 0),
                min_value=0.0, step=100.0, key=f"e_spend_{product_id}",
            )
        with _roi_c:
            cogs_php = st.number_input(
                "COGS (PHP)", value=float(p.get("cogs_php") or 0),
                min_value=0.0, step=100.0, key=f"e_cogs_{product_id}",
                help="Cost of goods sold = supplier cost × units sold",
            )
            _outcome_options = ["", "winner", "breakeven", "loser", "paused"]
            outcome_cur = p.get("outcome") or ""
            if outcome_cur not in _outcome_options:
                outcome_cur = ""
            outcome = st.selectbox(
                "Outcome",
                _outcome_options,
                index=_outcome_options.index(outcome_cur),
                format_func=lambda x: "— pending —" if x == "" else x.capitalize(),
                key=f"e_outcome_{product_id}",
            )

        # Compute derived metrics in-form (read-only display)
        _calc_roas = (revenue_php / ad_spend_php) if ad_spend_php else 0
        _calc_net = revenue_php - ad_spend_php - cogs_php
        # Float-safe band: treat near-zero as breakeven (use a peso-cent threshold)
        _roi_color = "var(--pra-success)" if _calc_net > 1.0 else (
            "var(--pra-danger)" if _calc_net < -1.0 else "var(--pra-warning)"
        )
        st.markdown(
            f"<div style='display:flex;gap:14px;margin-top:6px;padding:10px 14px;"
            f"background:var(--pra-subtle-bg);border:1px solid var(--pra-border);"
            f"border-radius:4px;font-size:0.85rem'>"
            f"<div><span style='color:var(--pra-text-muted)'>ROAS:</span> "
            f"<strong style='color:var(--pra-text);font-variant-numeric:tabular-nums'>"
            f"{_calc_roas:.2f}x</strong></div>"
            f"<div><span style='color:var(--pra-text-muted)'>Net profit:</span> "
            f"<strong style='color:{_roi_color};font-variant-numeric:tabular-nums'>"
            f"₱{_calc_net:,.0f}</strong></div>"
            f"</div>",
            unsafe_allow_html=True,
        )

        learnings = st.text_area(
            "Postmortem / learnings",
            value=p.get("learnings", "") or "",
            height=70,
            placeholder=(
                "What worked / what didn't · best-performing creative angle · "
                "audience that converted · what you'd do differently"
            ),
            key=f"e_learn_{product_id}",
        )

        # ---- LAUNCH CHECKLIST section (Phase 10.4) ----
        # Show when status is launching/launched. Helps user not skip critical steps.
        if status in ("launching", "launched"):
            _section_label("Launch checklist")
            import json as _json_lc
            try:
                _existing_lc = _json_lc.loads(p.get("launch_checklist") or "{}")
            except Exception:
                _existing_lc = {}
            _CHECKLIST_ITEMS = [
                ("creatives_ready",   "Creatives (image/video) ready"),
                ("ad_copy_finalized", "Ad copy finalized (Taglish/EN)"),
                ("fb_campaign_setup", "FB Ads campaign set up"),
                ("pixel_installed",   "FB Pixel installed + tested"),
                ("shopee_listing",    "Shopee listing live"),
                ("lazada_listing",    "Lazada listing live (optional)"),
                ("supplier_order",    "Supplier order placed (1688)"),
                ("inventory_ready",   "Inventory landed / ready"),
                ("logistics_ready",   "Logistics partner set (J&T / LBC)"),
                ("budget_allocated",  "Daily budget allocated"),
            ]
            _lc_state = {}
            for _key, _label in _CHECKLIST_ITEMS:
                _lc_state[_key] = st.checkbox(
                    _label,
                    value=bool(_existing_lc.get(_key, False)),
                    key=f"e_lc_{_key}_{product_id}",
                )
            _checked_n = sum(1 for v in _lc_state.values() if v)
            _total_n = len(_CHECKLIST_ITEMS)
            _pct = int((_checked_n / _total_n) * 100) if _total_n else 0
            _pct_color = "var(--pra-success)" if _pct == 100 else (
                "var(--pra-warning)" if _pct >= 60 else "var(--pra-text-muted)"
            )
            st.markdown(
                f"<div style='margin-top:8px;padding:8px 12px;background:var(--pra-subtle-bg);"
                f"border:1px solid var(--pra-border);border-radius:4px;font-size:0.82rem'>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;"
                f"margin-bottom:6px'>"
                f"<span style='color:var(--pra-text)'>Progress</span>"
                f"<span style='color:{_pct_color};font-weight:600;font-variant-numeric:tabular-nums'>"
                f"{_checked_n}/{_total_n} ({_pct}%)</span>"
                f"</div>"
                f"<div style='background:var(--pra-border);border-radius:2px;height:5px;overflow:hidden'>"
                f"<div style='background:{_pct_color};width:{_pct}%;height:100%'></div>"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
            launch_checklist_json = _json_lc.dumps(_lc_state)
        else:
            launch_checklist_json = p.get("launch_checklist") or None
            # Quiet hint if there's saved checklist data but we're not showing it
            if launch_checklist_json:
                st.caption(
                    "ℹ Launch checklist data is preserved. Switch status to "
                    "**launching** or **launched** to view/edit it."
                )

        # ---- NOTES section ----
        _section_label("Notes")
        notes = st.text_area(
            "Your notes",
            value=p.get("notes", ""), height=60, key=f"e_notes_{product_id}",
            placeholder="Checked Shopee · 5k sold · supplier found on 1688 · test order placed",
            label_visibility="collapsed",
        )

        col_save, col_delete = st.columns([3, 1])
        save_clicked = col_save.form_submit_button("Save changes", type="primary", width="stretch")
        delete_clicked = col_delete.form_submit_button("Delete", width="stretch")

    if save_clicked:
        db.update_testing_product(product_id, {
            "date_added": date_added, "status": status, "product_name": product_name,
            "brand_name": brand_name, "niche": niche_val,
            "hunted_by": hunted_by,
            "pain_point": pain_point, "emotional_benefits": emotional_benefits,
            "physical_effects": physical_effects, "main_ingredients": main_ingredients,
            "target_age": target_age, "target_gender": target_gender,
            "target_behavior": target_behavior, "target_interest": target_interest,
            "target_demographics": target_demographics, "notes": notes,
            # ROI tracker — save zero values too (0 = "no spend yet" is legitimate).
            "launch_date":      launch_date.strip() or None,
            "units_sold":       int(units_sold),
            "revenue_php":      float(revenue_php),
            "ad_spend_php":     float(ad_spend_php),
            "cogs_php":         float(cogs_php),
            "roas":             round(_calc_roas, 3),
            "net_profit_php":   round(_calc_net, 2),
            "outcome":          outcome or None,
            "learnings":        learnings.strip() or None,
            "launch_checklist": launch_checklist_json,
            # PH permits (Phase 16.2)
            "dti_permit_no":    dti_permit_no.strip() or None,
            "dti_expiry":       dti_expiry.strip() or None,
            "fda_cpr_no":       fda_cpr_no.strip() or None,
            "fda_expiry":       fda_expiry.strip() or None,
            "bir_or_no":        bir_or_no.strip() or None,
        })
        st.toast("✓ Product updated", icon=None)
        st.rerun()

    # ---- Delete confirmation: two-stage. Streamlit closes the @st.dialog on rerun,
    # so we must NOT call st.rerun() when arming the confirm — instead, set state
    # and let the same execution render the confirmation card below the form. ----
    _confirm_key = f"_confirm_del_test_{product_id}"
    if delete_clicked:
        st.session_state[_confirm_key] = True
        # Intentionally NO st.rerun() — that would close the dialog.
    if st.session_state.get(_confirm_key):
        st.markdown(
            "<div style='background:rgba(224,127,159,0.08);border:1px solid var(--pra-danger);"
            "border-radius:6px;padding:14px 18px;margin-top:10px'>"
            "<div style='color:var(--pra-danger);font-weight:600;font-size:0.9rem;margin-bottom:4px'>"
            "Delete this product?</div>"
            "<div style='color:var(--pra-text-muted);font-size:0.82rem'>"
            "This will permanently remove the entry from your testing lineup. "
            "This action cannot be undone.</div>"
            "</div>",
            unsafe_allow_html=True,
        )
        _ck1, _ck2 = st.columns(2)
        # Cancel: clear state. The natural button-click rerun closes the dialog,
        # but the user explicitly chose to back out — they can re-open the row to retry.
        if _ck1.button("Cancel", key=f"cancel_del_{product_id}", width="stretch"):
            st.session_state.pop(_confirm_key, None)
        # Confirm: delete + clear state + toast. The natural rerun closes the dialog,
        # which is the desired behavior (row is gone, dialog should close).
        if _ck2.button("Yes, delete", key=f"confirm_del_{product_id}",
                       width="stretch", type="primary"):
            db.delete_testing_product(product_id)
            st.session_state.pop(_confirm_key, None)
            st.toast("✕ Product deleted", icon=None)


# ============================================================
# Phase 16.4 — My Stores page (track your OWN Shopee/Lazada/TikTok shops)
# ============================================================

if current_page == "my_stores":
    st.subheader("My Stores · seller-side analytics")
    st.caption(
        "Track YOUR Shopee / Lazada / TikTok Shop stores. Log periodic snapshots "
        "(units sold, revenue, rating, followers) to see your own performance over time. "
        "Complements the ad-research data with what you're actually selling."
    )

    _stores = db.list_seller_stores()
    _ms_top_l, _ms_top_r = st.columns([3, 1])
    with _ms_top_l:
        st.markdown(
            f"<div style='color:var(--pra-text);font-size:0.95rem;margin-top:6px'>"
            f"<strong>{len(_stores)}</strong> store"
            f"{'s' if len(_stores) != 1 else ''} tracked"
            f"</div>",
            unsafe_allow_html=True,
        )
    with _ms_top_r:
        if st.button("➕ Add store", key="ms_add_btn", type="primary", width="stretch"):
            st.session_state["_ms_show_add"] = True
            st.rerun()

    # Add-store form (dialog-style inline)
    if st.session_state.get("_ms_show_add"):
        with st.form("ms_add_form"):
            st.markdown("**Add new store**")
            _ma1, _ma2 = st.columns(2)
            _new_plat = _ma1.selectbox(
                "Platform",
                ["shopee", "lazada", "tiktok_shop", "own_site"],
                format_func=lambda x: {
                    "shopee": "Shopee", "lazada": "Lazada",
                    "tiktok_shop": "TikTok Shop", "own_site": "Own Website",
                }.get(x, x),
            )
            _new_niche = _ma2.text_input("Niche", placeholder="e.g. capsule, cream")
            _new_name = st.text_input("Store name *",
                                      placeholder="e.g. GlowLab PH Official")
            _new_url = st.text_input("Store URL",
                                     placeholder="https://shopee.ph/glowlab")
            _new_notes = st.text_area("Notes", height=60,
                                      placeholder="Tagline, target market, anything to remember")
            _msa1, _msa2 = st.columns([3, 1])
            _ms_save = _msa1.form_submit_button("Save store", type="primary", width="stretch")
            _ms_cancel = _msa2.form_submit_button("Cancel", width="stretch")
        if _ms_cancel:
            st.session_state.pop("_ms_show_add", None)
            st.rerun()
        if _ms_save:
            if not _new_name.strip():
                st.error("Store name is required.")
            else:
                _new_id = db.insert_seller_store({
                    "platform": _new_plat, "store_name": _new_name,
                    "store_url": _new_url, "niche": _new_niche,
                    "notes": _new_notes,
                })
                st.session_state.pop("_ms_show_add", None)
                st.toast(f"✓ Store #{_new_id} added", icon=None)
                st.rerun()

    st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)

    if not _stores:
        st.info(
            "No stores tracked yet. Click **➕ Add store** to start logging your own "
            "Shopee / Lazada / TikTok Shop performance."
        )
    else:
        for _store in _stores:
            _snaps = db.list_seller_snapshots(_store["id"], limit=12)
            with st.expander(
                f"🏪 {_store['store_name']} · {_store['platform'].title()}"
                + (f" · {_store['niche']}" if _store.get("niche") else "")
                + (f" · {len(_snaps)} snapshots" if _snaps else ""),
                expanded=False,
            ):
                if _store.get("store_url"):
                    st.link_button("Open store →", _store["store_url"], width="stretch")

                # Snapshot history mini-chart
                if len(_snaps) >= 2:
                    import pandas as _pd_ms
                    _df_ms = _pd_ms.DataFrame([
                        {"date": s["snapshot_date"],
                         "units_sold": s.get("units_sold") or 0,
                         "revenue_php": s.get("revenue_php") or 0}
                        for s in reversed(_snaps)
                    ])
                    _df_ms = _df_ms.set_index("date")
                    st.line_chart(_df_ms, height=180)

                # Latest snapshot summary
                if _snaps:
                    _latest = _snaps[0]
                    _m1, _m2, _m3, _m4 = st.columns(4)
                    _m1.metric("Units sold", f"{int(_latest.get('units_sold') or 0):,}")
                    _m2.metric("Revenue", f"₱{float(_latest.get('revenue_php') or 0):,.0f}")
                    _m3.metric("Rating", f"{float(_latest.get('rating') or 0):.1f} ★"
                               if _latest.get("rating") else "—")
                    _m4.metric("Followers", f"{int(_latest.get('follower_count') or 0):,}")
                    st.caption(f"Latest snapshot: {_latest.get('snapshot_date')}")

                # Add snapshot form
                with st.form(f"ms_snap_{_store['id']}"):
                    st.markdown("**Add snapshot for today**")
                    _sc1, _sc2, _sc3, _sc4 = st.columns(4)
                    _sn_units = _sc1.number_input("Units sold", min_value=0, step=1,
                                                   key=f"ms_units_{_store['id']}")
                    _sn_rev = _sc2.number_input("Revenue (PHP)", min_value=0.0, step=100.0,
                                                 key=f"ms_rev_{_store['id']}")
                    _sn_rating = _sc3.number_input("Rating", min_value=0.0, max_value=5.0,
                                                    step=0.1, key=f"ms_rate_{_store['id']}")
                    _sn_reviews = _sc4.number_input("Review count", min_value=0, step=1,
                                                     key=f"ms_rev_n_{_store['id']}")
                    _sn_followers = st.number_input(
                        "Followers", min_value=0, step=10,
                        key=f"ms_followers_{_store['id']}",
                    )
                    _sn_notes = st.text_input(
                        "Snapshot notes", placeholder="optional context",
                        key=f"ms_notes_{_store['id']}",
                    )
                    _sn_a, _sn_b = st.columns([3, 1])
                    _sn_save = _sn_a.form_submit_button(
                        "Save snapshot", type="primary", width="stretch",
                    )
                    _sn_del = _sn_b.form_submit_button("Delete store", width="stretch")
                if _sn_save:
                    db.add_seller_snapshot(_store["id"], {
                        "units_sold": _sn_units, "revenue_php": _sn_rev,
                        "rating": _sn_rating, "review_count": _sn_reviews,
                        "follower_count": _sn_followers, "notes": _sn_notes,
                    })
                    st.toast("✓ Snapshot saved", icon=None)
                    st.rerun()
                if _sn_del:
                    db.delete_seller_store(_store["id"])
                    st.toast(f"✕ Deleted store '{_store['store_name']}'", icon=None)
                    st.rerun()


if current_page == "testing":
    _testing_products = db.list_testing_products()
    _t_total = len(_testing_products)

    # ---- Header: title + total + caption inline ----
    _th_l, _th_r = st.columns([3, 1])
    with _th_l:
        st.markdown(
            f"<div style='margin-bottom:4px'>"
            f"<span style='font-size:1.15rem;font-weight:600;color:var(--pra-text)'>Product testing lineup</span>"
            f"<span style='color:var(--pra-text-muted);margin-left:14px;font-size:0.85rem;"
            f"font-variant-numeric:tabular-nums'>{_t_total} product{'s' if _t_total != 1 else ''}</span>"
            f"</div>"
            f"<div style='color:var(--pra-text-muted);font-size:0.8rem'>"
            f"Track products to validate. Click pipeline cards or table rows to edit. "
            f"Export as PDF for supplier briefs."
            f"</div>",
            unsafe_allow_html=True,
        )
    st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)

    # ---- ROI summary panel (Phase 9.2) ----
    if _testing_products:
        _with_roi = [p for p in _testing_products if (p.get("revenue_php") or 0) or (p.get("ad_spend_php") or 0)]
        if _with_roi:
            _tot_rev = sum(float(p.get("revenue_php") or 0) for p in _with_roi)
            _tot_spend = sum(float(p.get("ad_spend_php") or 0) for p in _with_roi)
            _tot_cogs = sum(float(p.get("cogs_php") or 0) for p in _with_roi)
            _tot_net = _tot_rev - _tot_spend - _tot_cogs
            _avg_roas = (_tot_rev / _tot_spend) if _tot_spend else 0
            _n_winners = sum(1 for p in _with_roi if (p.get("outcome") or "") == "winner")
            _n_losers = sum(1 for p in _with_roi if (p.get("outcome") or "") == "loser")
            _win_rate = (_n_winners / len(_with_roi) * 100) if _with_roi else 0

            _roi_col = "var(--pra-success)" if _tot_net > 1.0 else (
                "var(--pra-danger)" if _tot_net < -1.0 else "var(--pra-warning)"
            )
            # Cleaner panel: flat panel bg (no gradient), tighter grid, responsive cell min-width
            _cell_style = (
                "min-width:0;padding:2px 0"
            )
            _label_style = (
                "color:var(--pra-text-muted);font-size:0.68rem;"
                "text-transform:uppercase;letter-spacing:0.1em;"
                "margin-bottom:4px;font-weight:500"
            )
            _value_style = (
                "font-size:1.05rem;font-weight:600;color:var(--pra-text);"
                "font-variant-numeric:tabular-nums;letter-spacing:-0.005em;"
                "overflow:hidden;text-overflow:ellipsis;white-space:nowrap"
            )
            st.markdown(
                f"<div style='background:var(--pra-panel);border:1px solid var(--pra-border);"
                f"border-left:2px solid var(--pra-accent);border-radius:6px;padding:14px 18px;"
                f"margin-bottom:18px'>"
                f"<div style='color:var(--pra-accent);font-size:0.66rem;letter-spacing:0.18em;"
                f"font-weight:700;text-transform:uppercase;margin-bottom:10px'>"
                f"ROI Summary · {len(_with_roi)} launched product{'s' if len(_with_roi) != 1 else ''}</div>"
                f"<div style='display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px'>"
                f"<div style='{_cell_style}'><div style='{_label_style}'>Revenue</div>"
                f"<div style='{_value_style}'>₱{_tot_rev:,.0f}</div></div>"
                f"<div style='{_cell_style}'><div style='{_label_style}'>Ad spend</div>"
                f"<div style='{_value_style}'>₱{_tot_spend:,.0f}</div></div>"
                f"<div style='{_cell_style}'><div style='{_label_style}'>Net profit</div>"
                f"<div style='{_value_style};color:{_roi_col}'>₱{_tot_net:,.0f}</div></div>"
                f"<div style='{_cell_style}'><div style='{_label_style}'>Avg ROAS</div>"
                f"<div style='{_value_style}'>{_avg_roas:.2f}x</div></div>"
                f"<div style='{_cell_style}'><div style='{_label_style}'>Win rate</div>"
                f"<div style='{_value_style}'>{_win_rate:.0f}% "
                f"<span style='font-size:0.74rem;color:var(--pra-text-muted);font-weight:400'>"
                f"({_n_winners}W·{_n_losers}L)</span></div></div>"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ---- Pipeline breakdown (kanban-style columns) ----
    if _testing_products:
        import html as _h_pl
        _STATUS_PIPELINE = [
            ("queued",   "Queued",   "#8FC2B4"),
            ("testing",  "Testing",  "#E6CC73"),
            ("passed",   "Passed",   "#7CC4A0"),
            ("launched", "Launched", "#D4AF37"),
            ("failed",   "Failed",   "#8B6B6B"),
        ]
        from collections import defaultdict as _dd_pl
        _by_status = _dd_pl(list)
        for p in _testing_products:
            _by_status[p.get("status", "queued")].append(p)

        _pcols = st.columns(len(_STATUS_PIPELINE))
        for _i, (_key, _label, _color) in enumerate(_STATUS_PIPELINE):
            with _pcols[_i]:
                _items = _by_status.get(_key, [])
                _cards_html = []
                for p in _items[:8]:  # cap to 8 per column for visual balance
                    _date = p.get("date_added", "")
                    _brand = (p.get("brand_name") or "").strip()
                    _niche = (p.get("niche") or "").strip()
                    _hunted = (p.get("hunted_by") or "").strip()
                    _meta_bits = []
                    if _niche:
                        _meta_bits.append(_niche.replace("_", " "))
                    if _date:
                        _meta_bits.append(_date)
                    _meta = " · ".join(_meta_bits) or "—"
                    _from_html = (
                        f"<div class='pra-pipeline-card-meta'>"
                        f"<span style='color:var(--pra-accent)'>from</span> {_h_pl.escape(_hunted)}"
                        f"</div>"
                    ) if _hunted else ""
                    _cards_html.append(
                        f"<div class='pra-pipeline-card'>"
                        f"<div class='pra-pipeline-card-name'>{_h_pl.escape(p.get('product_name','—'))}</div>"
                        f"<div class='pra-pipeline-card-meta'>{_h_pl.escape(_meta)}</div>"
                        + (f"<div class='pra-pipeline-card-brand'>{_h_pl.escape(_brand)}</div>" if _brand else "")
                        + _from_html
                        + "</div>"
                    )
                if len(_items) > 8:
                    _cards_html.append(
                        f"<div class='pra-pipeline-empty'>+ {len(_items) - 8} more</div>"
                    )
                if not _items:
                    _cards_html.append(
                        f"<div class='pra-pipeline-empty'>None yet</div>"
                    )

                st.markdown(
                    f"<div class='pra-pipeline-col'>"
                    f"<div class='pra-pipeline-col-header'>"
                    f"<span class='pra-pipeline-status' style='color:{_color}'>● {_label}</span>"
                    f"<span class='pra-pipeline-count'>{len(_items)}</span>"
                    f"</div>"
                    + "".join(_cards_html)
                    + "</div>",
                    unsafe_allow_html=True,
                )

        st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)

    # ---- Add new product form (minimal — just essentials) ----
    with st.expander("➕  Add a product to lineup", expanded=(_t_total == 0)):
        st.caption(
            "Quick-add the basics now. Click the row later to fill in marketing details "
            "(pain point, benefits, ingredients, target market) when ready."
        )
        with st.form("add_testing_form", clear_on_submit=True):
            f_name = st.text_input(
                "Product name *",
                placeholder="Glow Lean Fit Slimming Capsule",
            )
            row1_a, row1_b = st.columns(2)
            f_brand = row1_a.text_input(
                "Brand name",
                placeholder="GorgeousGlow Philippines",
            )
            _niche_opts = [""] + list(load_config().get("niches", {}).keys())
            f_niche = row1_b.selectbox(
                "Niche",
                _niche_opts,
                index=0,
                format_func=lambda x: "— pick a niche —" if x == "" else x.replace("_", " "),
            )
            row2_a, row2_b, row2_c = st.columns([1, 1, 1])
            from datetime import date as _date_cls
            f_date = row2_a.text_input("Date", value=_date_cls.today().isoformat())
            f_status = row2_b.selectbox("Status", db.TESTING_STATUSES, index=0)
            f_hunted_by = row2_c.text_input(
                "From (hunted by)",
                placeholder="Your name / VA / source",
                help="Who found this product? Useful if multiple people use this tool.",
            )
            f_notes = st.text_area(
                "Why are you testing this?",
                placeholder="One-liner: what caught your eye about this product?",
                height=70,
            )
            submitted = st.form_submit_button("Add to lineup", type="primary", width="stretch")
            if submitted:
                if not f_name.strip():
                    st.error("Product name is required.")
                else:
                    new_id = db.insert_testing_product({
                        "date_added": f_date,
                        "status": f_status,
                        "product_name": f_name.strip(),
                        "brand_name": f_brand.strip(),
                        "niche": f_niche,
                        "hunted_by": f_hunted_by.strip(),
                        "notes": f_notes.strip(),
                    })
                    st.success(
                        f"Added #{new_id} — {f_name.strip()}. "
                        f"Click the row below to fill in marketing details."
                    )
                    st.rerun()

    # ---- Filters + export ----
    fcol_a, fcol_b, fcol_c = st.columns([2, 2, 2])
    _status_filter = fcol_a.selectbox(
        "Filter by status", ["all"] + list(db.TESTING_STATUSES), index=0, key="testing_status_filter"
    )
    if _status_filter != "all":
        _testing_products = [x for x in _testing_products if x.get("status") == _status_filter]

    # PDF export
    import testing_pdf
    try:
        _pdf_bytes = testing_pdf.render_pdf(_testing_products)
        fcol_b.download_button(
            "Download as PDF",
            data=_pdf_bytes,
            file_name=f"product_testing_lineup_{datetime.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            type="primary",
            width="stretch",
            disabled=not _testing_products,
        )
    except Exception as e:
        fcol_b.caption(f"PDF render error: {e}")

    # CSV fallback export
    if _testing_products:
        import io as _io
        import csv as _csv
        _csv_buf = _io.StringIO()
        _writer = _csv.DictWriter(_csv_buf, fieldnames=list(_testing_products[0].keys()))
        _writer.writeheader()
        for row in _testing_products:
            _writer.writerow(row)
        fcol_c.download_button(
            "Download as CSV",
            data=_csv_buf.getvalue().encode("utf-8-sig"),
            file_name=f"product_testing_lineup_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            width="stretch",
        )

    # ---- Lineup table ----
    if not _testing_products:
        st.info(
            "No products in the testing lineup yet. "
            "Add one above to start tracking. "
            "Tip: brand winners from your Research tab make great testing candidates."
        )
    else:
        # Build a compact table view
        _df_test = pd.DataFrame([
            {
                "id": p.get("id"),
                "date": p.get("date_added"),
                "product": p.get("product_name"),
                "brand": p.get("brand_name") or "—",
                "niche": p.get("niche") or "—",
                "from": p.get("hunted_by") or "—",
                "status": p.get("status"),
                "pain_point": (p.get("pain_point") or "")[:80],
                "target": (
                    f"{p.get('target_age','')} · {p.get('target_gender','')}"
                    if (p.get('target_age') or p.get('target_gender'))
                    else "—"
                ),
            } for p in _testing_products
        ])
        _STATUS_BADGE = {
            "queued": "○ queued", "testing": "● testing", "passed": "● passed",
            "failed": "○ failed", "launched": "● launched",
        }
        _df_test["status"] = _df_test["status"].apply(lambda s: _STATUS_BADGE.get(s, s))

        col_cfg = {
            "id": st.column_config.NumberColumn("#", width="small", format="%d"),
            "date": st.column_config.TextColumn("Date", width="small"),
            "product": st.column_config.TextColumn("Product", width="medium"),
            "brand": st.column_config.TextColumn("Brand", width="medium"),
            "niche": st.column_config.TextColumn("Niche", width="small"),
            "from": st.column_config.TextColumn("From", width="small",
                       help="Who found / hunted this product"),
            "status": st.column_config.TextColumn("Status", width="small"),
            "pain_point": st.column_config.TextColumn("Pain point", width="large"),
            "target": st.column_config.TextColumn("Target", width="medium"),
        }
        _test_event = st.dataframe(
            _df_test, column_config=col_cfg, width="stretch", hide_index=True, height=480,
            on_select="rerun", selection_mode="single-row", key="testing_table",
        )
        st.caption("Click any row to edit, change status, or delete. PDF export uses the filtered view above.")

        if (_test_event and getattr(_test_event, "selection", None)
                and _test_event.selection.get("rows")):
            _sel = _test_event.selection["rows"][0]
            _sel_id = int(_df_test.iloc[_sel]["id"])
            _edit_testing_dialog(_sel_id)
